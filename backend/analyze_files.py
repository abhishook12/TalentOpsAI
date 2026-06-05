"""
Comprehensive file analysis for TalentOps AI ETL planning.
NO imports, NO database writes. Read-only analysis only.
"""
import os
import json
import re
from collections import Counter, defaultdict
from openpyxl import load_workbook

UPLOAD_DIR = "C:/TalentOpsAI/backend/uploads"

EMAIL_RE = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
PHONE_RE = re.compile(r'[\d]{7,}')
BAD_PHONE = {"0", "-", "--", "000-000-0000", "0000000000", "000000000", "n/a", "na", "none"}

def analyze_file(filepath):
    """Analyze a single xlsx file thoroughly."""
    result = {
        "filename": os.path.basename(filepath),
        "file_size_bytes": os.path.getsize(filepath),
        "sheets": [],
        "errors": [],
    }
    
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        result["errors"].append(f"Cannot open file: {e}")
        return result
    
    result["sheet_names"] = wb.sheetnames
    result["sheet_count"] = len(wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "sheet_name": sheet_name,
            "rows": [],
            "total_rows": 0,
            "blank_rows": 0,
            "has_headers": False,
            "detected_format": "unknown",
            "headers": [],
            "sample_rows": [],
            "unique_emails": set(),
            "unique_phones": set(),
            "unique_companies": set(),
            "unique_states": set(),
            "unique_names": set(),
            "rows_with_email": 0,
            "rows_without_email": 0,
            "rows_with_phone": 0,
            "rows_with_company": 0,
            "rows_with_state": 0,
            "rows_with_name": 0,
            "bad_phones": 0,
            "multi_email_rows": 0,
            "multi_phone_rows": 0,
            "possible_vertical_format": False,
            "column_value_samples": defaultdict(list),
            "email_columns": [],
            "phone_columns": [],
            "name_columns": [],
            "company_columns": [],
            "state_columns": [],
            "duplicate_emails_within_file": [],
        }
        
        all_rows = []
        for row_values in ws.iter_rows(values_only=True):
            all_rows.append([str(v).strip() if v is not None else "" for v in row_values])
        
        if not all_rows:
            sheet_info["errors"] = ["Empty sheet"]
            result["sheets"].append(sheet_info)
            continue
        
        # Detect headers
        first_row = all_rows[0]
        # Heuristic: if first row has no email-like values and looks like labels
        has_email_in_first = any(EMAIL_RE.match(str(v).strip()) for v in first_row if v)
        has_short_text = all(len(str(v)) < 50 for v in first_row if v)
        non_empty_first = [v for v in first_row if str(v).strip()]
        
        if not has_email_in_first and has_short_text and len(non_empty_first) >= 2:
            sheet_info["has_headers"] = True
            sheet_info["headers"] = first_row
            data_rows = all_rows[1:]
        else:
            sheet_info["has_headers"] = False
            sheet_info["headers"] = [f"Column_{i}" for i in range(len(first_row))]
            data_rows = all_rows
        
        sheet_info["total_rows"] = len(data_rows)
        
        # Sample first 5 and last 2 rows
        for i, row in enumerate(data_rows[:5]):
            sheet_info["sample_rows"].append({"row_index": i+1, "values": row})
        if len(data_rows) > 7:
            for i, row in enumerate(data_rows[-2:]):
                sheet_info["sample_rows"].append({"row_index": len(data_rows)-1+i, "values": row})
        
        # Detect which columns contain what type of data
        headers = sheet_info["headers"]
        email_col_indices = []
        phone_col_indices = []
        name_col_indices = []
        company_col_indices = []
        state_col_indices = []
        field_type_col_idx = None
        field_value_col_idx = None
        
        # First pass: by header names
        for idx, h in enumerate(headers):
            h_lower = h.lower().strip()
            if any(kw in h_lower for kw in ["email", "mail", "e-mail"]):
                email_col_indices.append(idx)
            if any(kw in h_lower for kw in ["phone", "mobile", "cell", "tel"]):
                phone_col_indices.append(idx)
            if any(kw in h_lower for kw in ["name", "contact", "recruiter", "person"]):
                name_col_indices.append(idx)
            if any(kw in h_lower for kw in ["company", "organization", "employer", "firm", "vendor"]):
                company_col_indices.append(idx)
            if any(kw in h_lower for kw in ["state", "province", "region"]):
                state_col_indices.append(idx)
            if h_lower in ["field type", "field_type", "type", "field"]:
                field_type_col_idx = idx
            if h_lower in ["field value", "field_value", "value"]:
                field_value_col_idx = idx
        
        # Second pass: by data content (for columns not yet identified)
        for idx in range(len(headers)):
            if idx in email_col_indices or idx in phone_col_indices:
                continue
            # Sample up to 100 rows
            sample_vals = [data_rows[i][idx] for i in range(min(100, len(data_rows))) if idx < len(data_rows[i]) and data_rows[i][idx]]
            email_matches = sum(1 for v in sample_vals if EMAIL_RE.match(str(v).strip()))
            phone_matches = sum(1 for v in sample_vals if PHONE_RE.search(str(v).replace("-","").replace(" ","").replace("(","").replace(")","")))
            
            if len(sample_vals) > 0:
                if email_matches / max(len(sample_vals), 1) > 0.5 and idx not in email_col_indices:
                    email_col_indices.append(idx)
                elif phone_matches / max(len(sample_vals), 1) > 0.5 and idx not in phone_col_indices:
                    phone_col_indices.append(idx)
        
        sheet_info["email_columns"] = [headers[i] for i in email_col_indices if i < len(headers)]
        sheet_info["phone_columns"] = [headers[i] for i in phone_col_indices if i < len(headers)]
        sheet_info["name_columns"] = [headers[i] for i in name_col_indices if i < len(headers)]
        sheet_info["company_columns"] = [headers[i] for i in company_col_indices if i < len(headers)]
        sheet_info["state_columns"] = [headers[i] for i in state_col_indices if i < len(headers)]
        
        # Detect vertical multi-value format
        if field_type_col_idx is not None and field_value_col_idx is not None:
            sheet_info["possible_vertical_format"] = True
            # Count unique field types
            field_types = Counter()
            for row in data_rows[:500]:
                if field_type_col_idx < len(row) and row[field_type_col_idx]:
                    field_types[row[field_type_col_idx].strip().lower()] += 1
            sheet_info["vertical_field_types"] = dict(field_types.most_common(20))
        
        # Also detect vertical format by checking if name+company repeat with different data
        if not sheet_info["possible_vertical_format"] and len(headers) <= 5:
            name_company_counts = Counter()
            for row in data_rows[:1000]:
                key_parts = []
                for ni in name_col_indices[:1]:
                    if ni < len(row): key_parts.append(row[ni])
                for ci in company_col_indices[:1]:
                    if ci < len(row): key_parts.append(row[ci])
                if key_parts:
                    name_company_counts[tuple(key_parts)] += 1
            repeated = sum(1 for k, v in name_company_counts.items() if v > 1)
            if repeated > len(name_company_counts) * 0.3:
                sheet_info["possible_vertical_format"] = True
                sheet_info["repeated_name_company_ratio"] = f"{repeated}/{len(name_company_counts)}"
        
        # Analyze each data row
        email_counter = Counter()
        
        for row in data_rows:
            is_blank = all(not str(v).strip() for v in row)
            if is_blank:
                sheet_info["blank_rows"] += 1
                continue
            
            # Count emails in this row
            row_emails = []
            for ei in email_col_indices:
                if ei < len(row) and row[ei] and EMAIL_RE.match(str(row[ei]).strip()):
                    row_emails.append(str(row[ei]).strip().lower())
                    sheet_info["unique_emails"].add(str(row[ei]).strip().lower())
            
            if row_emails:
                sheet_info["rows_with_email"] += 1
                for e in row_emails:
                    email_counter[e] += 1
            else:
                sheet_info["rows_without_email"] += 1
            
            if len(row_emails) > 1:
                sheet_info["multi_email_rows"] += 1
            
            # Count phones
            row_phones = []
            for pi in phone_col_indices:
                if pi < len(row) and row[pi]:
                    cleaned = str(row[pi]).replace("-","").replace(" ","").replace("(","").replace(")","").replace("+","").strip()
                    if cleaned and cleaned.lower() not in BAD_PHONE and PHONE_RE.search(cleaned):
                        row_phones.append(cleaned)
                        sheet_info["unique_phones"].add(cleaned)
                    elif cleaned and cleaned.lower() in BAD_PHONE:
                        sheet_info["bad_phones"] += 1
            
            if row_phones:
                sheet_info["rows_with_phone"] += 1
            if len(row_phones) > 1:
                sheet_info["multi_phone_rows"] += 1
            
            # Companies
            for ci in company_col_indices:
                if ci < len(row) and row[ci]:
                    sheet_info["unique_companies"].add(str(row[ci]).strip())
                    sheet_info["rows_with_company"] += 1
            
            # States
            for si in state_col_indices:
                if si < len(row) and row[si]:
                    sheet_info["unique_states"].add(str(row[si]).strip())
                    sheet_info["rows_with_state"] += 1
            
            # Names
            for ni in name_col_indices:
                if ni < len(row) and row[ni]:
                    sheet_info["unique_names"].add(str(row[ni]).strip())
                    sheet_info["rows_with_name"] += 1
        
        # Duplicate emails within file
        dup_emails = {e: c for e, c in email_counter.items() if c > 1}
        sheet_info["duplicate_emails_count"] = len(dup_emails)
        sheet_info["duplicate_email_examples"] = dict(list(dup_emails.items())[:10])
        
        # Detect format
        if sheet_info["possible_vertical_format"]:
            sheet_info["detected_format"] = "VERTICAL_MULTI_VALUE"
        elif len(email_col_indices) > 1 or len(phone_col_indices) > 1:
            sheet_info["detected_format"] = "WIDE_CONTACT"
        elif len(headers) <= 5 and sheet_info["has_headers"]:
            sheet_info["detected_format"] = "SIMPLE_ROW"
        else:
            sheet_info["detected_format"] = "STANDARD"
        
        # Convert sets to counts for JSON serialization
        sheet_info["unique_email_count"] = len(sheet_info["unique_emails"])
        sheet_info["unique_phone_count"] = len(sheet_info["unique_phones"])
        sheet_info["unique_company_count"] = len(sheet_info["unique_companies"])
        sheet_info["unique_state_count"] = len(sheet_info["unique_states"])
        sheet_info["unique_name_count"] = len(sheet_info["unique_names"])
        
        # Top companies
        company_counter = Counter()
        for row in data_rows:
            for ci in company_col_indices:
                if ci < len(row) and row[ci]:
                    company_counter[str(row[ci]).strip()] += 1
        sheet_info["top_companies"] = dict(company_counter.most_common(15))
        
        # Top states
        state_counter = Counter()
        for row in data_rows:
            for si in state_col_indices:
                if si < len(row) and row[si]:
                    state_counter[str(row[si]).strip()] += 1
        sheet_info["top_states"] = dict(state_counter.most_common(15))
        
        # Clean up non-serializable fields
        del sheet_info["unique_emails"]
        del sheet_info["unique_phones"]
        del sheet_info["unique_companies"]
        del sheet_info["unique_states"]
        del sheet_info["unique_names"]
        del sheet_info["rows"]
        del sheet_info["column_value_samples"]
        
        result["sheets"].append(sheet_info)
    
    wb.close()
    return result


def cross_file_analysis(file_results):
    """Analyze data overlap across all files."""
    all_emails = defaultdict(list)  # email -> list of filenames
    all_names = defaultdict(list)
    
    for fr in file_results:
        fname = fr["filename"]
        try:
            wb = load_workbook(os.path.join(UPLOAD_DIR, fname), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            
            if not rows:
                continue
            headers = [str(v).strip().lower() if v else "" for v in rows[0]]
            
            email_idx = None
            name_idx = None
            for i, h in enumerate(headers):
                if "email" in h or "mail" in h:
                    if email_idx is None: email_idx = i
                if "name" in h:
                    if name_idx is None: name_idx = i
            
            for row in rows[1:]:
                vals = [str(v).strip() if v else "" for v in row]
                if email_idx is not None and email_idx < len(vals) and vals[email_idx]:
                    e = vals[email_idx].lower().strip()
                    if EMAIL_RE.match(e):
                        all_emails[e].append(fname)
                if name_idx is not None and name_idx < len(vals) and vals[name_idx]:
                    all_names[vals[name_idx].strip()].append(fname)
        except Exception:
            continue
    
    # Emails appearing in multiple files
    cross_file_dupes = {e: files for e, files in all_emails.items() if len(set(files)) > 1}
    
    return {
        "total_unique_emails_across_all_files": len(all_emails),
        "total_unique_names_across_all_files": len(all_names),
        "emails_in_multiple_files": len(cross_file_dupes),
        "cross_file_duplicate_examples": {e: list(set(f)) for e, f in list(cross_file_dupes.items())[:10]},
    }


if __name__ == "__main__":
    files = sorted([f for f in os.listdir(UPLOAD_DIR) if f.endswith((".xlsx", ".csv"))])
    print(f"Found {len(files)} files in {UPLOAD_DIR}\n")
    
    file_results = []
    for f in files:
        path = os.path.join(UPLOAD_DIR, f)
        print(f"Analyzing: {f} ...")
        result = analyze_file(path)
        file_results.append(result)
    
    print("\nRunning cross-file analysis...")
    cross = cross_file_analysis(file_results)
    
    output = {
        "files": file_results,
        "cross_file_analysis": cross,
    }
    
    output_path = "C:/TalentOpsAI/backend/file_analysis_report.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, default=str)
    
    print(f"\nFull report saved to: {output_path}")
    
    # Print summary
    print("\n" + "="*80)
    print("FILE ANALYSIS SUMMARY")
    print("="*80)
    
    for fr in file_results:
        print(f"\n--- {fr['filename']} ---")
        print(f"  Size: {fr['file_size_bytes']:,} bytes")
        print(f"  Sheets: {fr.get('sheet_names', [])}")
        for s in fr["sheets"]:
            print(f"  Sheet: '{s['sheet_name']}'")
            print(f"    Has headers: {s['has_headers']}")
            print(f"    Headers: {s['headers']}")
            print(f"    Detected format: {s['detected_format']}")
            print(f"    Total rows: {s['total_rows']:,}")
            print(f"    Blank rows: {s['blank_rows']:,}")
            print(f"    Rows with email: {s['rows_with_email']:,}")
            print(f"    Rows WITHOUT email: {s['rows_without_email']:,}")
            print(f"    Rows with phone: {s['rows_with_phone']:,}")
            print(f"    Rows with company: {s['rows_with_company']:,}")
            print(f"    Rows with state: {s['rows_with_state']:,}")
            print(f"    Unique emails: {s['unique_email_count']:,}")
            print(f"    Unique phones: {s['unique_phone_count']:,}")
            print(f"    Unique companies: {s['unique_company_count']:,}")
            print(f"    Unique states: {s['unique_state_count']:,}")
            print(f"    Unique names: {s['unique_name_count']:,}")
            print(f"    Multi-email rows: {s['multi_email_rows']:,}")
            print(f"    Multi-phone rows: {s['multi_phone_rows']:,}")
            print(f"    Bad phones: {s['bad_phones']:,}")
            print(f"    Duplicate emails (within file): {s['duplicate_emails_count']:,}")
            if s.get("possible_vertical_format"):
                print(f"    *** VERTICAL FORMAT DETECTED ***")
            print(f"    Email columns: {s['email_columns']}")
            print(f"    Phone columns: {s['phone_columns']}")
            print(f"    Name columns: {s['name_columns']}")
            print(f"    Company columns: {s['company_columns']}")
            print(f"    State columns: {s['state_columns']}")
            print(f"    Top companies: {list(s.get('top_companies', {}).keys())[:5]}")
            print(f"    Top states: {list(s.get('top_states', {}).keys())[:5]}")
            if s["sample_rows"]:
                print(f"    Sample row 1: {s['sample_rows'][0]['values']}")
    
    print(f"\n{'='*80}")
    print("CROSS-FILE ANALYSIS")
    print(f"{'='*80}")
    print(f"  Total unique emails across all files: {cross['total_unique_emails_across_all_files']:,}")
    print(f"  Total unique names across all files: {cross['total_unique_names_across_all_files']:,}")
    print(f"  Emails found in multiple files: {cross['emails_in_multiple_files']:,}")
    if cross["cross_file_duplicate_examples"]:
        print(f"  Cross-file duplicate examples:")
        for e, flist in list(cross["cross_file_duplicate_examples"].items())[:5]:
            print(f"    {e} -> {flist}")
