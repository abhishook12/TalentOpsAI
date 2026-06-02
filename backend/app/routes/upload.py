from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, BackgroundTasks, Request, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any
import pandas as pd
import io, csv, uuid, re
from openpyxl import load_workbook
from app.database import get_db
from app.models.models import Candidate, Recruiter, UploadJob, ActionLog
from app.utils.column_mapper import detect_columns
from app.schemas.upload import AnalyzeResponse

router = APIRouter()

# ─── Helpers ───────────────────────────────────────────────────────────────────
def clean_email(email): return str(email).lower().strip()
def clean_phone(phone): return str(phone).replace("-","").replace(" ","").replace("(","").replace(")","").strip()
def clean_name(name): return str(name).strip().title()

def read_file(contents, filename):
    if filename.endswith(".csv"):
        text = contents.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        return [row for row in reader]
    else:
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        return [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]

def stream_file_rows(contents: bytes, filename: str):
    if filename.lower().endswith(".csv"):
        text = contents.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        return headers, reader

    if filename.lower().endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows, [])]

        def row_iter():
            for values in rows:
                yield {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers)) if headers[idx]}

        return headers, row_iter()

    # Legacy .xls fallback. This is slower, but it keeps compatibility.
    df = pd.read_excel(io.BytesIO(contents), dtype=str, keep_default_na=False)
    headers = list(df.columns)
    return headers, ({col: row[col] for col in headers} for _, row in df.iterrows())

def read_file_headers(contents: bytes, filename: str):
    if filename.lower().endswith(".csv"):
        text = contents.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return next(reader, [])

    if filename.lower().endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(values_only=True), [])]
        wb.close()
        return headers

    df = pd.read_excel(io.BytesIO(contents), dtype=str, keep_default_na=False, nrows=1)
    return list(df.columns)

# ─── Original Legacy Endpoints (keep working) ─────────────────────────────────

@router.post("/candidates")
async def upload_candidates(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    contents = await file.read()
    rows = read_file(contents, file.filename)
    if not rows or "candidate_name" not in rows[0] or "email" not in rows[0]:
        raise HTTPException(status_code=400, detail="Missing required columns: candidate_name, email")
    inserted = duplicates = 0
    for row in rows:
        email = clean_email(row.get("email",""))
        if not email: continue
        if db.query(Candidate).filter(Candidate.email == email).first():
            duplicates += 1; continue
        candidate = Candidate(
            candidate_name=clean_name(row.get("candidate_name","")),
            email=email,
            phone=clean_phone(row.get("phone","")) or None,
            visa_status=str(row["visa_status"]).strip() if row.get("visa_status") else None,
            skills=str(row["skills"]) if row.get("skills") else None,
            experience_years=float(row["experience_years"]) if row.get("experience_years") else None,
            location=str(row["location"]).strip() if row.get("location") else None,
            rate_per_hour=float(row["rate_per_hour"]) if row.get("rate_per_hour") else None,
            availability=str(row["availability"]).strip() if row.get("availability") else None,
        )
        db.add(candidate); inserted += 1
    db.commit()
    return {"message": "Upload complete", "inserted": inserted, "duplicates_skipped": duplicates, "total_rows": len(rows)}

@router.post("/recruiters")
async def upload_recruiters(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    contents = await file.read()
    rows = read_file(contents, file.filename)
    if not rows or "recruiter_name" not in rows[0] or "email" not in rows[0]:
        raise HTTPException(status_code=400, detail="Missing required columns: recruiter_name, email")
    inserted = duplicates = 0
    for row in rows:
        email = clean_email(row.get("email",""))
        if not email: continue
        if db.query(Recruiter).filter(Recruiter.email == email).first():
            duplicates += 1; continue
        recruiter = Recruiter(
            recruiter_name=clean_name(row.get("recruiter_name","")),
            email=email,
            phone=clean_phone(row.get("phone","")) or None,
            email2=str(row["email2"]).lower().strip() if row.get("email2") else None,
            phone2=clean_phone(row.get("phone2","")) or None,
            linkedin=str(row["linkedin"]).strip() if row.get("linkedin") else None,
            specialization=str(row["specialization"]).strip() if row.get("specialization") else None,
            notes=str(row["notes"]).strip() if row.get("notes") else None,
            is_active=True,
        )
        db.add(recruiter); inserted += 1
    db.commit()
    return {"message": "Upload complete", "inserted": inserted, "duplicates_skipped": duplicates, "total_rows": len(rows)}


# ─── NEW: Smart Analyze Endpoint ──────────────────────────────────────────────

@router.post("/analyze", response_model=AnalyzeResponse)
async def analyze_file(file: UploadFile = File(...)):
    """Parse the uploaded CSV/Excel, automatically detect column mapping,
    run basic validation, and return a preview with statistics."""
    if not file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Unsupported file type. Use CSV or Excel.")
    try:
        contents = await file.read()
        headers, rows = stream_file_rows(contents, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    # Detect column mapping using heuristic matcher
    column_map = detect_columns(list(headers))

    total_rows = 0
    email_col = column_map.get('email')
    duplicates = missing = invalid_emails = invalid_phones = corrupted = 0
    empty_cols = set(headers)
    non_empty_cols = set()
    seen_emails = set()
    email_re = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
    phone_re = re.compile(r'^(?:\+?1[\s.\-]?)?\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4}$')
    phone_cols = [c for k, c in column_map.items() if k.startswith('phone') and c]
    preview: List[Dict[str, Any]] = []

    for row in rows:
        total_rows += 1
        cleaned_row = {col: row.get(col) for col in headers}
        if total_rows <= 10:
            preview.append(cleaned_row)

        row_has_value = False
        for col, value in cleaned_row.items():
            if str(value).strip():
                row_has_value = True
                non_empty_cols.add(col)

        if not row_has_value:
            corrupted += 1

        if email_col:
            raw_email = str(cleaned_row.get(email_col, '')).strip()
            if not raw_email:
                missing += 1
            else:
                if raw_email in seen_emails:
                    duplicates += 1
                else:
                    seen_emails.add(raw_email)
                if not email_re.match(raw_email):
                    invalid_emails += 1

        for pc in phone_cols:
            raw_phone = str(cleaned_row.get(pc, '')).strip()
            if raw_phone and not phone_re.match(raw_phone):
                invalid_phones += 1

    empty_cols = [c for c in headers if c not in non_empty_cols]

    # Prepare preview (first 10 rows) with original headers
    analysis_id = str(uuid.uuid4())
    return AnalyzeResponse(
        analysis_id=analysis_id,
        total_rows=total_rows,
        duplicates=duplicates,
        missing_fields=missing,
        invalid_emails=invalid_emails,
        invalid_phones=invalid_phones,
        empty_columns=empty_cols,
        corrupted_rows=corrupted,
        column_map=column_map,
        original_headers=list(df.columns),
        preview=preview,
    )


# ─── NEW: Smart Import Async Endpoint ───────────────────────────────────────────
import os
import json
from app.services.etl_worker import process_smart_import

@router.post("/smart-import-async")
async def smart_import_async(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    mapping: str = Form(None),
    db: Session = Depends(get_db),
):
    """Accept a file + column mapping and enqueue an async background import job."""
    if not file.filename.lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Unsupported file type. Use CSV or Excel.")
    
    # Save file to disk securely
    job_id = str(uuid.uuid4())
    ext = file.filename.split('.')[-1]
    safe_filename = f"{job_id}.{ext}"
    os.makedirs("uploads", exist_ok=True)
    filepath = os.path.join("uploads", safe_filename)
    
    contents = await file.read()
    with open(filepath, "wb") as f:
        f.write(contents)

    try:
        headers = read_file_headers(contents, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    column_map = detect_columns(list(headers))
    if mapping:
        try:
            parsed_mapping = json.loads(mapping)
            if isinstance(parsed_mapping, dict):
                column_map = parsed_mapping
        except Exception:
            pass
    
    # Create UploadJob record
    new_job = UploadJob(
        job_id=job_id,
        filename=file.filename,
        status="queued",
        total_rows=0,
        processed_rows=0,
        inserted_rows=0,
        skipped_rows=0,
        error_count=0
    )
    db.add(new_job)

    # Log action
    user_email = request.headers.get("X-User-Email", "Anonymous")
    session_id = request.headers.get("X-Session-ID")
    ip_address = request.client.host if request.client else None
    action_log = ActionLog(
        user_email=user_email,
        session_id=session_id,
        action_type="UPLOAD_ETL",
        details=json.dumps({"filename": file.filename, "job_id": job_id, "rows": 0}),
        status="success",
        ip_address=ip_address
    )
    db.add(action_log)
    
    db.commit()

    # Launch background task
    background_tasks.add_task(process_smart_import, job_id, filepath, column_map)

    return {"message": "Job queued successfully", "job_id": job_id, "column_map": column_map}

@router.get("/jobs")
def get_jobs(db: Session = Depends(get_db)):
    jobs = db.query(UploadJob).order_by(UploadJob.started_at.desc()).limit(20).all()
    return jobs

@router.get("/jobs/{job_id}")
def get_job_status(job_id: str, db: Session = Depends(get_db)):
    job = db.query(UploadJob).filter(UploadJob.job_id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job_dict = {
        "job_id": job.job_id,
        "filename": job.filename,
        "status": job.status,
        "total_rows": job.total_rows,
        "processed_rows": job.processed_rows,
        "inserted_rows": job.inserted_rows,
        "skipped_rows": job.skipped_rows,
        "error_count": job.error_count,
        "started_at": job.started_at,
        "completed_at": job.completed_at,
        "errors": json.loads(job.errors) if job.errors else []
    }
    return job_dict

@router.post("/jobs/{job_id}/retry")
async def retry_job(job_id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    old_job = db.query(UploadJob).filter(UploadJob.job_id == job_id).first()
    if not old_job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    ext = old_job.filename.split('.')[-1]
    filepath = os.path.join("uploads", f"{job_id}.{ext}")
    if not os.path.exists(filepath):
        # Maybe it was the other extension, or doesn't exist
        fallback_ext = "xlsx" if ext == "csv" else "csv"
        filepath = os.path.join("uploads", f"{job_id}.{fallback_ext}")
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="Original uploaded file no longer available on server for retry")
            
    try:
        if filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath, dtype=str, keep_default_na=False)
        else:
            df = pd.read_excel(filepath, dtype=str, keep_default_na=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    column_map = detect_columns(list(df.columns))
    
    new_job_id = str(uuid.uuid4())
    
    # Copy file to new ID
    new_filepath = os.path.join("uploads", f"{new_job_id}.{filepath.split('.')[-1]}")
    import shutil
    shutil.copy(filepath, new_filepath)
    
    new_job = UploadJob(
        job_id=new_job_id,
        filename=old_job.filename,
        status="queued",
        total_rows=len(df),
        processed_rows=0,
        inserted_rows=0,
        skipped_rows=0,
        error_count=0
    )
    db.add(new_job)
    db.commit()

    background_tasks.add_task(process_smart_import, new_job_id, new_filepath, column_map)

    return {"message": "Job retried successfully", "new_job_id": new_job_id}
