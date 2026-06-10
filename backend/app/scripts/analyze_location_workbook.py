from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))

from app.database import SessionLocal
from app.models.models import Company, Recruiter
from app.utils.state_mapper import extract_state_detailed


SOURCE_FILE = Path(r"C:\Users\User\Desktop\for location by claude\1 the below all compny but location wise\Recruiter_Contacts_Master.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"


def norm_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def norm_key(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def norm_email(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip().lower()
    return value if "@" in value else None


def norm_phone(value: str | None) -> str | None:
    digits = re.sub(r"[^0-9]", "", value or "")
    return digits[-10:] if len(digits) >= 10 else None


def parse_state(*values: str | None):
    for value in values:
        if not value:
            continue
        state, reason = extract_state_detailed(value)
        if state:
            return state, reason
    return None, None


def parse_workbook():
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    merged_profiles: dict[str, dict] = {}
    row_count = 0
    sheet_counts = Counter()
    state_counts = Counter()

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [norm_text(str(cell)) if cell is not None else "" for cell in rows[0]]
        header_map = {header.lower(): index for index, header in enumerate(headers) if header}
        sheet_counts[sheet.title] += max(len(rows) - 1, 0)

        for row_index, row in enumerate(rows[1:], start=2):
            row_count += 1
            values = [norm_text(str(cell)) if cell is not None else "" for cell in row]
            name = values[header_map.get("name", 0)] if header_map.get("name", 0) < len(values) else ""
            email = ""
            for key in ("email", "email address", "emailid", "email id"):
                if key in header_map and header_map[key] < len(values):
                    email = values[header_map[key]]
                    break
            company = ""
            for key in ("company", "organization", "org"):
                if key in header_map and header_map[key] < len(values):
                    company = values[header_map[key]]
                    break
            if not company and len(values) >= 3:
                company = values[2]

            location = ""
            office_location = ""
            hq_location = ""
            country = ""
            phone = ""

            for key in ("location", "office location", "hq location", "country", "phone"):
                if key in header_map and header_map[key] < len(values):
                    value = values[header_map[key]]
                    if key == "location":
                        location = value
                    elif key == "office location":
                        office_location = value
                    elif key == "hq location":
                        hq_location = value
                    elif key == "country":
                        country = value
                    elif key == "phone":
                        phone = value

            if not location and len(values) >= 4:
                location = values[3]
            if not phone and len(values) >= 4 and "phone" in [h.lower() for h in headers]:
                phone = values[header_map.get("phone", 0)] if header_map.get("phone", 0) < len(values) else ""

            state, state_reason = parse_state(location, office_location, hq_location, country)
            if state:
                state_counts[state] += 1

            identity_email = norm_email(email)
            identity_phone = norm_phone(phone)
            identity_name_company = f"{norm_key(name)}::{norm_key(company)}" if name and company else None
            if identity_email:
                identity_key = f"email::{identity_email}"
            elif identity_phone:
                identity_key = f"phone::{identity_phone}"
            elif identity_name_company:
                identity_key = f"name_company::{identity_name_company}"
            else:
                identity_key = f"row::{sheet.title}::{row_index}"

            profile = merged_profiles.setdefault(
                identity_key,
                {
                    "identity_key": identity_key,
                    "name": name,
                    "email": identity_email,
                    "phone": identity_phone,
                    "company": company,
                    "locations": [],
                    "states": Counter(),
                    "sheet_names": set(),
                    "source_rows": [],
                },
            )

            if name and not profile["name"]:
                profile["name"] = name
            if identity_email and not profile["email"]:
                profile["email"] = identity_email
            if identity_phone and not profile["phone"]:
                profile["phone"] = identity_phone
            if company and not profile["company"]:
                profile["company"] = company

            if location:
                profile["locations"].append(location)
            if state:
                profile["states"][state] += 1
            profile["sheet_names"].add(sheet.title)
            profile["source_rows"].append(
                {
                    "sheet": sheet.title,
                    "row": row_index,
                    "name": name,
                    "email": identity_email,
                    "phone": identity_phone,
                    "company": company,
                    "location": location,
                    "office_location": office_location,
                    "hq_location": hq_location,
                    "country": country,
                    "state": state,
                    "state_reason": state_reason,
                }
            )

    profiles = list(merged_profiles.values())
    for profile in profiles:
        if profile["states"]:
            profile["dominant_state"], _ = profile["states"].most_common(1)[0]
        else:
            profile["dominant_state"] = None
        profile["sheet_names"] = sorted(profile["sheet_names"])
        profile["locations"] = list(dict.fromkeys(profile["locations"]))
        profile["source_row_count"] = len(profile["source_rows"])

    return profiles, row_count, sheet_counts, state_counts


def load_current_db():
    session = SessionLocal()
    try:
        company_rows = session.query(Company.company_id, Company.company_name).all()
        company_map = {company_id: company_name for company_id, company_name in company_rows}
        recruiters = session.query(
            Recruiter.recruiter_id,
            Recruiter.recruiter_name,
            Recruiter.email,
            Recruiter.phone,
            Recruiter.company_id,
            Recruiter.state,
        ).all()
        email_index = {}
        phone_index = {}
        name_company_index = defaultdict(list)
        for recruiter_id, recruiter_name, email, phone, company_id, state in recruiters:
            if email:
                email_index[email.strip().lower()] = {
                    "recruiter_id": recruiter_id,
                    "recruiter_name": recruiter_name,
                    "email": email,
                    "phone": phone,
                    "company_id": company_id,
                    "company_name": company_map.get(company_id, ""),
                    "state": state,
                }
            phone_key = norm_phone(phone)
            if phone_key:
                phone_index[phone_key] = {
                    "recruiter_id": recruiter_id,
                    "recruiter_name": recruiter_name,
                    "email": email,
                    "phone": phone,
                    "company_id": company_id,
                    "company_name": company_map.get(company_id, ""),
                    "state": state,
                }
            if recruiter_name and company_id:
                name_company_index[f"{norm_key(recruiter_name)}::{norm_key(company_map.get(company_id, ''))}"].append(
                    {
                        "recruiter_id": recruiter_id,
                        "recruiter_name": recruiter_name,
                        "email": email,
                        "phone": phone,
                        "company_id": company_id,
                        "company_name": company_map.get(company_id, ""),
                        "state": state,
                    }
                )
        return email_index, phone_index, name_company_index, company_map
    finally:
        session.close()


def cross_match_profiles(profiles, email_index, phone_index, name_company_index):
    stats = Counter()
    matched_examples = []
    unique_candidates = []
    existing_candidates = []
    evidence_state_counts = Counter()

    for profile in profiles:
        email = profile.get("email")
        phone = profile.get("phone")
        name_company_key = f"{norm_key(profile.get('name'))}::{norm_key(profile.get('company'))}" if profile.get("name") and profile.get("company") else None

        match_type = None
        matched_existing = None
        if email and email in email_index:
            match_type = "email"
            matched_existing = email_index[email]
        elif phone and phone in phone_index:
            match_type = "phone"
            matched_existing = phone_index[phone]
        elif name_company_key and name_company_key in name_company_index:
            match_type = "name_company"
            matched_existing = name_company_index[name_company_key][0]

        if profile.get("dominant_state"):
            evidence_state_counts[profile["dominant_state"]] += 1

        if matched_existing:
            stats[f"matched_{match_type}"] += 1
            existing_candidates.append(
                {
                    **profile,
                    "match_type": match_type,
                    "matched_recruiter_id": matched_existing["recruiter_id"],
                    "matched_recruiter_name": matched_existing["recruiter_name"],
                    "matched_company_name": matched_existing["company_name"],
                    "matched_state": matched_existing["state"],
                }
            )
            if len(matched_examples) < 20:
                matched_examples.append(
                    {
                        "name": profile.get("name"),
                        "email": email,
                        "company": profile.get("company"),
                        "state": profile.get("dominant_state"),
                        "match_type": match_type,
                        "matched_recruiter_id": matched_existing["recruiter_id"],
                        "matched_state": matched_existing["state"],
                    }
                )
            continue

        stats["unique"] += 1
        unique_candidates.append(profile)

    return stats, matched_examples, unique_candidates, existing_candidates, evidence_state_counts


def write_reports(profiles, stats, matched_examples, unique_candidates, existing_candidates, evidence_state_counts, sheet_counts, row_count):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = OUTPUT_DIR / "location_workbook_report.json"
    csv_path = OUTPUT_DIR / "location_workbook_unique_candidates.csv"
    matched_csv_path = OUTPUT_DIR / "location_workbook_matched_existing.csv"
    md_path = OUTPUT_DIR / "location_workbook_report.md"

    report = {
        "source_file": str(SOURCE_FILE),
        "source_rows": row_count,
        "merged_profiles": len(profiles),
        "stats": dict(stats),
        "matched_examples": matched_examples,
        "evidence_state_counts": dict(evidence_state_counts),
        "sheet_counts": dict(sheet_counts),
        "unique_candidates": len(unique_candidates),
        "existing_matches": len(existing_candidates),
    }
    json_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "identity_key",
                "name",
                "email",
                "phone",
                "company",
                "dominant_state",
                "locations",
                "sheet_names",
                "source_row_count",
            ],
        )
        writer.writeheader()
        for profile in unique_candidates:
            writer.writerow(
                {
                    "identity_key": profile.get("identity_key"),
                    "name": profile.get("name"),
                    "email": profile.get("email"),
                    "phone": profile.get("phone"),
                    "company": profile.get("company"),
                    "dominant_state": profile.get("dominant_state"),
                    "locations": " | ".join(profile.get("locations", [])),
                    "sheet_names": " | ".join(profile.get("sheet_names", [])),
                    "source_row_count": profile.get("source_row_count"),
                }
            )

    with matched_csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "identity_key",
                "name",
                "email",
                "phone",
                "company",
                "dominant_state",
                "match_type",
                "matched_recruiter_id",
                "matched_recruiter_name",
                "matched_company_name",
                "matched_state",
            ],
        )
        writer.writeheader()
        for profile in existing_candidates:
            writer.writerow(
                {
                    "identity_key": profile.get("identity_key"),
                    "name": profile.get("name"),
                    "email": profile.get("email"),
                    "phone": profile.get("phone"),
                    "company": profile.get("company"),
                    "dominant_state": profile.get("dominant_state"),
                    "match_type": profile.get("match_type"),
                    "matched_recruiter_id": profile.get("matched_recruiter_id"),
                    "matched_recruiter_name": profile.get("matched_recruiter_name"),
                    "matched_company_name": profile.get("matched_company_name"),
                    "matched_state": profile.get("matched_state"),
                }
            )

    lines = [
        "# Location Workbook Cross-Check Report",
        "",
        f"Source workbook: `{SOURCE_FILE}`",
        "",
        "## Parsed",
        f"- Source rows: `{row_count}`",
        f"- Merged profiles: `{len(profiles)}`",
        f"- Existing DB matches: `{len(existing_candidates)}`",
        f"- Unique candidates: `{len(unique_candidates)}`",
        "",
        "## Match breakdown",
    ]
    for key, value in stats.items():
        lines.append(f"- {key}: `{value}`")
    lines.extend([
        "",
        "## Strong state evidence by count",
    ])
    for state, count in evidence_state_counts.most_common(20):
        lines.append(f"- {state}: `{count}`")
    lines.extend([
        "",
        "## Files",
        f"- JSON report: `{json_path}`",
        f"- Unique candidates CSV: `{csv_path}`",
        f"- Existing matches CSV: `{matched_csv_path}`",
    ])
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    return report, json_path, csv_path, matched_csv_path, md_path


def main():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Workbook not found: {SOURCE_FILE}")

    profiles, row_count, sheet_counts, workbook_state_counts = parse_workbook()
    email_index, phone_index, name_company_index, _ = load_current_db()
    stats, matched_examples, unique_candidates, existing_candidates, evidence_state_counts = cross_match_profiles(
        profiles,
        email_index,
        phone_index,
        name_company_index,
    )
    report, json_path, csv_path, matched_csv_path, md_path = write_reports(
        profiles,
        stats,
        matched_examples,
        unique_candidates,
        existing_candidates,
        evidence_state_counts,
        sheet_counts,
        row_count,
    )
    print(json.dumps(report, indent=2, default=str))


if __name__ == "__main__":
    main()
