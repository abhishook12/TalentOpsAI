from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))

from app.database import SessionLocal
from app.models.models import Company, Recruiter
from app.utils.normalizer import normalize_text
from app.utils.state_mapper import extract_state_detailed


SOURCE_FILE = Path(r"C:\Users\User\Desktop\for location by claude\1 the below all compny but location wise\Recruiter_Contacts_Master.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
BATCH_ID = f"locwbuniq_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"


def norm_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def norm_key(value: str | None) -> str:
    return normalize_text(value or "")


def norm_email(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip().lower()
    return value if "@" in value else None


def norm_phone(value: str | None) -> str | None:
    digits = re.sub(r"[^0-9]", "", value or "")
    return digits[-10:] if len(digits) >= 10 else None


def parse_state(*values: str | None):
    for value in values:
        if not value:
            continue
        state, reason = extract_state_detailed(value)
        if state:
            return state, reason
    return None, None


def identity_key(name: str | None, email: str | None, phone: str | None, company: str | None):
    if email:
        return f"email::{email}"
    if phone:
        return f"phone::{phone}"
    if name and company:
        return f"name_company::{norm_key(name)}::{norm_key(company)}"
    return None


def parse_workbook():
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    profiles: dict[str, dict] = {}
    row_count = 0

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [norm_text(str(cell)) if cell is not None else "" for cell in rows[0]]
        header_map = {header.lower(): idx for idx, header in enumerate(headers) if header}

        for row_index, row in enumerate(rows[1:], start=2):
            row_count += 1
            values = [norm_text(str(cell)) if cell is not None else "" for cell in row]

            def cell(*names):
                for name in names:
                    idx = header_map.get(name)
                    if idx is not None and idx < len(values):
                        return values[idx]
                return ""

            name = cell("name")
            email = norm_email(cell("email", "email address", "email id"))
            phone = norm_phone(cell("phone", "mobile", "cell"))
            company = cell("company", "organization", "org") or (values[2] if len(values) > 2 else "")
            location = cell("location", "office location", "hq location")
            country = cell("country")
            office_location = cell("office location")
            hq_location = cell("hq location")

            state, state_reason = parse_state(location, office_location, hq_location, country)
            key = identity_key(name, email, phone, company)
            if not key:
                continue

            profile = profiles.setdefault(
                key,
                {
                    "identity_key": key,
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "company": company,
                    "locations": [],
                    "states": Counter(),
                    "sheet_names": set(),
                    "source_rows": [],
                },
            )

            if name and not profile["name"]:
                profile["name"] = name
            if email and not profile["email"]:
                profile["email"] = email
            if phone and not profile["phone"]:
                profile["phone"] = phone
            if company and not profile["company"]:
                profile["company"] = company
            if location:
                profile["locations"].append(location)
            if state:
                profile["states"][state] += 1
            profile["sheet_names"].add(sheet.title)
            profile["source_rows"].append(
                {
                    "sheet": sheet.title,
                    "row": row_index,
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "company": company,
                    "location": location,
                    "office_location": office_location,
                    "hq_location": hq_location,
                    "country": country,
                    "state": state,
                    "state_reason": state_reason,
                }
            )

    merged = list(profiles.values())
    for profile in merged:
        profile["sheet_names"] = sorted(profile["sheet_names"])
        profile["locations"] = list(dict.fromkeys([loc for loc in profile["locations"] if loc]))
        profile["source_row_count"] = len(profile["source_rows"])
        if profile["states"]:
            profile["dominant_state"], profile["dominant_state_count"] = profile["states"].most_common(1)[0]
        else:
            profile["dominant_state"], profile["dominant_state_count"] = None, 0
    return merged, row_count


def build_indexes(session):
    companies = session.query(Company).all()
    company_by_norm = {row.normalized_company_name or norm_key(row.company_name): row for row in companies if row.company_name}
    company_by_id = {row.company_id: row for row in companies}
    recruiters = session.query(Recruiter).all()

    email_index = {}
    phone_index = {}
    name_company_index = defaultdict(list)
    for recruiter in recruiters:
        if recruiter.email:
            email_index[recruiter.email.strip().lower()] = recruiter
        if recruiter.phone:
            phone_key = norm_phone(recruiter.phone)
            if phone_key:
                phone_index[phone_key] = recruiter
        company_name = company_by_id.get(recruiter.company_id).company_name if recruiter.company_id and company_by_id.get(recruiter.company_id) else ""
        if recruiter.recruiter_name and company_name:
            name_company_index[f"{norm_key(recruiter.recruiter_name)}::{norm_key(company_name)}"].append(recruiter)
    return company_by_norm, company_by_id, email_index, phone_index, name_company_index


def merge_metadata(existing_value, evidence):
    metadata = {}
    if existing_value:
        try:
            parsed = json.loads(existing_value) if isinstance(existing_value, str) else existing_value
            if isinstance(parsed, dict):
                metadata = dict(parsed)
        except Exception:
            metadata = {"raw_metadata": str(existing_value)}
    metadata["location_workbook_evidence"] = evidence
    return json.dumps(metadata, default=str)


def build_insert_email(profile: dict) -> str:
    email = profile.get("email")
    if email:
        return email
    phone = profile.get("phone")
    if phone:
        return f"no-email-{phone}@missing.local"
    token = hashlib.sha1((profile.get("identity_key") or "").encode("utf-8")).hexdigest()[:16]
    return f"no-email-{token}@missing.local"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch-size", type=int, default=500)
    parser.add_argument("--apply-existing-enrichment", action="store_true")
    args = parser.parse_args()

    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Workbook not found: {SOURCE_FILE}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    profiles, row_count = parse_workbook()

    session = SessionLocal()
    try:
        company_by_norm, company_by_id, email_index, phone_index, name_company_index = build_indexes(session)
        company_cache = dict(company_by_norm)

        existing_matches = []
        unique_candidates = []
        skipped = []
        ambiguous = []
        stats = Counter()

        for profile in profiles:
            if not profile.get("phone") and not profile.get("locations"):
                skipped.append(profile)
                stats["skipped_no_phone_or_location"] += 1
                continue

            email = profile.get("email")
            phone = profile.get("phone")
            key = f"{norm_key(profile.get('name'))}::{norm_key(profile.get('company'))}" if profile.get("name") and profile.get("company") else None
            match = None
            match_type = None
            if email and email in email_index:
                match = email_index[email]
                match_type = "email"
            elif phone and phone in phone_index:
                match = phone_index[phone]
                match_type = "phone"
            elif key and key in name_company_index:
                matches = name_company_index[key]
                if len(matches) == 1:
                    match = matches[0]
                    match_type = "name_company"
                else:
                    ambiguous.append(profile)
                    stats["ambiguous_existing"] += 1
                    continue

            if match:
                existing_matches.append({"profile": profile, "match_type": match_type, "matched_recruiter_id": match.recruiter_id})
                stats[f"matched_{match_type}"] += 1
            else:
                unique_candidates.append(profile)
                stats["unique"] += 1

        # Build company summaries for unique profiles
        company_profiles = defaultdict(list)
        for profile in unique_candidates:
            company_profiles[norm_key(profile.get("company"))].append(profile)

        new_company_objects = []
        created_company_keys = set()
        for company_key, profiles_in_company in company_profiles.items():
            if not company_key or company_key in company_cache:
                continue
            company_name = next((profile.get("company") for profile in profiles_in_company if profile.get("company")), "")
            state_counts = Counter()
            location_counts = Counter()
            for profile in profiles_in_company:
                if profile.get("dominant_state"):
                    state_counts[profile["dominant_state"]] += 1
                if profile.get("locations"):
                    location_counts[profile["locations"][0]] += 1
            dominant_state = state_counts.most_common(1)[0][0] if state_counts else None
            dominant_location = location_counts.most_common(1)[0][0] if location_counts else None
            company = Company(
                company_name=company_name,
                normalized_company_name=company_key,
                location=dominant_location,
                state=dominant_state,
                data_source="location_workbook",
                source_job_id=BATCH_ID,
                is_active=True,
            )
            new_company_objects.append(company)
            created_company_keys.add(company_key)
        if new_company_objects:
            session.add_all(new_company_objects)
            session.flush()
            for company in new_company_objects:
                company_cache[company.normalized_company_name or norm_key(company.company_name)] = company

        # Optional: enrich existing recruiters with location/state from workbook match only.
        enriched_count = 0
        if args.apply_existing_enrichment and existing_matches:
            for match in existing_matches:
                profile = match["profile"]
                recruiter = email_index.get(profile.get("email")) or (phone_index.get(profile.get("phone")) if profile.get("phone") else None)
                if not recruiter and profile.get("name") and profile.get("company"):
                    recruiter_matches = name_company_index.get(f"{norm_key(profile.get('name'))}::{norm_key(profile.get('company'))}", [])
                    recruiter = recruiter_matches[0] if len(recruiter_matches) == 1 else None
                if not recruiter:
                    continue
                company_name = profile.get("company")
                company_key = norm_key(company_name)
                company = company_cache.get(company_key)
                if not company and company_name:
                    company = Company(
                        company_name=company_name,
                        normalized_company_name=company_key,
                        location=profile.get("locations")[0] if profile.get("locations") else None,
                        state=profile.get("dominant_state"),
                        data_source="location_workbook",
                        source_job_id=BATCH_ID,
                        is_active=True,
                    )
                    session.add(company)
                    session.flush()
                    company_cache[company_key] = company
                if company and recruiter.company_id != company.company_id:
                    recruiter.company_id = company.company_id
                if not recruiter.location and profile.get("locations"):
                    recruiter.location = profile["locations"][0]
                if not recruiter.state and profile.get("dominant_state"):
                    recruiter.state = profile["dominant_state"]
                    recruiter.state_source = "location_workbook"
                    recruiter.state_confidence = "high"
                    recruiter.state_reason = "location_workbook_dominant_state"
                if recruiter.needs_review and (recruiter.state or recruiter.location):
                    recruiter.needs_review = False
                    recruiter.review_reason = None
                recruiter.metadata_json = merge_metadata(
                    recruiter.metadata_json,
                    {
                        "batch_id": BATCH_ID,
                        "source_file": str(SOURCE_FILE),
                        "match_type": match["match_type"],
                        "workbook_name": profile.get("name"),
                        "workbook_email": profile.get("email"),
                        "workbook_phone": profile.get("phone"),
                        "workbook_company": profile.get("company"),
                        "workbook_locations": profile.get("locations"),
                        "workbook_dominant_state": profile.get("dominant_state"),
                        "workbook_sheet_names": profile.get("sheet_names"),
                    },
                )
                recruiter.last_scan_at = datetime.now(timezone.utc)
                enriched_count += 1
            session.commit()

        recruiter_rows = []
        inserted_examples = []
        for profile in unique_candidates:
            company_name = profile.get("company")
            company_key = norm_key(company_name)
            company = company_cache.get(company_key)
            if not company and company_name:
                # Create late if this company only appears among unique candidates and wasn't pre-created.
                company = Company(
                    company_name=company_name,
                    normalized_company_name=company_key,
                    location=profile.get("locations")[0] if profile.get("locations") else None,
                    state=profile.get("dominant_state"),
                    data_source="location_workbook",
                    source_job_id=BATCH_ID,
                    is_active=True,
                )
                session.add(company)
                session.flush()
                company_cache[company_key] = company
            recruiter_rows.append(
                {
                    "recruiter_name": profile.get("name") or "Unknown",
                    "normalized_recruiter_name": norm_key(profile.get("name")) or None,
                    "email": build_insert_email(profile),
                    "phone": profile.get("phone"),
                    "company_id": company.company_id if company else None,
                    "location": profile.get("locations")[0] if profile.get("locations") else None,
                    "state": profile.get("dominant_state"),
                    "state_source": "location_workbook" if profile.get("dominant_state") else None,
                    "state_confidence": "high" if profile.get("dominant_state") else None,
                    "state_reason": "location_workbook_dominant_state" if profile.get("dominant_state") else None,
                    "needs_review": not bool(profile.get("dominant_state")),
                    "review_reason": None if profile.get("dominant_state") else "Workbook row qualified by phone/location but state could not be derived",
                    "data_source": "location_workbook",
                    "source_job_id": BATCH_ID,
                    "raw_data": json.dumps(profile.get("source_rows", []), default=str),
                    "metadata_json": json.dumps(
                        {
                            "batch_id": BATCH_ID,
                            "source_file": str(SOURCE_FILE),
                            "workbook_name": profile.get("name"),
                            "workbook_email": profile.get("email"),
                            "workbook_phone": profile.get("phone"),
                            "workbook_company": company_name,
                            "workbook_locations": profile.get("locations"),
                            "workbook_sheet_names": profile.get("sheet_names"),
                            "workbook_source_row_count": profile.get("source_row_count"),
                            "workbook_dominant_state": profile.get("dominant_state"),
                        },
                        default=str,
                    ),
                }
            )
            if len(inserted_examples) < 20:
                inserted_examples.append(
                    {
                        "name": profile.get("name"),
                        "email": profile.get("email"),
                        "company": company_name,
                        "phone": profile.get("phone"),
                        "state": profile.get("dominant_state"),
                        "location": profile.get("locations")[0] if profile.get("locations") else None,
                    }
                )

        inserted_recruiters = 0
        for i in range(0, len(recruiter_rows), args.batch_size):
            batch = recruiter_rows[i:i + args.batch_size]
            session.bulk_insert_mappings(Recruiter, batch)
            session.commit()
            inserted_recruiters += len(batch)

        state_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NOT NULL AND state != ''")).scalar() or 0
        unknown_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NULL OR state = ''")).scalar() or 0
        total_recruiters = session.execute(text("SELECT COUNT(*) FROM recruiters")).scalar() or 0
        total_companies = session.execute(text("SELECT COUNT(*) FROM companies")).scalar() or 0
        state_counts = session.execute(text("""
            SELECT state, COUNT(*) AS cnt
            FROM recruiters
            WHERE state IS NOT NULL AND state != ''
            GROUP BY state
            ORDER BY cnt DESC, state ASC
        """)).mappings().all()

        report = {
            "batch_id": BATCH_ID,
            "source_file": str(SOURCE_FILE),
            "source_row_count": row_count,
            "merged_profiles": len(profiles),
            "existing_matches": len(existing_matches),
            "unique_candidates": len(unique_candidates),
            "skipped_no_phone_or_location": len(skipped),
            "ambiguous_existing": len(ambiguous),
            "stats": dict(stats),
            "created_company_keys": sorted(created_company_keys),
            "new_company_count": len(new_company_objects),
            "enriched_existing_count": enriched_count,
            "inserted_recruiters": inserted_recruiters,
            "total_recruiters": total_recruiters,
            "total_companies": total_companies,
            "state_before": int(state_before),
            "unknown_before": int(unknown_before),
            "state_counts": {row["state"]: int(row["cnt"]) for row in state_counts},
            "inserted_examples": inserted_examples,
            "skipped_examples": skipped[:20],
            "ambiguous_examples": ambiguous[:20],
        }

        json_path = OUTPUT_DIR / "location_workbook_unique_apply_report.json"
        md_path = OUTPUT_DIR / "location_workbook_unique_apply_report.md"
        json_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
        md_lines = [
            "# Location Workbook Unique Apply Report",
            "",
            f"Source workbook: `{SOURCE_FILE}`",
            f"Batch id: `{BATCH_ID}`",
            "",
            "## Results",
            f"- Source rows: `{row_count}`",
            f"- Merged profiles: `{len(profiles)}`",
            f"- Existing matches: `{len(existing_matches)}`",
            f"- Unique candidates: `{len(unique_candidates)}`",
            f"- Skipped missing phone/location: `{len(skipped)}`",
            f"- Ambiguous existing: `{len(ambiguous)}`",
            f"- New companies created: `{len(new_company_objects)}`",
            f"- Existing recruiters enriched: `{enriched_count}`",
            f"- Recruiters inserted: `{inserted_recruiters}`",
            "",
            "## Files",
            f"- JSON: `{json_path}`",
            f"- Markdown: `{md_path}`",
        ]
        md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")
        print(json.dumps(report, indent=2, default=str))

    finally:
        session.close()


if __name__ == "__main__":
    main()
