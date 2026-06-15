from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))

from app.database import SessionLocal
from app.models.models import Company, Recruiter
from app.utils.state_mapper import extract_state_detailed


SOURCE_FILE = Path(r"C:\Users\User\Desktop\final updated sheet.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
REVIEW_QUEUE_JSON = OUTPUT_DIR / "workbook_review_queue.json"

STATE_CODES = {
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA",
    "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT",
    "VA", "WA", "WV", "WI", "WY", "DC",
}

COMPANY_KEYWORDS = [
    "llc", "inc", "group", "consulting", "solutions", "technology", "technologies",
    "services", "partners", "staffing", "resources", "search", "talent",
]

TITLE_KEYWORDS = [
    "recruiter", "manager", "sourcer", "talent", "director", "executive",
    "specialist", "head", "lead", "partner",
]

GENERIC_DOMAINS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com", "icloud.com",
    "proton.me", "protonmail.com", "mail.com", "msn.com", "live.com", "yandex.com", "gmx.com",
}

GENERIC_NAME_TOKENS = {
    "unknown",
    "n/a",
    "na",
    "none",
    "null",
    "-",
    "name",
    "recruiter",
    "contact",
    "person",
}

AUTO_FILL_CATEGORIES = {
    "exact_email_state",
    "exact_phone_state",
    "exact_name_company_state",
    "workbook_company_majority_90",
    "db_company_majority_90",
    "domain_company_agree",
}

MANUAL_REVIEW_CATEGORIES = {
    "company_majority_70_90",
    "domain_majority_multistate",
    "ambiguous_city",
    "conflicting_signals",
    "workbook_match_no_state",
    "workbook_email_no_state",
    "workbook_phone_no_state",
    "generic_email_only",
}

TRULY_UNKNOWN_CATEGORIES = {
    "no_workbook_match",
    "no_contact_evidence",
    "no_company_no_location",
}


def norm_company(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def norm_phone(value: str | None) -> str | None:
    digits = re.sub(r"[^0-9]", "", value or "")
    return digits[-10:] if len(digits) >= 10 else None


def normalize_name(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def name_score(value: str | None) -> tuple[int, int, int]:
    text = normalize_name(value)
    if not text:
        return (0, 0, 0)
    lowered = text.lower()
    tokens = [token for token in re.split(r"\s+", text) if token]
    if lowered in GENERIC_NAME_TOKENS or "@" in text or any(char.isdigit() for char in text):
        return (0, len(tokens), len(text))
    return (1 if len(tokens) >= 2 else 0, len(tokens), len(text))


def should_promote_name(existing: str | None, candidate: str | None, email: str | None = None) -> bool:
    existing_text = normalize_name(existing)
    candidate_text = normalize_name(candidate)
    if not candidate_text:
        return False
    candidate_lower = candidate_text.lower()
    if candidate_lower in GENERIC_NAME_TOKENS or "@" in candidate_text:
        return False
    if not existing_text:
        return True

    email_localpart = (email or "").strip().lower().split("@", 1)[0]
    existing_lower = existing_text.lower()
    if existing_lower in GENERIC_NAME_TOKENS or (email_localpart and existing_lower == email_localpart):
        return True

    existing_score = name_score(existing_text)
    candidate_score = name_score(candidate_text)
    if candidate_score > existing_score:
        return True

    existing_tokens = len([token for token in re.split(r"\s+", existing_text) if token])
    candidate_tokens = len([token for token in re.split(r"\s+", candidate_text) if token])
    if candidate_tokens >= 2 and existing_tokens <= 1:
        return True
    if len(candidate_text) >= len(existing_text) + 3 and candidate_tokens >= existing_tokens:
        return True
    return False


def parse_metadata(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return dict(value)
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {"value": parsed}
    except Exception:
        return {"raw_metadata": str(value)}


def extract_state_from_text(value: str | None):
    if not value:
        return None, None
    state, reason = extract_state_detailed(value)
    if state:
        return state, reason
    return None, None


def derive_workbook_state(person: dict) -> tuple[str | None, str | None, str | None]:
    for source_label in ("state", "location"):
        raw_value = person.get(source_label)
        if not raw_value:
            continue
        state, reason = extract_state_from_text(raw_value)
        if state:
            return state, source_label, reason
    return None, None, None


def parse_block(block):
    person = {
        "emails": [],
        "phones": [],
        "name": "",
        "company": "",
        "title": "",
        "state": "",
        "location": "",
        "notes": [],
        "sheet_names": set(),
    }
    for _, value in block:
        value = str(value).strip()
        if not value or value.lower() == "nan":
            continue
        if re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", value):
            person["emails"].append(value.lower())
            continue
        phone_match = re.search(r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", value)
        if phone_match:
            digits = re.sub(r"[^0-9]", "", phone_match.group(0))
            if len(digits) >= 10:
                person["phones"].append(digits[-10:])
            continue
        if len(value) == 2 and value.upper() in STATE_CODES:
            person["state"] = value.upper()
            continue
        if "," in value and len(value.split(",")[-1].strip()) == 2 and value.split(",")[-1].strip().upper() in STATE_CODES:
            person["location"] = value.title()
            person["state"] = value.split(",")[-1].strip().upper()
            continue
        if any(keyword in value.lower() for keyword in COMPANY_KEYWORDS):
            if not person["company"]:
                person["company"] = value.title()
            else:
                person["notes"].append(value)
            continue
        if any(keyword in value.lower() for keyword in TITLE_KEYWORDS):
            person["title"] = value.title()
            continue
        if len(value.split()) in [2, 3] and not re.search(r"\d", value) and len(value) < 30 and not person["name"]:
            person["name"] = value.title()
            continue
        person["notes"].append(value)

    if not person["name"] and person["notes"]:
        first_note = person["notes"][0]
        if len(first_note.split()) in [2, 3] and len(first_note) < 30:
            person["name"] = first_note.title()
            person["notes"].pop(0)

    if not person["company"] and person["emails"]:
        domain = person["emails"][0].split("@")[-1]
        if domain not in GENERIC_DOMAINS:
            person["company"] = domain.split(".")[0].title()

    return person


def person_key(person):
    if person["emails"]:
        return f"email::{person['emails'][0].lower()}"
    if person["phones"]:
        return f"phone::{person['phones'][0]}"
    if person["name"] and person["company"]:
        return f"name_company::{person['name'].lower()}::{norm_company(person['company'])}"
    if person["name"]:
        return f"name::{person['name'].lower()}"
    return None


def merge_person(existing, incoming, sheet_name):
    existing["emails"] = list(dict.fromkeys(existing["emails"] + incoming["emails"]))
    existing["phones"] = list(dict.fromkeys(existing["phones"] + incoming["phones"]))
    existing["sheet_names"].update(incoming.get("sheet_names", set()))
    existing["sheet_names"].add(sheet_name)
    if not existing["name"] and incoming["name"]:
        existing["name"] = incoming["name"]
    if not existing["company"] and incoming["company"]:
        existing["company"] = incoming["company"]
    if not existing["title"] and incoming["title"]:
        existing["title"] = incoming["title"]
    if not existing["state"] and incoming["state"]:
        existing["state"] = incoming["state"]
    if not existing["location"] and incoming["location"]:
        existing["location"] = incoming["location"]
    if incoming["notes"]:
        existing["notes"].extend(incoming["notes"])
    return existing


def load_workbook_people(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    people_by_key = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row = [str(cell).strip() if cell is not None else "" for cell in row]
            blocks = []
            current_block = []
            for index, cell in enumerate(row):
                if not cell or cell.lower() in {"nan", "none"}:
                    continue
                if not current_block or index - current_block[-1][0] <= 5:
                    current_block.append((index, cell))
                else:
                    blocks.append(current_block)
                    current_block = [(index, cell)]
            if current_block:
                blocks.append(current_block)

            for block in blocks:
                text = " ".join(cell.lower() for _, cell in block)
                if "contact #1" in text and "email id" in text and "@" not in text:
                    continue
                person = parse_block(block)
                if not (person["emails"] or person["phones"] or (person["name"] and person["company"])):
                    continue
                person["sheet_names"].add(sheet_name)
                key = person_key(person)
                if not key:
                    key = f"sheet::{sheet_name}::row::{len(people_by_key)}"
                if key in people_by_key:
                    people_by_key[key] = merge_person(people_by_key[key], person, sheet_name)
                else:
                    people_by_key[key] = person

    people = list(people_by_key.values())
    for person in people:
        person["sheet_names"] = sorted(person.get("sheet_names", set()))
    return people


def build_workbook_indices(people):
    by_email = {email: person for person in people for email in person["emails"]}
    by_phone = {}
    by_name_company = {}
    company_states = defaultdict(Counter)
    domain_states = defaultdict(Counter)

    for person in people:
        for phone in person["phones"]:
            by_phone[phone] = person
        if person["name"] and person["company"]:
            by_name_company[f"{person['name'].lower()}::{norm_company(person['company'])}"] = person
        state, _, _ = derive_workbook_state(person)
        if person["emails"] and state:
            domain = person["emails"][0].split("@")[-1].lower()
            if domain not in GENERIC_DOMAINS:
                domain_states[domain][state] += 1
        if person["company"] and state:
            company_states[norm_company(person["company"])][state] += 1

    strong_company = {}
    company_summary = {}
    for company_key, counts in company_states.items():
        total = sum(counts.values())
        if not total:
            continue
        sorted_counts = counts.most_common()
        state, count = sorted_counts[0]
        ratio = count / total
        company_summary[company_key] = {
            "total": total,
            "state": state,
            "count": count,
            "ratio": ratio,
            "states": dict(counts),
        }
        if total >= 10 and ratio >= 0.9:
            strong_company[company_key] = company_summary[company_key]

    strong_domain = {}
    for domain, counts in domain_states.items():
        total = sum(counts.values())
        if not total:
            continue
        sorted_counts = counts.most_common()
        state, count = sorted_counts[0]
        ratio = count / total
        if total >= 10 and ratio >= 0.9:
            strong_domain[domain] = {
                "total": total,
                "state": state,
                "count": count,
                "ratio": ratio,
                "states": dict(counts),
            }

    return by_email, by_phone, by_name_company, strong_company, strong_domain, company_summary


def build_db_company_majority(session):
    rows = session.execute(text("""
        SELECT company_id, state, COUNT(*) AS cnt
        FROM recruiters
        WHERE company_id IS NOT NULL AND state IS NOT NULL AND state != ''
        GROUP BY company_id, state
    """)).mappings().all()
    buckets = {}
    for row in rows:
        company_id = int(row["company_id"])
        state = row["state"]
        cnt = int(row["cnt"])
        bucket = buckets.setdefault(company_id, Counter())
        bucket[state] += cnt
    summary = {}
    for company_id, counts in buckets.items():
        total = sum(counts.values())
        state, count = counts.most_common(1)[0]
        ratio = count / total if total else 0
        summary[company_id] = {
            "total": total,
            "state": state,
            "count": count,
            "ratio": ratio,
            "states": dict(counts),
        }
    return summary


def merge_metadata(existing_value, evidence):
    metadata = parse_metadata(existing_value)
    metadata["state_recovery_evidence"] = evidence
    return json.dumps(metadata, default=str)


def state_source_rank(source: str | None, confidence: str | None) -> int:
    source = (source or "").strip()
    confidence = (confidence or "").strip().lower()
    if source in {"workbook_email", "workbook_phone", "workbook_name_company"}:
        return 100
    if source in {"workbook_company_majority", "workbook_domain_plus_company"}:
        return 90
    if source in {"db_company_majority", "company_majority_state", "workbook_domain_majority"}:
        return 80
    if source in {"location_workbook", "companies_state_workbook"}:
        return 70 if confidence == "high" else 60
    if source in {"email_domain", "notes", "metadata_json"}:
        return 50 if confidence == "medium" else 40
    if source in {"raw_data", "company_location", "recruiter_location"}:
        return 45 if confidence == "high" else 35
    return 0


def should_override_existing_state(recruiter, classification) -> bool:
    current_rank = state_source_rank(getattr(recruiter, "state_source", None), getattr(recruiter, "state_confidence", None))
    proposed_rank = state_source_rank(classification.get("source"), classification.get("confidence"))
    if proposed_rank > current_rank:
        return True
    if current_rank == 0 and classification.get("bucket") == "auto_fill_safe":
        return True
    return False


def _company_name_from_map(company_map, recruiter):
    if recruiter.company_id is None:
        return ""
    return company_map.get(recruiter.company_id, "") or ""


def classify_recruiter(
    recruiter,
    company_name: str,
    workbook_indices,
    workbook_company_summary,
    db_company_summary,
    company_map_by_key,
):
    by_email, by_phone, by_name_company, strong_company, strong_domain, company_summary = workbook_indices

    current_state = (recruiter.state or "").strip().upper() or None
    email = (recruiter.email or "").lower().strip()
    domain = email.split("@")[-1].lower() if "@" in email else None
    phone = norm_phone(recruiter.phone)
    company_key = norm_company(company_name)
    recruiter_name = (recruiter.recruiter_name or "").strip()
    name_company_key = f"{recruiter_name.lower()}::{company_key}" if recruiter_name and company_key else None
    workbook_person = by_email.get(email) if email else None
    current_company = company_map_by_key.get(recruiter.company_id) if recruiter.company_id else None

    workbook_person_state = None
    workbook_person_state_source = None
    workbook_person_state_reason = None
    workbook_location = None
    workbook_company = None
    workbook_phone = None
    workbook_name = None
    if workbook_person:
        workbook_person_state, workbook_person_state_source, workbook_person_state_reason = derive_workbook_state(workbook_person)
        workbook_location = workbook_person.get("location") or ""
        workbook_company = workbook_person.get("company") or ""
        workbook_phone_list = workbook_person.get("phones") or []
        workbook_phone = workbook_phone_list[0] if workbook_phone_list else None
        workbook_name = workbook_person.get("name") or ""

    exact_phone_person = by_phone.get(phone) if phone else None
    exact_name_company_person = by_name_company.get(name_company_key) if name_company_key else None

    # High confidence: exact workbook email, phone, or name+company with explicit state/location evidence.
    if workbook_person and workbook_person_state:
        return {
            "bucket": "auto_fill_safe",
            "category": "exact_email_state",
            "proposed_state": workbook_person_state,
            "confidence": "high",
            "source": "workbook_email",
            "reason": workbook_person_state_reason or "workbook_email_state",
            "evidence": {
                "matched_email": email,
                "workbook_name": workbook_name,
                "workbook_company": workbook_company,
                "workbook_location": workbook_location,
                "workbook_state": workbook_person_state,
                "workbook_state_source": workbook_person_state_source,
                "current_state": current_state,
            },
        }

    if exact_phone_person:
        phone_state, phone_state_source, phone_state_reason = derive_workbook_state(exact_phone_person)
        if phone_state:
            return {
                "bucket": "auto_fill_safe",
                "category": "exact_phone_state",
                "proposed_state": phone_state,
                "confidence": "high",
                "source": "workbook_phone",
                "reason": phone_state_reason or "workbook_phone_state",
                "evidence": {
                    "matched_phone": phone,
                    "workbook_name": exact_phone_person.get("name"),
                    "workbook_company": exact_phone_person.get("company"),
                    "workbook_location": exact_phone_person.get("location"),
                    "workbook_state": phone_state,
                    "workbook_state_source": phone_state_source,
                    "current_state": current_state,
                },
            }

    if exact_name_company_person:
        name_state, name_state_source, name_state_reason = derive_workbook_state(exact_name_company_person)
        if name_state:
            return {
                "bucket": "auto_fill_safe",
                "category": "exact_name_company_state",
                "proposed_state": name_state,
                "confidence": "high",
                "source": "workbook_name_company",
                "reason": name_state_reason or "workbook_name_company_state",
                "evidence": {
                    "matched_name": recruiter_name,
                    "matched_company": company_name,
                    "workbook_name": exact_name_company_person.get("name"),
                    "workbook_company": exact_name_company_person.get("company"),
                    "workbook_location": exact_name_company_person.get("location"),
                    "workbook_state": name_state,
                    "workbook_state_source": name_state_source,
                    "current_state": current_state,
                },
            }

    if company_key in strong_company:
        info = strong_company[company_key]
        return {
            "bucket": "auto_fill_safe",
            "category": "workbook_company_majority_90",
            "proposed_state": info["state"],
            "confidence": "high",
            "source": "workbook_company_majority",
            "reason": f"workbook_company_majority:{info['count']}/{info['total']}",
            "evidence": {
                "company_name": company_name,
                "majority_state": info["state"],
                "matching_rows": info["count"],
                "total_rows": info["total"],
                "ratio": round(info["ratio"], 3),
                "current_state": current_state,
            },
        }

    if recruiter.company_id in db_company_summary:
        info = db_company_summary[recruiter.company_id]
        if info["total"] >= 10 and info["ratio"] >= 0.9:
            return {
                "bucket": "auto_fill_safe",
                "category": "db_company_majority_90",
                "proposed_state": info["state"],
                "confidence": "high",
                "source": "db_company_majority",
                "reason": f"db_company_majority:{info['count']}/{info['total']}",
                "evidence": {
                    "company_name": company_name,
                    "majority_state": info["state"],
                    "matching_rows": info["count"],
                    "total_rows": info["total"],
                    "ratio": round(info["ratio"], 3),
                    "state_breakdown": info["states"],
                    "current_state": current_state,
                },
            }
        if info["total"] >= 5 and info["ratio"] >= 0.7:
            return {
                "bucket": "needs_manual_review",
                "category": "company_majority_70_90",
                "proposed_state": info["state"],
                "confidence": "medium",
                "source": "db_company_majority",
                "reason": f"db_company_majority:{info['count']}/{info['total']}",
                "evidence": {
                    "company_name": company_name,
                    "majority_state": info["state"],
                    "matching_rows": info["count"],
                    "total_rows": info["total"],
                    "ratio": round(info["ratio"], 3),
                    "state_breakdown": info["states"],
                    "current_state": current_state,
                },
            }

    if domain and domain in strong_domain:
        info = strong_domain[domain]
        company_info = strong_company.get(company_key) or company_summary.get(company_key)
        if company_info and company_info["state"] == info["state"]:
            return {
                "bucket": "auto_fill_safe",
                "category": "domain_company_agree",
                "proposed_state": info["state"],
                "confidence": "high",
                "source": "workbook_domain_plus_company",
                "reason": f"domain_company_agree:{info['count']}/{info['total']}",
                "evidence": {
                    "email_domain": domain,
                    "domain_state": info["state"],
                    "domain_rows": info["count"],
                    "domain_total": info["total"],
                    "company_state": company_info["state"],
                    "company_rows": company_info["count"],
                    "company_total": company_info["total"],
                    "current_state": current_state,
                },
            }
        return {
            "bucket": "needs_manual_review",
            "category": "domain_majority_multistate",
            "proposed_state": info["state"],
            "confidence": "medium",
            "source": "workbook_domain_majority",
            "reason": f"domain_majority_multistate:{info['count']}/{info['total']}",
            "evidence": {
                "email_domain": domain,
                "domain_state": info["state"],
                "domain_rows": info["count"],
                "domain_total": info["total"],
                "domain_state_breakdown": info["states"],
                "company_name": company_name,
                "current_state": current_state,
            },
        }

    if workbook_person:
        state, state_source, state_reason = derive_workbook_state(workbook_person)
        if state:
            return {
                "bucket": "needs_manual_review",
                "category": "conflicting_signals",
                "proposed_state": state,
                "confidence": "medium",
                "source": "workbook_partial_conflict",
                "reason": state_reason or "workbook_partial_state",
                "evidence": {
                    "matched_email": email,
                    "workbook_name": workbook_person.get("name"),
                    "workbook_company": workbook_person.get("company"),
                    "workbook_location": workbook_person.get("location"),
                    "workbook_state": state,
                    "workbook_state_source": state_source,
                    "current_state": current_state,
                },
            }
        if workbook_person.get("location"):
            location_text = workbook_person["location"]
            if any(token in location_text.upper() for token in ("METRO", "AREA", "COUNTY", "REGION", "GREATER")):
                return {
                    "bucket": "needs_manual_review",
                    "category": "ambiguous_city",
                    "proposed_state": None,
                    "confidence": "low",
                    "source": "workbook_ambiguous_location",
                    "reason": "ambiguous_city_or_metro",
                    "evidence": {
                        "matched_email": email,
                        "workbook_location": location_text,
                        "current_state": current_state,
                    },
                }
        return {
            "bucket": "needs_manual_review",
            "category": "workbook_match_no_state",
            "proposed_state": None,
            "confidence": "low",
            "source": "workbook_match",
            "reason": "workbook_match_found_no_state",
            "evidence": {
                "matched_email": email,
                "workbook_company": workbook_person.get("company"),
                "workbook_location": workbook_person.get("location"),
                "current_state": current_state,
            },
        }

    if email and domain in GENERIC_DOMAINS:
        return {
            "bucket": "truly_unknown",
            "category": "generic_email_only",
            "proposed_state": None,
            "confidence": "none",
            "source": "generic_email",
            "reason": "generic_email_domain_no_signal",
            "evidence": {
                "email_domain": domain,
                "current_state": current_state,
                "company_name": company_name,
            },
        }

    if not company_name and not (recruiter.location or "") and not (recruiter.notes or "") and not (recruiter.raw_data or "") and not (recruiter.metadata_json or ""):
        return {
            "bucket": "truly_unknown",
            "category": "no_company_no_location",
            "proposed_state": None,
            "confidence": "none",
            "source": "no_evidence",
            "reason": "no_company_no_location_no_workbook_match",
            "evidence": {
                "email": email,
                "phone": phone,
                "current_state": current_state,
            },
        }

    return {
        "bucket": "truly_unknown",
        "category": "no_workbook_match",
        "proposed_state": None,
        "confidence": "none",
        "source": "no_workbook_match",
        "reason": "no_workbook_match",
        "evidence": {
            "email": email,
            "phone": phone,
            "company_name": company_name,
            "current_state": current_state,
        },
    }


def snapshot_recruiter(recruiter, company_name, classification):
    return {
        "recruiter_id": recruiter.recruiter_id,
        "recruiter_name": recruiter.recruiter_name,
        "email": recruiter.email,
        "phone": recruiter.phone,
        "company_name": company_name,
        "current_state": recruiter.state,
        "needs_review": bool(recruiter.needs_review),
        "review_reason": recruiter.review_reason,
        "bucket": classification["bucket"],
        "category": classification["category"],
        "proposed_state": classification["proposed_state"],
        "confidence": classification["confidence"],
        "source": classification["source"],
        "reason": classification["reason"],
        "evidence": classification["evidence"],
    }


def load_review_queue_path():
    if REVIEW_QUEUE_JSON.exists():
        return REVIEW_QUEUE_JSON
    return None


def write_report(report, json_path: Path, md_path: Path):
    json_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")

    lines = [
        "# Workbook Review Queue Second Pass",
        "",
        f"Source workbook: `{SOURCE_FILE}`",
        "",
        f"Dry run: `{report['dry_run']}`",
        "",
        "## Counts",
        f"- Total inspected: `{report['total_inspected']}`",
        f"- Known state before: `{report['known_state_before']}`",
        f"- Known state after: `{report['known_state_after']}`",
        f"- Unknown state before: `{report['unknown_state_before']}`",
        f"- Unknown state after: `{report['unknown_state_after']}`",
        f"- Needs review before: `{report['needs_review_before']}`",
        f"- Needs review after: `{report['needs_review_after']}`",
        f"- Auto-filled count: `{report['auto_filled_count']}`",
        f"- Still review count: `{report['still_review_count']}`",
        f"- Truly unknown count: `{report['truly_unknown_count']}`",
        "",
        "## Recovery Sources",
    ]
    for source, count in report["source_counts"].items():
        lines.append(f"- {source}: `{count}`")
    lines.extend([
        "",
        "## State Counts",
    ])
    for state, count in report["state_counts"].items():
        lines.append(f"- {state}: `{count}`")
    lines.extend([
        "",
        "## Sample Auto-Filled",
    ])
    for item in report["auto_filled_samples"]:
        lines.append(f"- `{item['recruiter_id']}` {item['recruiter_name']} | {item['email']} | {item['proposed_state']} | {item['source']}")
    lines.extend([
        "",
        "## Sample Still Review",
    ])
    for item in report["still_review_samples"]:
        lines.append(f"- `{item['recruiter_id']}` {item['recruiter_name']} | {item['email']} | {item['category']}")
    lines.extend([
        "",
        "## Sample Truly Unknown",
    ])
    for item in report["truly_unknown_samples"]:
        lines.append(f"- `{item['recruiter_id']}` {item['recruiter_name']} | {item['email']} | {item['category']}")
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="Apply safe updates to the database.")
    parser.add_argument("--batch-size", type=int, default=500)
    parser.add_argument("--sample-size", type=int, default=20)
    args = parser.parse_args()

    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Workbook not found: {SOURCE_FILE}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print({"phase": "load_workbook"}, flush=True)
    people = load_workbook_people(SOURCE_FILE)
    workbook_indices = build_workbook_indices(people)

    session = SessionLocal()
    try:
        company_rows = session.query(Company.company_id, Company.company_name).all()
        company_map = {company_id: company_name for company_id, company_name in company_rows}
        company_map_by_key = {company_id: company_name for company_id, company_name in company_rows}
        db_company_summary = build_db_company_majority(session)

        # First, do a name-only recovery sweep across the whole local DB.
        name_only_updates = []
        all_recruiters = session.query(Recruiter).all()
        for recruiter in all_recruiters:
            company_name = company_map.get(recruiter.company_id, "") if recruiter.company_id else ""
            email = (recruiter.email or "").lower().strip()
            phone = norm_phone(recruiter.phone)
            company_key = norm_company(company_name)
            name_company_key = f"{normalize_name(recruiter.recruiter_name).lower()}::{company_key}" if recruiter.recruiter_name and company_key else None
            workbook_person = None
            if email and email in workbook_indices[0]:
                workbook_person = workbook_indices[0][email]
            elif phone and phone in workbook_indices[1]:
                workbook_person = workbook_indices[1][phone]
            elif name_company_key and name_company_key in workbook_indices[2]:
                workbook_person = workbook_indices[2][name_company_key]

            if not workbook_person:
                continue

            workbook_name = normalize_name(workbook_person.get("name"))
            if not should_promote_name(recruiter.recruiter_name, workbook_name, recruiter.email):
                continue

            evidence = {
                "source": "workbook_name_recovery",
                "workbook_name": workbook_name,
                "workbook_email": (workbook_person.get("emails") or [None])[0],
                "workbook_phone": (workbook_person.get("phones") or [None])[0],
                "workbook_company": workbook_person.get("company"),
                "workbook_location": workbook_person.get("location"),
                "workbook_state": derive_workbook_state(workbook_person)[0],
                "inspected_at": datetime.now(timezone.utc).isoformat(),
                "existing_name": recruiter.recruiter_name,
            }
            name_only_updates.append(
                {
                    "id": recruiter.recruiter_id,
                    "recruiter_name": workbook_name[:150],
                    "normalized_recruiter_name": workbook_name.lower().replace(" ", "")[:150],
                    "metadata_json": merge_metadata(recruiter.metadata_json, evidence),
                    "last_scan_at": datetime.now(timezone.utc),
                }
            )

        if name_only_updates:
            update_sql = text("""
                UPDATE recruiters
                SET recruiter_name = COALESCE(:recruiter_name, recruiter_name),
                    normalized_recruiter_name = COALESCE(:normalized_recruiter_name, normalized_recruiter_name),
                    metadata_json = :metadata_json,
                    last_scan_at = :last_scan_at
                WHERE recruiter_id = :id
            """)
            for i in range(0, len(name_only_updates), args.batch_size if 'args' in locals() else 500):
                session.execute(update_sql, name_only_updates[i:i + (args.batch_size if 'args' in locals() else 500)])
                session.commit()

        recruiter_filter = session.query(Recruiter).filter(
            (Recruiter.state.is_(None)) | (Recruiter.state == "") | (Recruiter.needs_review == True)
        )
        targets = recruiter_filter.all()

        known_state_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NOT NULL AND state != ''")).scalar() or 0
        unknown_state_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NULL OR state = ''")).scalar() or 0
        needs_review_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE needs_review = true")).scalar() or 0

        state_counts_before = session.execute(text("""
            SELECT state, COUNT(*) AS cnt
            FROM recruiters
            WHERE state IS NOT NULL AND state != ''
            GROUP BY state
            ORDER BY cnt DESC, state ASC
        """)).mappings().all()

        auto_updates = []
        review_updates = []
        source_counts = Counter()
        bucket_counts = Counter()
        auto_state_fills = 0
        auto_review_cleanups = 0
        auto_conflicts = 0
        auto_filled_samples = []
        still_review_samples = []
        truly_unknown_samples = []
        all_inspected = []

        for recruiter in targets:
            company_name = company_map.get(recruiter.company_id, "") if recruiter.company_id else ""
            classification = classify_recruiter(
                recruiter,
                company_name,
                workbook_indices,
                workbook_indices[5],  # workbook company summary
                db_company_summary,
                company_map,
            )
            bucket_counts[classification["bucket"]] += 1
            source_counts[classification["source"]] += 1
            record = snapshot_recruiter(recruiter, company_name, classification)
            all_inspected.append(record)

            current_state = (recruiter.state or "").strip().upper() or None
            proposed_state = classification["proposed_state"]
            evidence = dict(classification["evidence"])
            evidence["category"] = classification["category"]
            evidence["bucket"] = classification["bucket"]
            evidence["source"] = classification["source"]
            evidence["confidence"] = classification["confidence"]
            evidence["reason"] = classification["reason"]
            evidence["inspected_at"] = datetime.now(timezone.utc).isoformat()

            metadata_json = merge_metadata(recruiter.metadata_json, evidence)
            proposed_name = normalize_name(evidence.get("workbook_name") or evidence.get("matched_name") or evidence.get("workbook_person_name"))
            name_promoted = should_promote_name(recruiter.recruiter_name, proposed_name, recruiter.email)

            if classification["bucket"] == "auto_fill_safe":
                if current_state is None:
                    auto_state_fills += 1
                    auto_updates.append(
                        {
                            "id": recruiter.recruiter_id,
                            "recruiter_name": proposed_name if name_promoted else recruiter.recruiter_name,
                            "normalized_recruiter_name": normalize_name(proposed_name).lower().replace(" ", "") if name_promoted and proposed_name else recruiter.normalized_recruiter_name,
                            "state": proposed_state,
                            "state_source": classification["source"],
                            "state_confidence": classification["confidence"],
                            "state_reason": classification["reason"],
                            "needs_review": False,
                            "review_reason": None,
                            "metadata_json": metadata_json,
                            "last_scan_at": datetime.now(timezone.utc),
                        }
                    )
                elif current_state == proposed_state:
                    auto_review_cleanups += 1
                    auto_updates.append(
                        {
                            "id": recruiter.recruiter_id,
                            "recruiter_name": proposed_name if name_promoted else recruiter.recruiter_name,
                            "normalized_recruiter_name": normalize_name(proposed_name).lower().replace(" ", "") if name_promoted and proposed_name else recruiter.normalized_recruiter_name,
                            "state": current_state,
                            "state_source": classification["source"],
                            "state_confidence": classification["confidence"],
                            "state_reason": classification["reason"],
                            "needs_review": False if recruiter.needs_review else recruiter.needs_review,
                            "review_reason": None if recruiter.needs_review else recruiter.review_reason,
                            "metadata_json": metadata_json,
                            "last_scan_at": datetime.now(timezone.utc),
                        }
                    )
                else:
                    if should_override_existing_state(recruiter, classification):
                        auto_conflicts += 1
                        override_reason = (
                            f"Workbook override: db={current_state}({getattr(recruiter, 'state_source', None) or 'none'}) "
                            f"-> proposed={proposed_state}({classification['source']})"
                        )
                        auto_updates.append(
                            {
                                "id": recruiter.recruiter_id,
                                "recruiter_name": proposed_name if name_promoted else recruiter.recruiter_name,
                                "normalized_recruiter_name": normalize_name(proposed_name).lower().replace(" ", "") if name_promoted and proposed_name else recruiter.normalized_recruiter_name,
                                "state": proposed_state,
                                "state_source": classification["source"],
                                "state_confidence": classification["confidence"],
                                "state_reason": f"{classification['reason']} | {override_reason}",
                                "needs_review": False,
                                "review_reason": None,
                                "metadata_json": metadata_json,
                                "last_scan_at": datetime.now(timezone.utc),
                            }
                        )
                        if len(auto_filled_samples) < args.sample_size:
                            auto_filled_samples.append(record)
                        continue
                    auto_conflicts += 1
                    review_updates.append(
                        {
                            "id": recruiter.recruiter_id,
                            "needs_review": True,
                            "review_reason": f"State conflict: db={current_state} proposed={proposed_state}",
                            "metadata_json": metadata_json,
                            "last_scan_at": datetime.now(timezone.utc),
                        }
                    )
                    if len(still_review_samples) < args.sample_size:
                        still_review_samples.append(record)
                    continue

                if len(auto_filled_samples) < args.sample_size:
                    auto_filled_samples.append(record)
                continue

            if classification["bucket"] == "needs_manual_review":
                review_updates.append(
                    {
                        "id": recruiter.recruiter_id,
                        "recruiter_name": proposed_name if name_promoted else recruiter.recruiter_name,
                        "normalized_recruiter_name": normalize_name(proposed_name).lower().replace(" ", "") if name_promoted and proposed_name else recruiter.normalized_recruiter_name,
                        "needs_review": True,
                        "review_reason": classification["reason"][:500],
                        "metadata_json": metadata_json,
                        "last_scan_at": datetime.now(timezone.utc),
                    }
                )
                if len(still_review_samples) < args.sample_size:
                    still_review_samples.append(record)
                continue

            review_updates.append(
                {
                    "id": recruiter.recruiter_id,
                    "recruiter_name": proposed_name if name_promoted else recruiter.recruiter_name,
                    "normalized_recruiter_name": normalize_name(proposed_name).lower().replace(" ", "") if name_promoted and proposed_name else recruiter.normalized_recruiter_name,
                    "needs_review": True,
                    "review_reason": classification["reason"][:500],
                    "metadata_json": metadata_json,
                    "last_scan_at": datetime.now(timezone.utc),
                }
            )
            if len(truly_unknown_samples) < args.sample_size:
                truly_unknown_samples.append(record)

        if args.apply and auto_updates:
            update_sql = text("""
                UPDATE recruiters
                SET state = :state,
                    recruiter_name = COALESCE(:recruiter_name, recruiter_name),
                    normalized_recruiter_name = COALESCE(:normalized_recruiter_name, normalized_recruiter_name),
                    state_source = :state_source,
                    state_confidence = :state_confidence,
                    state_reason = :state_reason,
                    needs_review = :needs_review,
                    review_reason = :review_reason,
                    metadata_json = :metadata_json,
                    last_scan_at = :last_scan_at
                WHERE recruiter_id = :id
            """)
            for i in range(0, len(auto_updates), args.batch_size):
                session.execute(update_sql, auto_updates[i:i + args.batch_size])
                session.commit()

        if args.apply and review_updates:
            update_sql = text("""
                UPDATE recruiters
                SET needs_review = :needs_review,
                    recruiter_name = COALESCE(:recruiter_name, recruiter_name),
                    normalized_recruiter_name = COALESCE(:normalized_recruiter_name, normalized_recruiter_name),
                    review_reason = :review_reason,
                    metadata_json = :metadata_json,
                    last_scan_at = :last_scan_at
                WHERE recruiter_id = :id
            """)
            for i in range(0, len(review_updates), args.batch_size):
                session.execute(update_sql, review_updates[i:i + args.batch_size])
                session.commit()

        known_state_after = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NOT NULL AND state != ''")).scalar() or 0
        unknown_state_after = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NULL OR state = ''")).scalar() or 0
        needs_review_after = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE needs_review = true")).scalar() or 0

        state_counts_after_rows = session.execute(text("""
            SELECT state, COUNT(*) AS cnt
            FROM recruiters
            WHERE state IS NOT NULL AND state != ''
            GROUP BY state
            ORDER BY cnt DESC, state ASC
        """)).mappings().all()

        auto_filled_count = len(auto_updates) if args.apply else sum(1 for item in all_inspected if item["bucket"] == "auto_fill_safe")
        still_review_count = sum(1 for item in all_inspected if item["bucket"] == "needs_manual_review")
        truly_unknown_count = sum(1 for item in all_inspected if item["bucket"] == "truly_unknown")

        state_counts_after = {row["state"]: int(row["cnt"]) for row in state_counts_after_rows}

        report = {
            "source_file": str(SOURCE_FILE),
            "dry_run": not args.apply,
            "total_inspected": len(targets),
            "known_state_before": int(known_state_before),
            "known_state_after": int(known_state_after),
            "unknown_state_before": int(unknown_state_before),
            "unknown_state_after": int(unknown_state_after),
            "needs_review_before": int(needs_review_before),
            "needs_review_after": int(needs_review_after),
            "auto_filled_count": int(auto_filled_count),
            "auto_state_fills": int(auto_state_fills),
            "auto_review_cleanups": int(auto_review_cleanups),
            "auto_conflicts": int(auto_conflicts),
            "still_review_count": int(still_review_count),
            "truly_unknown_count": int(truly_unknown_count),
            "bucket_counts": dict(bucket_counts),
            "source_counts": dict(source_counts),
            "state_counts": state_counts_after,
            "auto_filled_samples": auto_filled_samples,
            "still_review_samples": still_review_samples,
            "truly_unknown_samples": truly_unknown_samples,
            "review_queue_json": str(load_review_queue_path()) if load_review_queue_path() else None,
            "workbook_people_parsed": len(people),
            "workbook_company_summary_count": len(workbook_indices[5]),
            "db_company_summary_count": len(db_company_summary),
            "state_counts_before": {row["state"]: int(row["cnt"]) for row in state_counts_before},
            "source_breakdown": {
                "auto_fill_safe": sum(1 for item in all_inspected if item["bucket"] == "auto_fill_safe"),
                "needs_manual_review": sum(1 for item in all_inspected if item["bucket"] == "needs_manual_review"),
                "truly_unknown": sum(1 for item in all_inspected if item["bucket"] == "truly_unknown"),
            },
        }

        report_name = "attack_workbook_review_queue_dry_run" if not args.apply else "attack_workbook_review_queue_report"
        json_path = OUTPUT_DIR / f"{report_name}.json"
        md_path = OUTPUT_DIR / f"{report_name}.md"
        write_report(report, json_path, md_path)
        print(json.dumps(report, indent=2, default=str))

    finally:
        session.close()


if __name__ == "__main__":
    main()
