from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models.models import Company, Recruiter
from app.utils.normalizer import normalize_text
from app.utils.state_mapper import extract_state_detailed


CONTACT_WORKBOOK = Path(
    r"C:\Users\User\Desktop\for location by claude\1 the below all compny but location wise\Recruiter_Contacts_Master.xlsx"
)
STATE_WORKBOOK = Path(r"C:\Users\User\Downloads\Companies data state wise (2).xlsx")
GENERIC_LOCATION_VALUES = {"united states", "usa", "us", "u.s.", "u.s.a.", "country", "remote"}
BATCH_KEY = f"local_workbook_location_backfill_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"


def norm_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def norm_key(value: str | None) -> str:
    return normalize_text(value or "")


def norm_email(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip().lower()
    return value if "@" in value else None


def norm_phone(value: str | None) -> str | None:
    digits = re.sub(r"[^0-9]", "", value or "")
    return digits[-10:] if len(digits) >= 10 else None


def parse_state(*values: str | None) -> str | None:
    for value in values:
        if not value:
            continue
        state, _ = extract_state_detailed(value)
        if state:
            return state
    return None


def clean_location(value: str | None) -> tuple[str, str] | None:
    if not value:
        return None
    value = norm_text(value)
    if not value:
        return None
    lower = value.lower()
    if lower in {"n/a", "na", "nil", "-", "#error!"}:
        return None
    if lower in GENERIC_LOCATION_VALUES:
        return None
    if "@" in value or "www." in lower or "http" in lower:
        return None
    if re.search(r"(^|[^A-Za-z])[0-9]{3}[-\.\s]?[0-9]{3}[-\.\s]?[0-9]{4}([^0-9]|$)", value):
        return None
    if not re.search(r"[A-Za-z]", value):
        return None
    state = parse_state(value)
    if not state:
        return None
    return value, state


def merge_metadata(existing_value: str | None, evidence: dict) -> str:
    metadata = {}
    if existing_value:
        try:
            parsed = json.loads(existing_value) if isinstance(existing_value, str) else existing_value
            if isinstance(parsed, dict):
                metadata = dict(parsed)
        except Exception:
            metadata = {"raw_metadata": str(existing_value)}
    metadata["local_workbook_location_backfill"] = evidence
    return json.dumps(metadata, default=str)


def build_indexes(session):
    company_by_id = {cid: cname for cid, cname in session.query(Company.company_id, Company.company_name).all()}
    email_index = {}
    phone_index = {}
    name_company_index = defaultdict(list)
    recruiters_by_id = {}
    for recruiter in session.query(Recruiter).all():
        recruiters_by_id[recruiter.recruiter_id] = recruiter
        if recruiter.email:
            email_index[recruiter.email.strip().lower()] = recruiter
        phone = norm_phone(recruiter.phone)
        if phone:
            phone_index[phone] = recruiter
        company_name = company_by_id.get(recruiter.company_id, "")
        if recruiter.recruiter_name and company_name:
            key = f"{norm_key(recruiter.recruiter_name)}::{norm_key(company_name)}"
            name_company_index[key].append(recruiter)
    return email_index, phone_index, name_company_index, recruiters_by_id


def match_recruiter(profile: dict, email_index, phone_index, name_company_index):
    if profile.get("email") and profile["email"] in email_index:
        return email_index[profile["email"]]
    if profile.get("phone") and profile["phone"] in phone_index:
        return phone_index[profile["phone"]]
    if profile.get("name") and profile.get("company"):
        key = f"{norm_key(profile['name'])}::{norm_key(profile['company'])}"
        matches = name_company_index.get(key, [])
        if len(matches) == 1:
            return matches[0]
    return None


def collect_contact_workbook_candidates(candidates, email_index, phone_index, name_company_index):
    workbook = load_workbook(CONTACT_WORKBOOK, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [norm_text(str(cell)) if cell is not None else "" for cell in rows[0]]
        header_map = {header.lower(): idx for idx, header in enumerate(headers) if header}
        for row in rows[1:]:
            values = [norm_text(str(cell)) if cell is not None else "" for cell in row]

            def cell(*names):
                for name in names:
                    idx = header_map.get(name)
                    if idx is not None and idx < len(values):
                        return values[idx]
                return ""

            profile = {
                "name": cell("name"),
                "email": norm_email(cell("email", "email address", "email id")),
                "phone": norm_phone(cell("phone", "mobile", "cell")),
                "company": cell("company", "organization", "org") or (values[2] if len(values) > 2 else ""),
            }
            location = (
                clean_location(cell("hq location"))
                or clean_location(cell("office location"))
                or clean_location(cell("location"))
            )
            if not location:
                continue

            recruiter = match_recruiter(profile, email_index, phone_index, name_company_index)
            if recruiter and (not recruiter.location or not recruiter.location.strip()):
                candidates[recruiter.recruiter_id].append(
                    {
                        "source": "contact_workbook",
                        "location": location[0],
                        "state": location[1],
                    }
                )


def collect_state_workbook_candidates(candidates, email_index, phone_index, name_company_index):
    workbook = load_workbook(STATE_WORKBOOK, read_only=True, data_only=True)
    header_hints = {
        "name",
        "email",
        "email address",
        "email id",
        "company",
        "organization",
        "org",
        "state",
        "location",
        "location ",
        "office location",
        "hq location",
        "country",
        "phone",
        "mobile",
        "cell",
    }
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        first_row = [norm_text(str(cell)) if cell is not None else "" for cell in rows[0]]
        headerish = any(value.lower() in header_hints for value in first_row)
        header_map = None
        start_row = 2 if headerish else 1
        if headerish:
            header_map = {norm_key(header): idx for idx, header in enumerate(first_row) if header}

        for row in rows[start_row - 1 :]:
            values = [norm_text(str(cell)) if cell is not None else "" for cell in row]

            def cell(*names, default=""):
                if header_map:
                    for name in names:
                        idx = header_map.get(norm_key(name))
                        if idx is not None and idx < len(values):
                            return values[idx]
                return default

            profile = {
                "name": cell("name") if header_map else (values[0] if len(values) > 0 else ""),
                "email": norm_email(cell("email", "email address", "email id")) if header_map else norm_email(values[1] if len(values) > 1 else ""),
                "phone": norm_phone(cell("phone", "mobile", "cell")) if header_map else None,
                "company": cell("company", "organization", "org") if header_map else (values[2] if len(values) > 2 else sheet.title),
            }
            location = clean_location(cell("location", "location ", "office location", "hq location")) or clean_location(cell("state"))
            if not location:
                continue

            recruiter = match_recruiter(profile, email_index, phone_index, name_company_index)
            if recruiter and (not recruiter.location or not recruiter.location.strip()):
                candidates[recruiter.recruiter_id].append(
                    {
                        "source": "state_workbook",
                        "location": location[0],
                        "state": location[1],
                    }
                )


def main():
    if not CONTACT_WORKBOOK.exists():
        raise FileNotFoundError(f"Missing workbook: {CONTACT_WORKBOOK}")
    if not STATE_WORKBOOK.exists():
        raise FileNotFoundError(f"Missing workbook: {STATE_WORKBOOK}")

    session = SessionLocal()
    try:
        email_index, phone_index, name_company_index, recruiters_by_id = build_indexes(session)
        candidates = defaultdict(list)

        collect_contact_workbook_candidates(candidates, email_index, phone_index, name_company_index)
        collect_state_workbook_candidates(candidates, email_index, phone_index, name_company_index)

        agreed = []
        conflicts = []
        for recruiter_id, sources in candidates.items():
            unique_values = {(item["location"], item["state"]) for item in sources}
            if len(unique_values) == 1:
                location, state = next(iter(unique_values))
                agreed.append((recruiter_id, location, state, sources))
            else:
                conflicts.append((recruiter_id, sources))

        updated = []
        for recruiter_id, location, state, sources in agreed:
            recruiter = recruiters_by_id.get(recruiter_id)
            if not recruiter or (recruiter.location and recruiter.location.strip()):
                continue
            recruiter.location = location[:255]
            recruiter.location_confidence = "high"
            recruiter.state = state
            recruiter.state_source = "local_workbook_location_backfill"
            recruiter.state_confidence = "high"
            recruiter.state_reason = "Agreed clean location from local workbook sources"
            recruiter.metadata_json = merge_metadata(
                recruiter.metadata_json,
                {
                    "batch_key": BATCH_KEY,
                    "sources": sources,
                },
            )
            updated.append((recruiter.recruiter_id, recruiter.recruiter_name, recruiter.location, recruiter.state))

        session.commit()

        print(f"candidate_recruiters={len(candidates)}")
        print(f"agreed_updates={len(updated)}")
        print(f"conflicts_skipped={len(conflicts)}")
        for recruiter_id, recruiter_name, location, state in updated[:50]:
            print(f"{recruiter_id} | {recruiter_name} | {location} | {state}")

    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


if __name__ == "__main__":
    main()
