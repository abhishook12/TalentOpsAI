import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))

from app.database import SessionLocal
from app.models.models import Company, Recruiter


SOURCE_FILE = Path(r"C:\Users\User\Desktop\final updated sheet.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"

STATE_CODES = {
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA",
    "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT",
    "VA", "WA", "WV", "WI", "WY", "DC",
}

COMPANY_KEYWORDS = [
    "llc", "inc", "group", "consulting", "solutions", "technology", "technologies",
    "services", "partners", "staffing", "resources", "search", "talent",
]

TITLE_KEYWORDS = [
    "recruiter", "manager", "sourcer", "talent", "director", "executive",
    "specialist", "head", "lead", "partner",
]

GENERIC_DOMAINS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com", "icloud.com",
    "proton.me", "protonmail.com", "mail.com", "msn.com", "live.com", "yandex.com", "gmx.com",
}


def norm_company(value):
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def norm_phone(value):
    digits = re.sub(r"[^0-9]", "", value or "")
    return digits[-10:] if len(digits) >= 10 else None


def parse_block(block):
    person = {
        "emails": [],
        "phones": [],
        "name": "",
        "company": "",
        "title": "",
        "state": "",
        "location": "",
        "notes": [],
        "sheet_names": set(),
    }
    for _, value in block:
        value = str(value).strip()
        if not value or value.lower() == "nan":
            continue
        if re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", value):
            person["emails"].append(value.lower())
            continue
        phone_match = re.search(r"\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}", value)
        if phone_match:
            digits = re.sub(r"[^0-9]", "", phone_match.group(0))
            if len(digits) >= 10:
                person["phones"].append(digits[-10:])
            continue
        if len(value) == 2 and value.upper() in STATE_CODES:
            person["state"] = value.upper()
            continue
        if "," in value and len(value.split(",")[-1].strip()) == 2 and value.split(",")[-1].strip().upper() in STATE_CODES:
            person["location"] = value.title()
            person["state"] = value.split(",")[-1].strip().upper()
            continue
        if any(keyword in value.lower() for keyword in COMPANY_KEYWORDS):
            if not person["company"]:
                person["company"] = value.title()
            else:
                person["notes"].append(value)
            continue
        if any(keyword in value.lower() for keyword in TITLE_KEYWORDS):
            person["title"] = value.title()
            continue
        if len(value.split()) in [2, 3] and not re.search(r"\d", value) and len(value) < 30 and not person["name"]:
            person["name"] = value.title()
            continue
        person["notes"].append(value)

    if not person["name"] and person["notes"]:
        first_note = person["notes"][0]
        if len(first_note.split()) in [2, 3] and len(first_note) < 30:
            person["name"] = first_note.title()
            person["notes"].pop(0)

    if not person["company"] and person["emails"]:
        domain = person["emails"][0].split("@")[-1]
        if domain not in GENERIC_DOMAINS:
            person["company"] = domain.split(".")[0].title()

    return person


def person_key(person):
    if person["emails"]:
        return f"email::{person['emails'][0].lower()}"
    if person["phones"]:
        return f"phone::{person['phones'][0]}"
    if person["name"] and person["company"]:
        return f"name_company::{person['name'].lower()}::{norm_company(person['company'])}"
    if person["name"]:
        return f"name::{person['name'].lower()}"
    return None


def merge_person(existing, incoming, sheet_name):
    existing["emails"] = list(dict.fromkeys(existing["emails"] + incoming["emails"]))
    existing["phones"] = list(dict.fromkeys(existing["phones"] + incoming["phones"]))
    existing["sheet_names"].update(incoming.get("sheet_names", set()))
    existing["sheet_names"].add(sheet_name)
    if not existing["name"] and incoming["name"]:
        existing["name"] = incoming["name"]
    if not existing["company"] and incoming["company"]:
        existing["company"] = incoming["company"]
    if not existing["title"] and incoming["title"]:
        existing["title"] = incoming["title"]
    if not existing["state"] and incoming["state"]:
        existing["state"] = incoming["state"]
    if not existing["location"] and incoming["location"]:
        existing["location"] = incoming["location"]
    if incoming["notes"]:
        existing["notes"].extend(incoming["notes"])
    return existing


def load_workbook_people(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    people_by_key = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row = [str(cell).strip() if cell is not None else "" for cell in row]
            blocks = []
            current_block = []
            for index, cell in enumerate(row):
                if not cell or cell.lower() in {"nan", "none"}:
                    continue
                if not current_block or index - current_block[-1][0] <= 5:
                    current_block.append((index, cell))
                else:
                    blocks.append(current_block)
                    current_block = [(index, cell)]
            if current_block:
                blocks.append(current_block)

            for block in blocks:
                text = " ".join(cell.lower() for _, cell in block)
                if "contact #1" in text and "email id" in text and "@" not in text:
                    continue
                person = parse_block(block)
                if not (person["emails"] or person["phones"] or (person["name"] and person["company"])):
                    continue
                person["sheet_names"].add(sheet_name)
                key = person_key(person)
                if not key:
                    key = f"sheet::{sheet_name}::row::{len(people_by_key)}"
                if key in people_by_key:
                    people_by_key[key] = merge_person(people_by_key[key], person, sheet_name)
                else:
                    people_by_key[key] = person

    people = list(people_by_key.values())
    for person in people:
        person["sheet_names"] = sorted(person.get("sheet_names", set()))
    return people


def build_indices(people):
    by_email = {email: person for person in people for email in person["emails"]}
    by_phone = {}
    by_name_company = {}
    company_states = defaultdict(Counter)
    domain_states = defaultdict(Counter)

    for person in people:
        for phone in person["phones"]:
            by_phone[phone] = person
        if person["name"] and person["company"]:
            by_name_company[f"{person['name'].lower()}::{norm_company(person['company'])}"] = person
        if person["emails"] and person["state"]:
            domain = person["emails"][0].split("@")[-1].lower()
            if domain not in GENERIC_DOMAINS:
                domain_states[domain][person["state"]] += 1
        if person["company"] and person["state"]:
            company_states[norm_company(person["company"])][person["state"]] += 1

    strong_company = {}
    for company_key, counts in company_states.items():
        total = sum(counts.values())
        state, count = counts.most_common(1)[0]
        if total >= 3 and count / total >= 0.8:
            strong_company[company_key] = {"state": state, "total": total, "count": count}

    strong_domain = {}
    for domain, counts in domain_states.items():
        total = sum(counts.values())
        state, count = counts.most_common(1)[0]
        if total >= 3 and count / total >= 0.8:
            strong_domain[domain] = {"state": state, "total": total, "count": count}

    return by_email, by_phone, by_name_company, strong_company, strong_domain


def parse_metadata(value):
    if not value:
        return {}
    if isinstance(value, dict):
        return dict(value)
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {"value": parsed}
    except Exception:
        return {"raw_metadata": str(value)}


def classify_recruiter(recruiter, company_name, indices):
    by_email, by_phone, by_name_company, strong_company, strong_domain = indices

    email = (recruiter.email or "").lower()
    domain = email.split("@")[-1].lower() if "@" in email else None
    phone = norm_phone(recruiter.phone)
    company_key = norm_company(company_name)
    name_company_key = f"{(recruiter.recruiter_name or '').lower()}::{company_key}" if recruiter.recruiter_name and company_name else None

    workbook_person = by_email.get(email) if email else None
    if workbook_person and workbook_person.get("state"):
        return {
            "category": "workbook_exact_email",
            "priority": 1,
            "state": workbook_person["state"],
            "review_reason": "Workbook exact email match already recovered; queued for confirmation only if needed.",
            "evidence": {
                "matched_email": recruiter.email,
                "workbook_state": workbook_person["state"],
                "workbook_company": workbook_person.get("company"),
                "workbook_location": workbook_person.get("location"),
            },
        }

    if company_key in strong_company:
        info = strong_company[company_key]
        return {
            "category": "workbook_company_majority",
            "priority": 2,
            "state": info["state"],
            "review_reason": f"Workbook company-majority match {info['count']}/{info['total']}",
            "evidence": {
                "company_name": company_name,
                "majority_state": info["state"],
                "matching_rows": info["count"],
                "total_rows": info["total"],
            },
        }

    if domain and domain in strong_domain:
        info = strong_domain[domain]
        return {
            "category": "workbook_domain_majority",
            "priority": 3,
            "state": info["state"],
            "review_reason": f"Workbook domain-majority match {info['count']}/{info['total']}",
            "evidence": {
                "email_domain": domain,
                "majority_state": info["state"],
                "matching_rows": info["count"],
                "total_rows": info["total"],
            },
        }

    if phone and phone in by_phone and by_phone[phone].get("state"):
        workbook_person = by_phone[phone]
        return {
            "category": "workbook_phone_exact",
            "priority": 4,
            "state": workbook_person["state"],
            "review_reason": "Workbook phone exact match",
            "evidence": {
                "matched_phone": phone,
                "workbook_state": workbook_person["state"],
                "workbook_company": workbook_person.get("company"),
            },
        }

    if name_company_key and name_company_key in by_name_company and by_name_company[name_company_key].get("state"):
        workbook_person = by_name_company[name_company_key]
        return {
            "category": "workbook_name_company_exact",
            "priority": 5,
            "state": workbook_person["state"],
            "review_reason": "Workbook exact name+company match",
            "evidence": {
                "workbook_state": workbook_person["state"],
                "workbook_company": workbook_person.get("company"),
            },
        }

    if workbook_person:
        if workbook_person.get("emails"):
            return {
                "category": "workbook_email_no_state",
                "priority": 10,
                "state": None,
                "review_reason": "Workbook email matched but no explicit state present",
                "evidence": {
                    "matched_email": recruiter.email,
                    "workbook_company": workbook_person.get("company"),
                    "workbook_location": workbook_person.get("location"),
                },
            }
        if workbook_person.get("phones"):
            return {
                "category": "workbook_phone_no_state",
                "priority": 11,
                "state": None,
                "review_reason": "Workbook phone matched but no explicit state present",
                "evidence": {
                    "matched_email": recruiter.email,
                    "workbook_company": workbook_person.get("company"),
                    "workbook_location": workbook_person.get("location"),
                },
            }
        return {
            "category": "workbook_match_no_state",
            "priority": 12,
            "state": None,
            "review_reason": "Workbook match found but no state could be extracted",
            "evidence": {
                "matched_email": recruiter.email,
                "workbook_company": workbook_person.get("company"),
                "workbook_location": workbook_person.get("location"),
            },
        }

    if recruiter.email:
        if "@" in recruiter.email and recruiter.email.split("@")[-1].lower() in GENERIC_DOMAINS:
            return {
                "category": "generic_email_domain",
                "priority": 20,
                "state": None,
                "review_reason": "Generic email domain provides no useful state signal",
                "evidence": {"email_domain": recruiter.email.split("@")[-1].lower()},
            }
        return {
            "category": "no_workbook_match",
            "priority": 21,
            "state": None,
            "review_reason": "No safe workbook match",
            "evidence": {"email": recruiter.email},
        }

    return {
        "category": "no_contact_evidence",
        "priority": 22,
        "state": None,
        "review_reason": "No contact evidence in workbook",
        "evidence": {
            "phone": recruiter.phone,
            "company_name": company_name,
        },
    }


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    people = load_workbook_people(SOURCE_FILE)
    indices = build_indices(people)

    session = SessionLocal()
    try:
        companies = {company_id: company_name for company_id, company_name in session.query(Company.company_id, Company.company_name).all()}
        unknown = session.query(Recruiter).filter((Recruiter.state.is_(None)) | (Recruiter.state == "")).all()

        queue = []
        updates = []
        category_counts = Counter()
        source_counts = Counter()

        for recruiter in unknown:
            company_name = companies.get(recruiter.company_id, "") if recruiter.company_id else ""
            result = classify_recruiter(recruiter, company_name, indices)
            category_counts[result["category"]] += 1

            queue_entry = {
                "recruiter_id": recruiter.recruiter_id,
                "recruiter_name": recruiter.recruiter_name,
                "email": recruiter.email,
                "phone": recruiter.phone,
                "company_name": company_name,
                "category": result["category"],
                "priority": result["priority"],
                "proposed_state": result["state"],
                "review_reason": result["review_reason"],
                "evidence": result["evidence"],
            }
            queue.append(queue_entry)

            if result["state"] is None:
                meta = parse_metadata(recruiter.metadata_json)
                meta["workbook_review_queue"] = {
                    "category": result["category"],
                    "priority": result["priority"],
                    "source_file": str(SOURCE_FILE),
                    "review_reason": result["review_reason"],
                    "evidence": result["evidence"],
                    "queued_at": datetime.utcnow().isoformat(),
                }
                updates.append(
                    {
                        "id": recruiter.recruiter_id,
                        "needs_review": True,
                        "review_reason": result["review_reason"][:500],
                        "metadata_json": json.dumps(meta, default=str),
                    }
                )
                source_counts[result["category"]] += 1

        if updates:
            update_sql = text("""
                UPDATE recruiters
                SET needs_review = :needs_review,
                    review_reason = :review_reason,
                    metadata_json = :metadata_json
                WHERE recruiter_id = :id
            """)
            for i in range(0, len(updates), 500):
                batch = updates[i:i + 500]
                session.execute(
                    update_sql,
                    batch,
                )
                session.commit()

        queue.sort(key=lambda row: (row["priority"], row["category"], row["recruiter_name"] or ""))

        csv_path = OUTPUT_DIR / "workbook_review_queue.csv"
        json_path = OUTPUT_DIR / "workbook_review_queue.json"
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "recruiter_id",
                    "recruiter_name",
                    "email",
                    "phone",
                    "company_name",
                    "category",
                    "priority",
                    "proposed_state",
                    "review_reason",
                    "evidence",
                ],
            )
            writer.writeheader()
            for row in queue:
                serializable = dict(row)
                serializable["evidence"] = json.dumps(serializable["evidence"], default=str)
                writer.writerow(serializable)

        with json_path.open("w", encoding="utf-8") as handle:
            json.dump(queue, handle, indent=2, default=str)

        report = {
            "source_file": str(SOURCE_FILE),
            "parsed_people": len(people),
            "unknown_recruiters": len(unknown),
            "queued_for_review": len(queue),
            "queued_updates_applied": len(updates),
            "category_counts": dict(category_counts),
            "source_counts": dict(source_counts),
            "csv_path": str(csv_path),
            "json_path": str(json_path),
        }
        print(json.dumps(report, indent=2))

    finally:
        session.close()


if __name__ == "__main__":
    main()
