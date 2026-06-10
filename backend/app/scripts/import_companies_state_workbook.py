from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))

from app.database import SessionLocal
from app.models.models import Company, Recruiter
from app.utils.normalizer import normalize_text
from app.utils.state_mapper import extract_state_detailed, normalize_state


DEFAULT_SOURCE_FILE = Path(r"C:\Users\User\Downloads\Companies data state wise (2).xlsx")
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
BATCH_ID = f"csw_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"

GENERIC_DOMAINS = {
    "gmail.com",
    "yahoo.com",
    "hotmail.com",
    "outlook.com",
    "aol.com",
    "icloud.com",
    "proton.me",
    "protonmail.com",
    "mail.com",
    "msn.com",
    "live.com",
    "yandex.com",
    "gmx.com",
}

HEADER_HINTS = {
    "name",
    "email",
    "email address",
    "email id",
    "company",
    "organization",
    "org",
    "state",
    "location",
    "location ",
    "office location",
    "hq location",
    "country",
    "phone",
    "mobile",
    "cell",
}


def norm_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def norm_key(value: str | None) -> str:
    return normalize_text(value or "")


def norm_email(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip().lower()
    return value if "@" in value else None


def norm_phone(value: str | None) -> str | None:
    digits = re.sub(r"[^0-9]", "", value or "")
    return digits[-10:] if len(digits) >= 10 else None


def has_header(row_values: list[str]) -> bool:
    normalized = {norm_key(value) for value in row_values if value}
    return bool(normalized & HEADER_HINTS)


def parse_state(*values: str | None):
    for value in values:
        if not value:
            continue
        state, reason = extract_state_detailed(value)
        if state:
            return state, reason
    return None, None


def is_generic_domain(email: str | None) -> bool:
    if not email or "@" not in email:
        return False
    return email.split("@", 1)[1].lower() in GENERIC_DOMAINS


def is_plausible_company(value: str | None) -> bool:
    if not value:
        return False
    value = value.strip()
    if not value:
        return False
    lower_value = value.lower()
    if "@" in lower_value or ".com" in lower_value or ".net" in lower_value or ".org" in lower_value:
        return False
    if re.search(r"\d", value):
        return False
    if len(value) > 80:
        return False
    return True


def stable_synthetic_email(identity: str) -> str:
    digest = hashlib.sha1(identity.encode("utf-8")).hexdigest()[:16]
    return f"no-email-{digest}@missing.local"


def merge_metadata(existing_value, evidence: dict, key: str):
    metadata = {}
    if existing_value:
        try:
            parsed = json.loads(existing_value) if isinstance(existing_value, str) else existing_value
            if isinstance(parsed, dict):
                metadata = dict(parsed)
        except Exception:
            metadata = {"raw_metadata": str(existing_value)}
    metadata[key] = evidence
    return json.dumps(metadata, default=str)


def row_to_profile(row_values: list[str], header_map: dict[str, int] | None, sheet_title: str, row_number: int):
    values = [norm_text(value) for value in row_values]

    def cell(*names, default=""):
        if header_map:
            for name in names:
                idx = header_map.get(norm_key(name))
                if idx is not None and idx < len(values):
                    return values[idx]
        return default

    if header_map:
        name = cell("name")
        email = norm_email(cell("email", "email address", "email id"))
        phone = norm_phone(cell("phone", "mobile", "cell"))
        company_raw = cell("company", "organization", "org")
        state = cell("state")
        location = cell("location", "location ", "office location", "hq location")
        country = cell("country")
    else:
        name = values[0] if len(values) > 0 else ""
        email = norm_email(values[1] if len(values) > 1 else "")
        company_raw = values[2] if len(values) > 2 else sheet_title
        phone = None
        state = None
        location = None
        country = None

    company = company_raw if is_plausible_company(company_raw) else None

    state_from_value, state_reason = parse_state(state, location, country)
    location_text = location or state or country or ""
    qualifies = bool(state_from_value or phone or location_text)
    identity_parts = [
        f"sheet={sheet_title}",
        f"row={row_number}",
        f"name={name}",
        f"email={email or ''}",
        f"phone={phone or ''}",
        f"company={company or ''}",
    ]
    identity = "|".join(identity_parts)

    return {
        "sheet": sheet_title,
        "row_number": row_number,
        "name": name,
        "email": email,
        "phone": phone,
        "company": company,
        "state": state_from_value,
        "state_reason": state_reason,
        "location": location_text or None,
        "country": country,
        "qualifies": qualifies,
        "identity": identity,
        "raw_values": values,
    }


def parse_workbook(source_file: Path):
    workbook = load_workbook(source_file, read_only=True, data_only=True)
    source_row_count = 0
    sheet_summaries = {}
    profiles: dict[str, dict] = {}
    sheet_rows: dict[str, list[dict]] = defaultdict(list)

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        first_row = [norm_text(str(cell)) if cell is not None else "" for cell in rows[0]]
        header_row = has_header(first_row)
        header_map = None
        start_row = 2 if header_row else 1
        if header_row:
            header_map = {
                norm_key(header): idx
                for idx, header in enumerate(first_row)
                if header
            }

        sheet_company_counter = Counter()
        sheet_state_counter = Counter()
        sheet_location_rows = 0
        sheet_qualifying_rows = 0

        for row_number, row in enumerate(rows[start_row - 1 :], start=start_row):
            source_row_count += 1
            row_values = ["" if cell is None else str(cell) for cell in row]
            if not any(norm_text(value) for value in row_values):
                continue

            profile = row_to_profile(row_values, header_map, sheet.title, row_number)
            sheet_rows[sheet.title].append(profile)

            if profile["company"]:
                sheet_company_counter[norm_key(profile["company"])] += 1
            if profile["state"]:
                sheet_state_counter[profile["state"]] += 1
            if profile["location"]:
                sheet_location_rows += 1
            if profile["qualifies"]:
                sheet_qualifying_rows += 1

            if not profile["identity"]:
                continue

            existing = profiles.setdefault(
                profile["identity"],
                {
                    "identity": profile["identity"],
                    "sheet_names": set(),
                    "names": Counter(),
                    "emails": Counter(),
                    "phones": Counter(),
                    "companies": Counter(),
                    "states": Counter(),
                    "locations": [],
                    "rows": [],
                    "qualifies": False,
                },
            )
            existing["sheet_names"].add(sheet.title)
            if profile["name"]:
                existing["names"][profile["name"]] += 1
            if profile["email"]:
                existing["emails"][profile["email"]] += 1
            if profile["phone"]:
                existing["phones"][profile["phone"]] += 1
            if profile["company"]:
                existing["companies"][profile["company"]] += 1
            if profile["state"]:
                existing["states"][profile["state"]] += 1
            if profile["location"]:
                existing["locations"].append(profile["location"])
            existing["qualifies"] = existing["qualifies"] or profile["qualifies"]
            existing["rows"].append(profile)

        dominant_company = None
        dominant_company_count = 0
        if sheet_company_counter:
            dominant_company, dominant_company_count = sheet_company_counter.most_common(1)[0]
            dominant_company = None if dominant_company_count < 1 else dominant_company

        dominant_state = None
        dominant_state_count = 0
        if sheet_state_counter:
            dominant_state, dominant_state_count = sheet_state_counter.most_common(1)[0]

        sheet_summaries[sheet.title] = {
            "sheet_name": sheet.title,
            "header_row": header_row,
            "row_count": max(len(rows) - (1 if header_row else 0), 0),
            "qualifying_rows": sheet_qualifying_rows,
            "location_rows": sheet_location_rows,
            "company_counter": dict(sheet_company_counter),
            "state_counter": dict(sheet_state_counter),
            "dominant_company": dominant_company,
            "dominant_company_count": dominant_company_count,
            "dominant_state": dominant_state,
            "dominant_state_count": dominant_state_count,
            "sample_rows": sheet_rows[sheet.title][:5],
        }

    for profile in profiles.values():
        profile["sheet_names"] = sorted(profile["sheet_names"])
        profile["primary_name"], profile["primary_name_count"] = (profile["names"].most_common(1)[0] if profile["names"] else (None, 0))
        profile["primary_email"], profile["primary_email_count"] = (profile["emails"].most_common(1)[0] if profile["emails"] else (None, 0))
        profile["primary_phone"], profile["primary_phone_count"] = (profile["phones"].most_common(1)[0] if profile["phones"] else (None, 0))
        profile["primary_company"], profile["primary_company_count"] = (profile["companies"].most_common(1)[0] if profile["companies"] else (None, 0))
        profile["primary_state"], profile["primary_state_count"] = (profile["states"].most_common(1)[0] if profile["states"] else (None, 0))
        profile["location_text"] = profile["locations"][0] if profile["locations"] else None
        profile["row_count"] = len(profile["rows"])

    return list(profiles.values()), sheet_summaries, source_row_count


def build_db_indexes(session):
    companies = session.query(Company).all()
    recruiters = session.query(Recruiter).all()

    company_by_norm = {}
    for company in companies:
        if company.company_name:
            company_by_norm[norm_key(company.company_name)] = company

    recruiter_by_email = {}
    recruiter_by_phone = {}
    recruiter_by_name_company = defaultdict(list)
    recruiter_by_name_email = defaultdict(list)

    for recruiter in recruiters:
        if recruiter.email:
            recruiter_by_email[recruiter.email.strip().lower()] = recruiter
        phone = norm_phone(recruiter.phone)
        if phone:
            recruiter_by_phone[phone] = recruiter

        company_name = None
        if recruiter.company_id:
            company = next((row for row in companies if row.company_id == recruiter.company_id), None)
            if company and company.company_name:
                company_name = company.company_name
        if recruiter.recruiter_name and company_name:
            recruiter_by_name_company[f"{norm_key(recruiter.recruiter_name)}::{norm_key(company_name)}"].append(recruiter)
        if recruiter.recruiter_name and recruiter.email:
            recruiter_by_name_email[f"{norm_key(recruiter.recruiter_name)}::{recruiter.email.strip().lower()}"].append(recruiter)

    return {
        "companies": companies,
        "company_by_norm": company_by_norm,
        "recruiters": recruiters,
        "recruiter_by_email": recruiter_by_email,
        "recruiter_by_phone": recruiter_by_phone,
        "recruiter_by_name_company": recruiter_by_name_company,
        "recruiter_by_name_email": recruiter_by_name_email,
    }


def choose_company_name(profile: dict, sheet_summary: dict) -> str | None:
    if profile.get("primary_company") and profile["primary_company_count"] >= 1:
        return profile["primary_company"]
    if sheet_summary.get("dominant_company"):
        return sheet_summary["dominant_company"]
    return None


def choose_state(profile: dict, sheet_summary: dict) -> tuple[str | None, str | None, str | None]:
    state = profile.get("primary_state")
    reason = None
    confidence = None

    if state:
        confidence = "high" if profile.get("primary_state_count", 0) >= 2 else "medium"
        reason = f"profile_state_majority:{profile.get('primary_state_count', 0)}"
    elif sheet_summary.get("dominant_state"):
        state = sheet_summary["dominant_state"]
        confidence = "high" if sheet_summary.get("dominant_state_count", 0) >= 2 else "medium"
        reason = f"sheet_state_majority:{sheet_summary.get('dominant_state_count', 0)}"

    return state, confidence, reason


def is_generic_or_missing_email(email: str | None) -> bool:
    return not email or is_generic_domain(email)


def match_existing_recruiter(indexes, profile: dict, company_name: str | None):
    email = profile.get("primary_email")
    phone = profile.get("primary_phone")
    recruiter = None
    if email and email.lower() in indexes["recruiter_by_email"]:
        recruiter = indexes["recruiter_by_email"][email.lower()]
        match_type = "email"
    elif phone and phone in indexes["recruiter_by_phone"]:
        recruiter = indexes["recruiter_by_phone"][phone]
        match_type = "phone"
    elif profile.get("primary_name") and company_name:
        key = f"{norm_key(profile['primary_name'])}::{norm_key(company_name)}"
        if key in indexes["recruiter_by_name_company"]:
            recruiter = indexes["recruiter_by_name_company"][key][0]
            match_type = "name_company"
        else:
            match_type = None
    elif profile.get("primary_name") and email:
        key = f"{norm_key(profile['primary_name'])}::{email.lower()}"
        if key in indexes["recruiter_by_name_email"]:
            recruiter = indexes["recruiter_by_name_email"][key][0]
            match_type = "name_email"
        else:
            match_type = None
    else:
        match_type = None
    return recruiter, match_type


def safe_update_recruiter(recruiter: Recruiter, profile: dict, company: Company | None, state: str | None, confidence: str | None, state_reason: str | None, source_sheet: str):
    changed = False
    evidence = {
        "source_sheet": source_sheet,
        "source_row_count": profile.get("row_count"),
        "sheet_names": profile.get("sheet_names"),
        "name": profile.get("primary_name"),
        "email": profile.get("primary_email"),
        "phone": profile.get("primary_phone"),
        "company": profile.get("primary_company"),
        "state": state,
        "state_reason": state_reason,
        "location_text": profile.get("location_text"),
    }

    if profile.get("primary_name") and (not recruiter.recruiter_name or recruiter.recruiter_name == "Unknown"):
        recruiter.recruiter_name = profile["primary_name"][:150]
        changed = True

    if profile.get("primary_phone"):
        if not recruiter.phone:
            recruiter.phone = profile["primary_phone"]
            changed = True

    if profile.get("location_text") and not recruiter.location:
        recruiter.location = profile["location_text"][:255]
        changed = True

    if state and not recruiter.state:
        recruiter.state = state
        recruiter.state_source = "companies_state_workbook"
        recruiter.state_confidence = confidence
        recruiter.state_reason = state_reason
        changed = True
    elif state and recruiter.state and recruiter.state != state:
        evidence["existing_state"] = recruiter.state

    if company and not recruiter.company_id:
        recruiter.company_id = company.company_id
        changed = True

    if recruiter.needs_review and recruiter.state:
        recruiter.needs_review = False
        changed = True

    if recruiter.review_reason and recruiter.state:
        recruiter.review_reason = None
        changed = True

    recruiter.metadata_json = merge_metadata(recruiter.metadata_json, evidence, "companies_state_workbook_evidence")
    recruiter.data_source = recruiter.data_source or "companies_state_workbook"
    recruiter.source_job_id = recruiter.source_job_id or BATCH_ID
    recruiter.last_scan_at = datetime.now(timezone.utc).replace(tzinfo=None)
    return changed


def safe_update_company(company: Company, dominant_state: str | None, evidence: dict):
    changed = False
    if dominant_state and not company.state:
        company.state = dominant_state
        changed = True
    company.metadata_json = merge_metadata(company.metadata_json, evidence, "companies_state_workbook_evidence")
    company.data_source = company.data_source or "companies_state_workbook"
    company.source_job_id = company.source_job_id or BATCH_ID
    return changed


def create_new_company(company_name: str, state: str | None, sheet_summary: dict):
    return Company(
        company_name=company_name[:255],
        normalized_company_name=norm_key(company_name),
        state=state,
        data_source="companies_state_workbook",
        source_job_id=BATCH_ID,
        is_active=True,
        metadata_json=json.dumps(
            {
                "companies_state_workbook_evidence": {
                    "sheet_name": sheet_summary["sheet_name"],
                    "dominant_state": state,
                    "dominant_state_count": sheet_summary.get("dominant_state_count", 0),
                    "row_count": sheet_summary.get("row_count", 0),
                }
            },
            default=str,
        ),
    )


def make_report_path(name: str, suffix: str):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR / f"{name}_{BATCH_ID}.{suffix}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_FILE)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--batch-size", type=int, default=250)
    args = parser.parse_args()

    if not args.source.exists():
        raise FileNotFoundError(f"Workbook not found: {args.source}")

    profiles, sheet_summaries, source_row_count = parse_workbook(args.source)
    session = SessionLocal()

    try:
        indexes = build_db_indexes(session)
        company_updates = []
        company_creates = []
        recruiter_updates = []
        recruiter_creates = []
        skipped = []
        conflicts = []
        state_counter = Counter()
        state_by_source = Counter()

        total_recruiters_before = session.execute(text("SELECT COUNT(*) FROM recruiters")).scalar() or 0
        total_companies_before = session.execute(text("SELECT COUNT(*) FROM companies")).scalar() or 0
        known_state_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NOT NULL AND state <> ''")).scalar() or 0
        unknown_state_before = total_recruiters_before - known_state_before

        for profile in profiles:
            first_sheet = profile["sheet_names"][0] if profile["sheet_names"] else None
            sheet_summary = sheet_summaries.get(first_sheet, {}) if first_sheet else {}

            company_name = choose_company_name(profile, sheet_summary)
            if not company_name:
                company_name = profile.get("primary_company")
            if not company_name:
                company_name = sheet_summary.get("dominant_company")

            state, confidence, state_reason = choose_state(profile, sheet_summary)
            if state:
                state_counter[state] += 1
                state_by_source[profile.get("primary_company") or "unknown"] += 1

            qualifies = bool(state or profile.get("primary_phone") or profile.get("location_text"))
            if not qualifies:
                skipped.append(
                    {
                        "identity": profile["identity"],
                        "name": profile.get("primary_name"),
                        "email": profile.get("primary_email"),
                        "company": company_name,
                        "reason": "no_state_no_location_no_phone",
                    }
                )
                continue

            company = indexes["company_by_norm"].get(norm_key(company_name)) if company_name else None
            if company is None and company_name and state:
                sheet_name = profile["sheet_names"][0] if profile["sheet_names"] else "unknown"
                company = create_new_company(company_name, state, sheet_summaries.get(sheet_name, {"sheet_name": sheet_name}))
                session.add(company)
                session.flush()
                indexes["company_by_norm"][norm_key(company_name)] = company
                company_creates.append(
                    {
                        "company_name": company_name,
                        "state": state,
                        "sheet_name": sheet_name,
                    }
                )

            if company and state and not company.state:
                company_updates.append(
                    {
                        "company_name": company.company_name,
                        "company_id": company.company_id,
                        "before_state": company.state,
                        "after_state": state,
                        "sheet_name": first_sheet,
                    }
                )
                safe_update_company(
                    company,
                    state,
                    {
                        "sheet_name": first_sheet,
                        "dominant_state": state,
                        "dominant_state_count": profile.get("primary_state_count", 0),
                        "profile_rows": profile.get("row_count", 0),
                    },
                )

            recruiter, match_type = match_existing_recruiter(indexes, profile, company_name)
            if recruiter:
                before_state = recruiter.state
                changed = safe_update_recruiter(recruiter, profile, company, state, confidence, state_reason, first_sheet or "unknown")
                if changed:
                    recruiter_updates.append(
                        {
                            "recruiter_id": recruiter.recruiter_id,
                            "name": recruiter.recruiter_name,
                            "email": recruiter.email,
                            "match_type": match_type,
                            "before_state": before_state,
                            "after_state": recruiter.state,
                            "company": company_name,
                        }
                    )
            else:
                email = profile.get("primary_email")
                if is_generic_or_missing_email(email) and not profile.get("primary_phone"):
                    email = stable_synthetic_email(profile["identity"])
                elif not email and profile.get("primary_phone"):
                    email = f"no-email-{profile['primary_phone']}@missing.local"
                elif not email:
                    email = stable_synthetic_email(profile["identity"])

                if not email:
                    skipped.append(
                        {
                            "identity": profile["identity"],
                            "name": profile.get("primary_name"),
                            "company": company_name,
                            "reason": "could_not_generate_email",
                        }
                    )
                    continue

                if email.lower() in indexes["recruiter_by_email"]:
                    recruiter = indexes["recruiter_by_email"][email.lower()]
                    changed = safe_update_recruiter(recruiter, profile, company, state, confidence, state_reason, first_sheet or "unknown")
                    if changed:
                        recruiter_updates.append(
                            {
                                "recruiter_id": recruiter.recruiter_id,
                                "name": recruiter.recruiter_name,
                                "email": recruiter.email,
                                "match_type": "email_after_generation",
                                "before_state": None,
                                "after_state": recruiter.state,
                                "company": company_name,
                            }
                        )
                    continue

                recruiter = Recruiter(
                    recruiter_name=(profile.get("primary_name") or "Unknown")[:150],
                    normalized_recruiter_name=norm_key(profile.get("primary_name")),
                    email=email.lower(),
                    phone=profile.get("primary_phone"),
                    company_id=company.company_id if company else None,
                    location=(profile.get("location_text") or None)[:255] if profile.get("location_text") else None,
                    state=state,
                    normalized_city=None,
                    location_confidence="high" if state else "manual_review",
                    state_source="companies_state_workbook" if state else None,
                    state_confidence=confidence if state else None,
                    state_reason=state_reason if state else None,
                    needs_review=False if state else True,
                    data_source="companies_state_workbook",
                    source_job_id=BATCH_ID,
                    metadata_json=json.dumps(
                        {
                            "companies_state_workbook_evidence": {
                                "sheet_name": first_sheet,
                                "sheet_names": profile["sheet_names"],
                                "primary_company": profile.get("primary_company"),
                                "primary_state": state,
                                "state_reason": state_reason,
                                "row_count": profile.get("row_count", 0),
                            }
                        },
                        default=str,
                    ),
                )
                session.add(recruiter)
                recruiter_creates.append(
                    {
                        "name": recruiter.recruiter_name,
                        "email": recruiter.email,
                        "company": company_name,
                        "state": recruiter.state,
                        "sheet_name": first_sheet,
                    }
                )
                indexes["recruiter_by_email"][recruiter.email] = recruiter

        if args.apply:
            session.commit()
        else:
            session.rollback()

        total_recruiters_after = session.execute(text("SELECT COUNT(*) FROM recruiters")).scalar() or 0
        total_companies_after = session.execute(text("SELECT COUNT(*) FROM companies")).scalar() or 0
        known_state_after = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NOT NULL AND state <> ''")).scalar() or 0
        unknown_state_after = total_recruiters_after - known_state_after

        report = {
            "batch_id": BATCH_ID,
            "source_file": str(args.source),
            "source_row_count": source_row_count,
            "sheet_count": len(sheet_summaries),
            "sheet_summaries": sheet_summaries,
            "total_recruiters_before": int(total_recruiters_before),
            "total_recruiters_after": int(total_recruiters_after),
            "total_companies_before": int(total_companies_before),
            "total_companies_after": int(total_companies_after),
            "known_state_before": int(known_state_before),
            "known_state_after": int(known_state_after),
            "unknown_state_before": int(unknown_state_before),
            "unknown_state_after": int(unknown_state_after),
            "recruiters_created": len(recruiter_creates),
            "recruiters_updated": len(recruiter_updates),
            "companies_created": len(company_creates),
            "companies_updated": len(company_updates),
            "skipped": len(skipped),
            "state_counts": dict(state_counter.most_common()),
            "recruiter_create_samples": recruiter_creates[:20],
            "recruiter_update_samples": recruiter_updates[:20],
            "company_create_samples": company_creates[:20],
            "company_update_samples": company_updates[:20],
            "skipped_samples": skipped[:20],
            "conflicts": conflicts[:20],
        }

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        json_path = make_report_path("companies_state_workbook_report", "json")
        md_path = make_report_path("companies_state_workbook_report", "md")
        csv_path = make_report_path("companies_state_workbook_recruiters", "csv")

        json_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
        md_lines = [
            f"# Companies State Workbook Report",
            "",
            f"- Batch ID: `{BATCH_ID}`",
            f"- Source file: `{args.source}`",
            f"- Source rows parsed: `{source_row_count}`",
            f"- Recruiters created: `{len(recruiter_creates)}`",
            f"- Recruiters updated: `{len(recruiter_updates)}`",
            f"- Companies created: `{len(company_creates)}`",
            f"- Companies updated: `{len(company_updates)}`",
            f"- Skipped rows: `{len(skipped)}`",
            f"- Known state before: `{known_state_before}`",
            f"- Known state after: `{known_state_after}`",
            f"- Unknown state before: `{unknown_state_before}`",
            f"- Unknown state after: `{unknown_state_after}`",
            "",
            "## State Counts",
        ]
        for state, count in state_counter.most_common(20):
            md_lines.append(f"- {state}: `{count}`")
        md_lines.extend(
            [
                "",
                "## Recruiter Samples",
            ]
        )
        for sample in recruiter_creates[:20]:
            md_lines.append(f"- `{sample}`")
        md_lines.extend(
            [
                "",
                "## Company Samples",
            ]
        )
        for sample in company_updates[:20]:
            md_lines.append(f"- `{sample}`")
        md_path.write_text("\n".join(md_lines), encoding="utf-8")

        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=["name", "email", "company", "state", "sheet_name"])
            writer.writeheader()
            for sample in recruiter_creates:
                writer.writerow(sample)

        print(json.dumps(report, indent=2, default=str))
        print(f"\nReports written to:\n- {json_path}\n- {md_path}\n- {csv_path}")

    finally:
        session.close()


if __name__ == "__main__":
    main()
