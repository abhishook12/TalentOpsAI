from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

sys.path.append(str(Path(__file__).resolve().parents[2]))

from app.database import SessionLocal
from app.models.models import Company, Recruiter
from app.utils.normalizer import normalize_text
from app.utils.state_mapper import extract_state_detailed


DEFAULT_SOURCE_FILE = Path(r"C:\Users\User\Desktop\for location by claude\1 the below all compny but location wise\Recruiter_Contacts_Master.xlsx")
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
BATCH_ID = f"location_workbook_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"

GENERIC_DOMAINS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com", "icloud.com",
    "proton.me", "protonmail.com", "mail.com", "msn.com", "live.com", "yandex.com", "gmx.com",
}

GENERIC_NAME_TOKENS = {
    "unknown",
    "n/a",
    "na",
    "none",
    "null",
    "-",
    "name",
    "recruiter",
    "contact",
    "person",
}


def norm_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def norm_key(value: str | None) -> str:
    return normalize_text(value or "")


def norm_email(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip().lower()
    return value if "@" in value else None


def norm_phone(value: str | None) -> str | None:
    digits = re.sub(r"[^0-9]", "", value or "")
    return digits[-10:] if len(digits) >= 10 else None


def parse_state(*values: str | None):
    for value in values:
        if not value:
            continue
        state, reason = extract_state_detailed(value)
        if state:
            return state, reason
    return None, None


def profile_identity(name: str | None, email: str | None, phone: str | None, company: str | None):
    if email:
        return f"email::{email}"
    if phone:
        return f"phone::{phone}"
    if name and company:
        return f"name_company::{norm_key(name)}::{norm_key(company)}"
    return None


def name_score(value: str | None) -> tuple[int, int, int]:
    text = norm_text(value)
    if not text:
        return (0, 0, 0)
    lowered = text.lower()
    tokens = [token for token in re.split(r"\s+", text) if token]
    if lowered in GENERIC_NAME_TOKENS or "@" in text or any(char.isdigit() for char in text):
        return (0, len(tokens), len(text))
    return (1 if len(tokens) >= 2 else 0, len(tokens), len(text))


def choose_preferred_name(names: list[str], email: str | None = None) -> str | None:
    cleaned: list[str] = []
    seen = set()
    email_localpart = (email or "").strip().lower().split("@", 1)[0]
    for name in names:
        text = norm_text(name)
        if not text:
            continue
        lowered = text.lower()
        if lowered in GENERIC_NAME_TOKENS or "@" in text:
            continue
        if email_localpart and lowered == email_localpart:
            continue
        if lowered not in seen:
            seen.add(lowered)
            cleaned.append(text)
    if not cleaned:
        return None
    cleaned.sort(key=name_score, reverse=True)
    return cleaned[0]


def build_placeholder_email(profile: dict) -> str:
    email = profile.get("email")
    if email:
        return email
    phone = profile.get("phone")
    if phone:
        return f"no-email-{phone}@missing.local"
    token_source = profile.get("identity_key") or profile.get("preferred_name") or profile.get("name") or json.dumps(profile.get("source_rows", []), default=str)
    token = hashlib.sha1(str(token_source).encode("utf-8")).hexdigest()[:16]
    return f"no-email-{token}@missing.local"


def truncate(value: str | None, limit: int) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text[:limit] if text else None


def should_promote_name(existing: str | None, candidate: str | None, email: str | None = None) -> bool:
    existing_text = norm_text(existing)
    candidate_text = norm_text(candidate)
    if not candidate_text:
        return False
    candidate_lower = candidate_text.lower()
    if candidate_lower in GENERIC_NAME_TOKENS or "@" in candidate_text:
        return False
    if not existing_text:
        return True

    existing_lower = existing_text.lower()
    email_localpart = (email or "").strip().lower().split("@", 1)[0]
    if existing_lower in GENERIC_NAME_TOKENS or (email_localpart and existing_lower == email_localpart):
        return True

    existing_score = name_score(existing_text)
    candidate_score = name_score(candidate_text)
    if candidate_score > existing_score:
        return True

    existing_tokens = len([token for token in re.split(r"\s+", existing_text) if token])
    candidate_tokens = len([token for token in re.split(r"\s+", candidate_text) if token])
    if candidate_tokens >= 2 and existing_tokens <= 1:
        return True
    if len(candidate_text) >= len(existing_text) + 3 and candidate_tokens >= existing_tokens:
        return True
    return False


def parse_workbook(source_file: Path):
    workbook = load_workbook(source_file, read_only=True, data_only=True)
    profiles: dict[str, dict] = {}
    source_row_count = 0
    sheet_counts = Counter()

    def row_text(row):
        return [norm_text(str(cell)) if cell is not None else "" for cell in row]

    def header_score(values: list[str]) -> int:
        haystack = " | ".join(value.lower() for value in values if value).lower()
        score = 0
        for token in ("company", "company name", "pv name", "email", "location", "notes", "phone", "country"):
            if token in haystack:
                score += 1
        return score

    def is_header_row(values: list[str]) -> bool:
        haystack = " | ".join(value.lower() for value in values if value)
        if "email" not in haystack:
            return False
        if not any(token in haystack for token in ("company name", "pv name", "contact name")):
            return False
        return header_score(values) >= 2

    for sheet in workbook.worksheets:
        rows = [row_text(row) for row in sheet.iter_rows(values_only=True)]
        if not rows:
            continue

        header_row_index = None
        header_map: dict[str, int] = {}
        for idx, candidate_row in enumerate(rows[:5]):
            if is_header_row(candidate_row):
                header_row_index = idx
                header_map = {
                    header.lower(): col_idx
                    for col_idx, header in enumerate(candidate_row)
                    if header
                }
                break

        data_start_index = (header_row_index + 1) if header_row_index is not None else 0
        sheet_counts[sheet.title] += max(len(rows) - data_start_index, 0)

        for row_index, values in enumerate(rows[data_start_index:], start=data_start_index + 1):
            source_row_count += 1

            def cell(*names, default=""):
                for name in names:
                    idx = header_map.get(name)
                    if idx is not None and idx < len(values):
                        return values[idx]
                return default

            name = cell("name", "pv name", "contact name") or (values[1] if len(values) > 1 else "")
            email = norm_email(cell("email", "email address", "email id") or (values[2] if len(values) > 2 else ""))
            phone = norm_phone(cell("phone", "mobile", "cell"))
            company = cell("company", "company name", "organization", "org") or (values[0] if len(values) > 0 else "")
            location = cell("location", "office location", "hq location") or (values[3] if len(values) > 3 else "")
            country = cell("country")
            office_location = cell("office location")
            hq_location = cell("hq location")

            state, state_reason = parse_state(location, office_location, hq_location, country)
            identity = profile_identity(name, email, phone, company)
            if not identity:
                continue

            profile = profiles.setdefault(
                identity,
                {
                    "identity_key": identity,
                    "name": name,
                    "names": [],
                    "email": email,
                    "phone": phone,
                    "company": company,
                    "locations": [],
                    "states": Counter(),
                    "sheet_names": set(),
                    "source_rows": [],
                },
            )

            if name and not profile["name"]:
                profile["name"] = name
            if email and not profile["email"]:
                profile["email"] = email
            if phone and not profile["phone"]:
                profile["phone"] = phone
            if company and not profile["company"]:
                profile["company"] = company
            if name:
                profile["names"].append(name)
            if location:
                profile["locations"].append(location)
            if state:
                profile["states"][state] += 1
            profile["sheet_names"].add(sheet.title)
            profile["source_rows"].append(
                {
                    "sheet": sheet.title,
                    "row": row_index,
                    "name": name,
                    "email": email,
                    "phone": phone,
                    "company": company,
                    "location": location,
                    "office_location": office_location,
                    "hq_location": hq_location,
                    "country": country,
                    "state": state,
                    "state_reason": state_reason,
                }
            )

    company_groups: dict[str, list[dict]] = defaultdict(list)
    for profile in profiles.values():
        company_groups[norm_key(profile.get("company"))].append(profile)

    for profile in profiles.values():
        profile["sheet_names"] = sorted(profile["sheet_names"])
        profile["locations"] = list(dict.fromkeys([loc for loc in profile["locations"] if loc]))
        profile["names"] = list(dict.fromkeys([norm_text(name) for name in profile.get("names", []) if norm_text(name)]))
        profile["preferred_name"] = choose_preferred_name(profile["names"], profile.get("email")) or profile.get("name")
        profile["source_row_count"] = len(profile["source_rows"])
        if profile["states"]:
            profile["dominant_state"], profile["dominant_state_count"] = profile["states"].most_common(1)[0]
        else:
            profile["dominant_state"], profile["dominant_state_count"] = None, 0
        profile["company_group_size"] = len(company_groups[norm_key(profile.get("company"))])

    return list(profiles.values()), source_row_count, sheet_counts, company_groups


def build_db_indexes(session):
    company_rows = session.query(Company).all()
    company_by_norm = {row.normalized_company_name or norm_key(row.company_name): row for row in company_rows if row.company_name}
    company_by_id = {row.company_id: row for row in company_rows}

    recruiters = session.query(Recruiter).all()

    email_index = {}
    phone_index = {}
    name_company_index = defaultdict(list)
    for row in recruiters:
        if row.email:
            email_index[row.email.strip().lower()] = row
        phone = norm_phone(row.phone)
        if phone:
            phone_index[phone] = row
        company_name = company_by_id.get(row.company_id).company_name if row.company_id and company_by_id.get(row.company_id) else ""
        if row.recruiter_name and company_name:
            name_company_index[f"{norm_key(row.recruiter_name)}::{norm_key(company_name)}"].append(row)

    return company_by_norm, company_by_id, email_index, phone_index, name_company_index


def get_or_create_company(session, company_name: str | None, company_state: str | None, company_location: str | None, company_cache: dict[str, Company]):
    if not company_name:
        return None, False
    norm_company = norm_key(company_name)
    if norm_company in company_cache:
        company = company_cache[norm_company]
        changed = False
        if not company.state and company_state:
            company.state = company_state
            changed = True
        if not company.location and company_location:
            company.location = company_location
            changed = True
        return company, changed

    company = Company(
        company_name=company_name,
        normalized_company_name=norm_company,
        location=company_location,
        state=company_state,
        data_source="location_workbook",
        source_job_id=BATCH_ID,
        is_active=True,
    )
    session.add(company)
    session.flush()
    company_cache[norm_company] = company
    return company, True


def merge_metadata(existing_value, evidence):
    metadata = {}
    if existing_value:
        try:
            parsed = json.loads(existing_value) if isinstance(existing_value, str) else existing_value
            if isinstance(parsed, dict):
                metadata = dict(parsed)
        except Exception:
            metadata = {"raw_metadata": str(existing_value)}
    metadata["location_workbook_evidence"] = evidence
    return json.dumps(metadata, default=str)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE_FILE)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--batch-size", type=int, default=250)
    args = parser.parse_args()

    if not args.source.exists():
        raise FileNotFoundError(f"Workbook not found: {args.source}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    profiles, source_row_count, sheet_counts, company_groups = parse_workbook(args.source)

    session = SessionLocal(expire_on_commit=False)
    try:
        company_cache = {}
        db_company_norm_map, db_company_by_id, email_index, phone_index, name_company_index = build_db_indexes(session)
        company_cache.update(db_company_norm_map)

        existing_matches = []
        unique_candidates = []
        skipped_no_qualification = []
        ambiguous_existing = []
        stats = Counter()
        source_state_counts = Counter()
        company_state_summary = {}
        state_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NOT NULL AND state != ''")).scalar() or 0
        unknown_before = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NULL OR state = ''")).scalar() or 0
        company_count_before = session.execute(text("SELECT COUNT(*) FROM companies")).scalar() or 0

        workbook_company_seed: dict[str, dict] = {}
        for profile in profiles:
            company_name = profile.get("company")
            if not company_name:
                continue
            company_key = norm_key(company_name)
            if not company_key or company_key in company_cache or company_key in workbook_company_seed:
                continue
            workbook_company_seed[company_key] = {
                "company_name": company_name[:255],
                "normalized_company_name": company_key[:255],
                "location": profile.get("locations")[0] if profile.get("locations") else None,
                "state": profile.get("dominant_state"),
                "data_source": "location_workbook",
                "source_job_id": BATCH_ID,
                "is_active": True,
            }

        seeded_company_rows = list(workbook_company_seed.values())
        if args.apply and seeded_company_rows:
            for chunk_start in range(0, len(seeded_company_rows), args.batch_size):
                chunk = seeded_company_rows[chunk_start:chunk_start + args.batch_size]
                session.bulk_insert_mappings(Company, chunk)
                session.commit()
            company_cache = {
                (row.normalized_company_name or norm_key(row.company_name)): row
                for row in session.query(Company).all()
                if row.company_name
            }
        companies_created = len(seeded_company_rows) if args.apply else 0

        for profile in profiles:
            if not profile.get("email") and not profile.get("phone") and not profile.get("locations"):
                skipped_no_qualification.append(profile)
                stats["skipped_no_email_phone_or_location"] += 1
                continue

            email = profile.get("email")
            phone = profile.get("phone")
            key = f"{norm_key(profile.get('name'))}::{norm_key(profile.get('company'))}" if profile.get("name") and profile.get("company") else None
            match = None
            match_type = None
            if email and email in email_index:
                match = email_index[email]
                match_type = "email"
            elif phone and phone in phone_index:
                match = phone_index[phone]
                match_type = "phone"
            elif key and key in name_company_index:
                matches = name_company_index[key]
                if len(matches) == 1:
                    match = matches[0]
                    match_type = "name_company"
                else:
                    ambiguous_existing.append(profile)
                    stats["ambiguous_existing"] += 1
                    continue

            if match:
                existing_matches.append(
                    {
                        "profile": profile,
                        "matched_recruiter_id": match.recruiter_id,
                        "matched_recruiter_name": match.recruiter_name,
                        "match_type": match_type,
                    }
                )
                stats[f"matched_{match_type}"] += 1
            else:
                unique_candidates.append(profile)
                stats["unique"] += 1

            if profile.get("dominant_state"):
                source_state_counts[profile["dominant_state"]] += 1
            company_key = norm_key(profile.get("company"))
            if company_key:
                company_state_summary.setdefault(company_key, Counter())
                if profile.get("dominant_state"):
                    company_state_summary[company_key][profile["dominant_state"]] += 1

        inserted_recruiters = 0
        updated_recruiters = 0
        companies_created = 0
        companies_updated = 0
        inserted_examples = []
        enriched_examples = []

        if args.apply:
            # First update existing recruiters with safe workbook evidence.
            for chunk_start in range(0, len(existing_matches), args.batch_size):
                chunk = existing_matches[chunk_start:chunk_start + args.batch_size]
                for item in chunk:
                    profile = item["profile"]
                    recruiter = email_index.get(profile.get("email"))
                    if not recruiter and profile.get("phone"):
                        recruiter = phone_index.get(profile.get("phone"))
                    if not recruiter and profile.get("name") and profile.get("company"):
                        recruiter_matches = name_company_index.get(f"{norm_key(profile.get('name'))}::{norm_key(profile.get('company'))}", [])
                        recruiter = recruiter_matches[0] if len(recruiter_matches) == 1 else None
                    if not recruiter:
                        continue

                    company_name = profile.get("company")
                    company_key = norm_key(company_name)
                    company = company_cache.get(company_key) if company_key else None
                    company_changed = False
                    if company and recruiter.company_id != company.company_id:
                        recruiter.company_id = company.company_id
                        company_changed = True

                    profile_state = profile.get("dominant_state")
                    workbook_location = profile.get("locations")[0] if profile.get("locations") else None
                    preferred_name = profile.get("preferred_name") or profile.get("name")
                    evidence = {
                        "batch_id": BATCH_ID,
                        "source_file": str(args.source),
                        "match_type": item["match_type"],
                        "workbook_name": preferred_name,
                        "workbook_all_names": profile.get("names"),
                        "workbook_email": profile.get("email"),
                        "workbook_phone": profile.get("phone"),
                        "workbook_company": profile.get("company"),
                        "workbook_locations": profile.get("locations"),
                        "workbook_dominant_state": profile_state,
                        "workbook_sheet_names": profile.get("sheet_names"),
                        "existing_name": recruiter.recruiter_name,
                    }

                    changed = False
                    if should_promote_name(recruiter.recruiter_name, preferred_name, recruiter.email):
                        recruiter.recruiter_name = preferred_name[:150]
                        recruiter.normalized_recruiter_name = norm_key(preferred_name)
                        changed = True

                    if not recruiter.location and workbook_location:
                        recruiter.location = workbook_location
                        changed = True

                    if not recruiter.state and profile_state:
                        recruiter.state = profile_state
                        recruiter.state_source = "location_workbook"
                        recruiter.state_confidence = "high"
                        recruiter.state_reason = "location_workbook_dominant_state"
                        changed = True

                    if recruiter.state and profile_state and recruiter.state != profile_state:
                        evidence["state_conflict"] = {
                            "db_state": recruiter.state,
                            "workbook_state": profile_state,
                        }

                    if recruiter.needs_review and (recruiter.state or recruiter.location):
                        recruiter.needs_review = False
                        recruiter.review_reason = None
                        changed = True

                    recruiter.metadata_json = merge_metadata(recruiter.metadata_json, evidence)
                    recruiter.last_scan_at = datetime.now(timezone.utc)
                    if changed or company_changed:
                        updated_recruiters += 1
                        if len(enriched_examples) < 20:
                            enriched_examples.append(
                                {
                                    "recruiter_id": recruiter.recruiter_id,
                                    "recruiter_name": recruiter.recruiter_name,
                                    "email": recruiter.email,
                                    "company": company_name,
                                    "match_type": item["match_type"],
                                    "state": recruiter.state,
                                    "location": recruiter.location,
                                }
                            )

                session.commit()

            # Then insert only truly unique candidates with the minimum qualification.
            for profile in unique_candidates:
                company_name = profile.get("company")
                profile_state = profile.get("dominant_state")
                workbook_location = profile.get("locations")[0] if profile.get("locations") else None
                preferred_name = profile.get("preferred_name") or profile.get("name") or "Unknown"
                if not profile.get("email") and not profile.get("phone") and not workbook_location:
                    continue

                company_key = norm_key(company_name)
                company = company_cache.get(company_key) if company_key else None

                recruiter = Recruiter(
                    recruiter_name=truncate(preferred_name, 150) or "Unknown",
                    normalized_recruiter_name=truncate(norm_key(preferred_name), 150) or None,
                    email=truncate(build_placeholder_email(profile), 150) or f"no-email-{hashlib.sha1((profile.get('identity_key') or str(profile.get('source_rows', []))).encode('utf-8')).hexdigest()[:16]}@missing.local",
                    phone=truncate(profile.get("phone"), 30),
                    company_id=company.company_id if company else None,
                    location=truncate(workbook_location, 255),
                    state=truncate(profile_state, 2),
                    state_source=truncate("location_workbook" if profile_state else None, 150),
                    state_confidence=truncate("high" if profile_state else None, 50),
                    state_reason=truncate("location_workbook_dominant_state" if profile_state else None, 500),
                    needs_review=not bool(profile_state),
                    review_reason=None if profile_state else "Workbook row qualified by phone/location but state could not be derived",
                    data_source="location_workbook",
                    source_job_id=BATCH_ID,
                    raw_data=json.dumps(profile.get("source_rows", []), default=str),
                    metadata_json=json.dumps(
                        {
                            "batch_id": BATCH_ID,
                        "source_file": str(args.source),
                            "workbook_name": preferred_name,
                            "workbook_all_names": profile.get("names"),
                            "workbook_email": profile.get("email"),
                            "workbook_phone": profile.get("phone"),
                            "workbook_company": company_name,
                            "workbook_locations": profile.get("locations"),
                            "workbook_sheet_names": profile.get("sheet_names"),
                            "workbook_source_row_count": profile.get("source_row_count"),
                            "workbook_dominant_state": profile_state,
                        },
                        default=str,
                    ),
                )
                session.add(recruiter)
                inserted_recruiters += 1
                if len(inserted_examples) < 20:
                    inserted_examples.append(
                        {
                            "name": recruiter.recruiter_name,
                            "email": recruiter.email,
                            "company": company_name,
                            "phone": recruiter.phone,
                            "state": recruiter.state,
                            "location": recruiter.location,
                        }
                    )

                if inserted_recruiters % args.batch_size == 0:
                    session.commit()

            session.commit()

        state_after = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NOT NULL AND state != ''")).scalar() or 0
        unknown_after = session.execute(text("SELECT COUNT(*) FROM recruiters WHERE state IS NULL OR state = ''")).scalar() or 0
        company_count_after = session.execute(text("SELECT COUNT(*) FROM companies")).scalar() or 0
        state_counts = session.execute(text("""
            SELECT state, COUNT(*) AS cnt
            FROM recruiters
            WHERE state IS NOT NULL AND state != ''
            GROUP BY state
            ORDER BY cnt DESC, state ASC
        """)).mappings().all()

        report = {
            "batch_id": BATCH_ID,
            "source_file": str(args.source),
            "apply": bool(args.apply),
            "source_row_count": source_row_count,
            "merged_profiles": len(profiles),
            "existing_matches": len(existing_matches),
            "unique_candidates": len(unique_candidates),
            "skipped_no_email_phone_or_location": len(skipped_no_qualification),
            "ambiguous_existing": len(ambiguous_existing),
            "stats": dict(stats),
            "source_state_counts": dict(source_state_counts),
            "state_before": int(state_before),
            "state_after": int(state_after),
            "unknown_before": int(unknown_before),
            "unknown_after": int(unknown_after),
            "company_count_before": int(company_count_before),
            "company_count_after": int(company_count_after),
            "inserted_recruiters": int(inserted_recruiters),
            "updated_recruiters": int(updated_recruiters),
            "companies_created": int(companies_created),
            "companies_updated": int(companies_updated),
            "inserted_examples": inserted_examples,
            "enriched_examples": enriched_examples,
            "skipped_examples": skipped_no_qualification[:20],
            "state_counts": {row["state"]: int(row["cnt"]) for row in state_counts},
            "top_company_state_breakdown": {
                company_key: dict(counter.most_common())
                for company_key, counter in sorted(company_state_summary.items(), key=lambda item: (-sum(item[1].values()), item[0]))[:25]
            },
        }

        json_path = OUTPUT_DIR / "location_workbook_apply_report.json"
        md_path = OUTPUT_DIR / "location_workbook_apply_report.md"
        json_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")

        lines = [
            "# Location Workbook Apply Report",
            "",
            f"Source workbook: `{args.source}`",
            f"Batch id: `{BATCH_ID}`",
            "",
            "## Results",
            f"- Source rows: `{source_row_count}`",
            f"- Merged profiles: `{len(profiles)}`",
            f"- Existing matches: `{len(existing_matches)}`",
            f"- Unique candidates: `{len(unique_candidates)}`",
            f"- Skipped missing email/phone/location: `{len(skipped_no_qualification)}`",
            f"- Ambiguous existing matches: `{len(ambiguous_existing)}`",
            f"- Inserted recruiters: `{inserted_recruiters}`",
            f"- Updated recruiters: `{updated_recruiters}`",
            f"- Companies created: `{companies_created}`",
            f"- Companies updated: `{companies_updated}`",
            "",
            "## Database impact",
            f"- State before: `{state_before}`",
            f"- State after: `{state_after}`",
            f"- Unknown before: `{unknown_before}`",
            f"- Unknown after: `{unknown_after}`",
            f"- Companies before: `{company_count_before}`",
            f"- Companies after: `{company_count_after}`",
            "",
            "## Top state evidence",
        ]
        for state, count in Counter(source_state_counts).most_common(20):
            lines.append(f"- {state}: `{count}`")
        lines.extend([
            "",
            "## Files",
            f"- JSON: `{json_path}`",
            f"- Markdown: `{md_path}`",
        ])
        md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(json.dumps(report, indent=2, default=str))

    finally:
        session.close()


if __name__ == "__main__":
    main()
