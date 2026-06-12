from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from app.database import SessionLocal
from app.utils.phone_normalizer import format_us_phone, normalize_us_phone_digits

WORKBOOK_PATH = Path(r"C:\Users\User\Desktop\final updated sheet.xlsx")
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_RE = re.compile(r"\+?\d[\d\s()./-]{7,}\d")
PLACEHOLDERS = {"", "-", "—", "--", "n/a", "na", "none", "null", "#value!", "0"}


@dataclass
class Candidate:
    row_index: int
    email: str
    name: str
    company: str
    phones: list[str]
    raw_cells: list[str]


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def is_placeholder(value: str) -> bool:
    return clean(value).lower() in PLACEHOLDERS


def extract_emails(cells: list[str]) -> list[tuple[int, str]]:
    found = []
    seen = set()
    for idx, cell in enumerate(cells):
        for match in EMAIL_RE.findall(cell):
            lowered = match.lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            found.append((idx, match))
    return found


def extract_phones(cells: list[str]) -> list[str]:
    phones = []
    seen = set()
    for cell in cells:
        for match in PHONE_RE.findall(cell):
            formatted = format_us_phone(match) or clean(match)
            digits = normalize_us_phone_digits(formatted)
            if not digits or len(digits) < 10 or digits in seen:
                continue
            seen.add(digits)
            phones.append(formatted)
    return phones


def textish(value: str) -> bool:
    value = clean(value)
    if not value or is_placeholder(value):
        return False
    if EMAIL_RE.search(value) or PHONE_RE.search(value):
        return False
    if value.lower().startswith("http") or "linkedin.com" in value.lower() or "www." in value.lower():
        return False
    return any(ch.isalpha() for ch in value)


def infer_name_company(cells: list[str], email_idx: int) -> tuple[str, str]:
    left = [(idx, clean(cells[idx])) for idx in range(max(0, email_idx - 2), email_idx) if textish(cells[idx])]
    right = [(idx, clean(cells[idx])) for idx in range(email_idx + 1, min(len(cells), email_idx + 4)) if textish(cells[idx])]

    if len(left) >= 2:
        return left[-1][1], left[-2][1]
    if len(left) == 1 and len(right) >= 1:
        return left[-1][1], right[0][1]
    if len(right) >= 2:
        return right[0][1], right[1][1]
    if len(left) == 1:
        return left[0][1], ""
    if len(right) == 1:
        return right[0][1], ""
    return "", ""


def iter_candidates() -> list[Candidate]:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)

    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    candidates: list[Candidate] = []

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [clean(value) for value in row]
        email_hits = extract_emails(cells)
        if not email_hits:
            continue
        row_phones = extract_phones(cells)
        for email_idx, email in email_hits:
            name, company = infer_name_company(cells, email_idx)
            if not name:
                continue
            candidates.append(Candidate(
                row_index=row_index,
                email=email,
                name=name,
                company=company,
                phones=row_phones,
                raw_cells=cells[:12],
            ))
    wb.close()
    return candidates


def apply_repairs(limit: int | None = None) -> dict:
    db = SessionLocal()
    processed = 0
    updated = 0
    examples = []
    try:
        candidates = iter_candidates()
        if limit is not None:
            candidates = candidates[:limit]

        for candidate in candidates:
            processed += 1
            recruiter = db.execute(text("""
                select recruiter_id, recruiter_name, email, phone, phone2, phone3, phone4, metadata_json
                from recruiters
                where lower(email) = lower(:email)
                limit 1
            """), {"email": candidate.email}).mappings().first()
            if not recruiter:
                continue

            metadata = {}
            if recruiter["metadata_json"]:
                try:
                    metadata = json.loads(recruiter["metadata_json"])
                except Exception:
                    metadata = {}

            all_phones = metadata.get("all_phones") or []
            for phone in candidate.phones:
                digits = normalize_us_phone_digits(phone)
                if not any(normalize_us_phone_digits(existing) == digits for existing in all_phones):
                    all_phones.append(phone)
            metadata["all_phones"] = all_phones

            workbook_repairs = metadata.get("workbook_repairs") or []
            workbook_repairs.append({
                "source_file": str(WORKBOOK_PATH),
                "row": candidate.row_index,
                "name": candidate.name,
                "company": candidate.company,
            })
            metadata["workbook_repairs"] = workbook_repairs[-10:]

            phone_slots = [None, None, None, None]
            for idx, phone in enumerate(candidate.phones[:4]):
                phone_slots[idx] = phone

            db.execute(text("""
                update recruiters
                set recruiter_name = :name,
                    phone = coalesce(:phone, phone),
                    phone2 = coalesce(:phone2, phone2),
                    phone3 = coalesce(:phone3, phone3),
                    phone4 = coalesce(:phone4, phone4),
                    metadata_json = :metadata_json,
                    updated_at = now()
                where recruiter_id = :recruiter_id
            """), {
                "name": candidate.name,
                "phone": phone_slots[0],
                "phone2": phone_slots[1],
                "phone3": phone_slots[2],
                "phone4": phone_slots[3],
                "metadata_json": json.dumps(metadata, default=str),
                "recruiter_id": recruiter["recruiter_id"],
            })
            updated += 1
            if len(examples) < 20:
                examples.append({
                    "row": candidate.row_index,
                    "email": candidate.email,
                    "name": candidate.name,
                    "company": candidate.company,
                    "phones": candidate.phones,
                })

        db.commit()
        return {"processed": processed, "updated": updated, "examples": examples}
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    print(json.dumps(apply_repairs(), indent=2))
