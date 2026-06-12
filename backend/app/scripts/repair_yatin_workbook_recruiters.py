from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.models import Company, Recruiter
from app.utils.normalizer import normalize_text
from app.utils.phone_normalizer import format_us_phone, normalize_us_phone_digits

WORKBOOK_PATH = Path(r"C:\Users\User\Desktop\for yatin sir , may 28 ,2026.xlsx")
PLACEHOLDERS = {"", "n/a", "na", "none", "null", "-", "—"}
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_RE = re.compile(r"\+?\d[\d\s()./-]{7,}\d")


def clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_placeholder(value: str) -> bool:
    return value.strip().lower() in PLACEHOLDERS


def looks_like_url(value: str) -> bool:
    lowered = value.lower()
    return lowered.startswith("http://") or lowered.startswith("https://") or "linkedin.com/" in lowered or "www." in lowered


def extract_emails(cells: list[str]) -> list[str]:
    emails: list[str] = []
    seen = set()
    for cell in cells:
        for match in EMAIL_RE.findall(cell):
            lowered = match.strip().lower()
            if lowered in seen:
                continue
            seen.add(lowered)
            emails.append(match.strip())
    return emails


def extract_phones(cells: list[str]) -> list[str]:
    phones: list[str] = []
    seen = set()
    for cell in cells:
        for match in PHONE_RE.findall(cell):
            formatted = format_us_phone(match) or match.strip()
            compare = normalize_us_phone_digits(formatted)
            if not compare or len(compare) < 10 or compare in seen:
                continue
            seen.add(compare)
            phones.append(formatted)
    return phones


def text_candidates(cells: list[str], email_cells: set[int], phone_cells: set[int]) -> list[tuple[int, str]]:
    items: list[tuple[int, str]] = []
    for index, value in enumerate(cells):
        if not value or is_placeholder(value) or looks_like_url(value):
            continue
        if index in email_cells or index in phone_cells:
            continue
        if EMAIL_RE.search(value):
            continue
        if PHONE_RE.search(value):
            continue
        if not re.search(r"[A-Za-z]", value):
            continue
        items.append((index, value))
    return items


def parse_row(cells: list[str]) -> dict | None:
    email_cells = {index for index, value in enumerate(cells) if EMAIL_RE.search(value)}
    emails = extract_emails(cells)
    if not emails:
        return None

    phone_cells = {index for index, value in enumerate(cells) if PHONE_RE.search(value)}
    phones = extract_phones(cells)
    text_cells = text_candidates(cells, email_cells, phone_cells)
    if not text_cells:
        return None

    first_email_index = min(email_cells)
    name = ""
    company = ""

    if first_email_index == 1 and cells[0] and not is_placeholder(cells[0]):
        name = cells[0]
        remaining = [value for index, value in text_cells if index > first_email_index and value != name]
        company = remaining[0] if remaining else ""
    elif first_email_index == 2 and len(cells) > 1 and cells[0] and cells[1]:
        company = cells[0] if not is_placeholder(cells[0]) else ""
        name = cells[1] if not is_placeholder(cells[1]) else ""
    else:
        before = [value for index, value in text_cells if index < first_email_index]
        after = [value for index, value in text_cells if index > first_email_index]
        name = before[-1] if before else text_cells[0][1]
        company = next((value for value in after if value != name), "")

    name = clean_cell(name)
    company = clean_cell(company)
    if not name:
        return None

    return {
        "name": name,
        "company": company,
        "emails": emails,
        "phones": phones,
    }


def load_metadata(value) -> dict:
    if not value:
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def load_raw_rows(value) -> list[dict]:
    if not value:
        return []
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            return [item for item in parsed if isinstance(item, dict)]
    except Exception:
        return []
    return []


def find_company(db: Session, company_name: str) -> Company | None:
    company_name = clean_cell(company_name)
    if not company_name:
        return None
    normalized = normalize_text(company_name)
    company = db.query(Company).filter(Company.normalized_company_name == normalized).first()
    if company:
        return company
    company = db.query(Company).filter(Company.company_name.ilike(company_name)).first()
    return company


def assign_contacts(recruiter: Recruiter, emails: list[str], phones: list[str]) -> None:
    unique_emails: list[str] = []
    seen_emails = set()
    for value in emails:
        lowered = value.strip().lower()
        if not lowered or lowered in seen_emails:
            continue
        seen_emails.add(lowered)
        unique_emails.append(value.strip())

    unique_phones: list[str] = []
    seen_phones = set()
    for value in phones:
        digits = normalize_us_phone_digits(value)
        if not digits or digits in seen_phones:
            continue
        seen_phones.add(digits)
        unique_phones.append(format_us_phone(value) or value.strip())

    email_slots = ["email", "email2", "email3", "email4"]
    phone_slots = ["phone", "phone2", "phone3", "phone4"]

    if unique_emails:
        current_primary = clean_cell(getattr(recruiter, "email", ""))
        if current_primary.startswith("no-email-") and current_primary.endswith("@missing.local"):
            recruiter.email = unique_emails[0]

    current_emails = []
    for slot in email_slots:
        value = clean_cell(getattr(recruiter, slot, ""))
        if value:
            current_emails.append(value)
    for value in unique_emails:
        if any(existing.lower() == value.lower() for existing in current_emails):
            continue
        for slot in email_slots:
            if not clean_cell(getattr(recruiter, slot, "")):
                setattr(recruiter, slot, value)
                current_emails.append(value)
                break

    current_phones = []
    for slot in phone_slots:
        value = clean_cell(getattr(recruiter, slot, ""))
        if value:
            current_phones.append(value)
    for value in unique_phones:
        compare = normalize_us_phone_digits(value)
        if any(normalize_us_phone_digits(existing) == compare for existing in current_phones):
            continue
        for slot in phone_slots:
            if not clean_cell(getattr(recruiter, slot, "")):
                setattr(recruiter, slot, value)
                current_phones.append(value)
                break


def append_repair_metadata(recruiter: Recruiter, row_index: int, profile: dict) -> None:
    metadata = load_metadata(recruiter.metadata_json)
    repairs = metadata.setdefault("workbook_repairs", [])
    repair_entry = {
        "source_file": str(WORKBOOK_PATH),
        "row": row_index,
        "name": profile["name"],
        "company": profile["company"],
        "emails": profile["emails"],
        "phones": profile["phones"],
        "repaired_at": datetime.now(timezone.utc).isoformat(),
    }
    if repair_entry not in repairs:
        repairs.append(repair_entry)
    metadata["all_emails"] = sorted(
        {value for value in metadata.get("all_emails", []) if clean_cell(value)}
        | {value for value in profile["emails"] if clean_cell(value)},
        key=str.lower,
    )
    metadata["all_phones"] = sorted(
        {value for value in metadata.get("all_phones", []) if clean_cell(value)}
        | {format_us_phone(value) or value for value in profile["phones"] if clean_cell(value)},
    )
    recruiter.metadata_json = json.dumps(metadata, default=str)


def repair_workbook() -> None:
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_by_index = {
        row_index: [clean_cell(value) for value in row[:8]]
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1)
    }
    wb.close()

    db = SessionLocal()
    updated = 0
    deleted = 0
    skipped = 0
    processed = 0
    repaired_examples: list[tuple[int, str, str]] = []
    try:
        malformed_rows = db.execute(text("""
            SELECT recruiter_id, metadata_json, raw_data
            FROM recruiters
            WHERE email LIKE 'no-email-%@missing.local'
              AND metadata_json::text LIKE :source_file
        """), {"source_file": f"%{WORKBOOK_PATH}%"}).all()

        malformed_by_row: dict[int, list[int]] = {}
        for recruiter_id, metadata_json, raw_data in malformed_rows:
            metadata = load_metadata(metadata_json)
            row_index = None
            raw_rows = load_raw_rows(raw_data)
            if raw_rows:
                row_index = raw_rows[0].get("row")
            if not row_index:
                evidence = metadata.get("location_workbook_evidence") or {}
                raw_rows = evidence.get("source_rows") or []
                if raw_rows:
                    row_index = raw_rows[0].get("row")
            if not row_index:
                continue
            malformed_by_row.setdefault(int(row_index), []).append(recruiter_id)

        for row_index, cells in rows_by_index.items():
            profile = parse_row(cells)
            if not profile:
                continue

            exact_recruiter = db.query(Recruiter).filter(Recruiter.email.ilike(profile["emails"][0])).first()
            malformed_ids = malformed_by_row.get(row_index, [])
            if not exact_recruiter and not malformed_ids:
                continue

            processed += 1
            target = exact_recruiter
            if not target and malformed_ids:
                target = db.query(Recruiter).filter(Recruiter.recruiter_id == malformed_ids[0]).first()
            if not target:
                skipped += 1
                continue

            company = find_company(db, profile["company"]) if profile["company"] else None
            if profile["name"]:
                target.recruiter_name = profile["name"]
            if company:
                target.company_id = company.company_id
            assign_contacts(target, profile["emails"], profile["phones"])
            append_repair_metadata(target, row_index, profile)
            updated += 1
            if len(repaired_examples) < 20:
                repaired_examples.append((row_index, target.recruiter_name, target.email))

            for malformed_id in malformed_ids:
                if malformed_id == target.recruiter_id:
                    continue
                malformed = db.query(Recruiter).filter(Recruiter.recruiter_id == malformed_id).first()
                if malformed is None:
                    continue
                db.delete(malformed)
                deleted += 1

        db.commit()
        print(json.dumps({
            "processed_rows": processed,
            "updated_recruiters": updated,
            "deleted_malformed_duplicates": deleted,
            "skipped": skipped,
            "examples": repaired_examples,
        }, indent=2))
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    repair_workbook()
