"""
Adaptive multi-format file parser for TalentOps ETL.

Handles:
- Multiple sheets in xlsx files
- Headerless sheets (positional mapping)
- Multiple column layouts (Format A/B/C/D and future formats)
- Blank row skipping
- Content-based column detection (email regex, phone regex)
- Raw data preservation

This module does NOT touch the database. It is a pure parsing layer.
"""

import csv
import io
import json
import os
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

# ──────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
PHONE_RE = re.compile(r"[\d]{7,}")
BAD_PHONE_VALUES = frozenset({
    "0", "-", "--", "---", "000-000-0000", "0000000000", "000000000",
    "n/a", "na", "none", "null", "", "0000000", "1111111111",
})

# Keyword groups for heuristic column detection
COLUMN_KEYWORDS: dict[str, list[str]] = {
    "name": [
        "name", "full_name", "full name", "recruiter_name", "recruiter name",
        "candidate_name", "contact_name", "contact name", "person",
    ],
    "email": [
        "email", "mail", "e-mail", "email_id", "work_email", "work email",
        "primary_email", "primary email", "email address", "email_address",
    ],
    "email2": [
        "email2", "email 2", "alt_email", "alternate_email", "secondary_email",
        "secondary email", "personal_email", "personal email", "other email",
    ],
    "email3": [
        "email3", "email 3", "third_email", "third email"
    ],
    "email4": [
        "email4", "email 4", "fourth_email", "fourth email"
    ],
    "phone": [
        "phone", "mobile", "cell", "tel", "telephone", "contact",
        "phone_number", "phone number", "mobile_number", "work phone",
    ],
    "phone2": [
        "phone2", "phone 2", "alt_phone", "alternate_phone", "secondary_phone",
        "secondary phone", "personal phone", "home phone", "other phone",
    ],
    "phone3": [
        "phone3", "phone 3", "third_phone", "third phone"
    ],
    "phone4": [
        "phone4", "phone 4", "fourth_phone", "fourth phone"
    ],
    "company": [
        "company", "organization", "org", "vendor", "client", "employer",
        "firm", "company_name", "company name", "staffing company",
    ],
    "location": [
        "location", "city", "address", "address_line", "office_location",
        "city_state", "city/state",
    ],
    "state": [
        "state", "state_abbr", "province", "region", "us_state",
    ],
    "linkedin": [
        "linkedin", "linkedin_url", "profile_link", "profile_url",
        "linkedin url", "linkedin profile",
    ],
    "title": [
        "title", "role", "position", "job_title", "job title", "designation",
    ],
    "specialization": [
        "specialization", "skills", "expertise", "focus_area", "focus area",
        "specialty", "vertical",
    ],
    "notes": [
        "notes", "remarks", "comments", "additional_info", "additional info",
        "description",
    ],
}


# ──────────────────────────────────────────────────────────────
# Data classes
# ──────────────────────────────────────────────────────────────

@dataclass
class ColumnMapping:
    """Mapping from logical field names to column headers/indices."""
    name: str | None = None
    email: str | None = None
    email2: str | None = None
    email3: str | None = None
    email4: str | None = None
    phone: str | None = None
    phone2: str | None = None
    phone3: str | None = None
    phone4: str | None = None
    company: str | None = None
    location: str | None = None
    state: str | None = None
    linkedin: str | None = None
    title: str | None = None
    specialization: str | None = None
    notes: str | None = None
    # Columns that didn't map to any known field
    unmapped_columns: list[str] = field(default_factory=list)
    # How the mapping was determined
    detection_method: str = "header_keywords"  # header_keywords | content_analysis | positional

    def to_dict(self) -> dict[str, str | None]:
        return {
            "name": self.name, "email": self.email, "email2": self.email2,
            "email3": self.email3, "email4": self.email4,
            "phone": self.phone, "phone2": self.phone2, 
            "phone3": self.phone3, "phone4": self.phone4, 
            "company": self.company,
            "location": self.location, "state": self.state, "linkedin": self.linkedin,
            "title": self.title, "specialization": self.specialization, "notes": self.notes,
        }

    @property
    def has_email(self) -> bool:
        return self.email is not None

    @property
    def confidence(self) -> str:
        """Return confidence level of this mapping."""
        if self.detection_method == "positional":
            return "low"
        if self.email and self.name:
            return "high"
        if self.email or self.name:
            return "medium"
        return "low"


@dataclass
class ParsedRow:
    """A single parsed recruiter row with mapped fields and preserved originals."""
    # Core mapped fields
    name: str | None = None
    email: str | None = None
    email2: str | None = None
    email3: str | None = None
    email4: str | None = None
    phone: str | None = None
    phone2: str | None = None
    phone3: str | None = None
    phone4: str | None = None
    company: str | None = None
    location: str | None = None
    state: str | None = None
    linkedin: str | None = None
    title: str | None = None
    specialization: str | None = None
    notes: str | None = None

    # Preservation fields
    raw_data: dict[str, Any] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)
    source_file: str = ""
    source_sheet: str = ""
    row_index: int = 0

    # Review flags
    needs_review: bool = False
    review_reasons: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name, "email": self.email, "email2": self.email2,
            "email3": self.email3, "email4": self.email4,
            "phone": self.phone, "phone2": self.phone2, 
            "phone3": self.phone3, "phone4": self.phone4,
            "company": self.company,
            "location": self.location, "state": self.state, "linkedin": self.linkedin,
            "title": self.title, "specialization": self.specialization, "notes": self.notes,
            "raw_data": self.raw_data,
            "metadata_json": self.metadata,
            "source_file": self.source_file, "source_sheet": self.source_sheet,
            "row_index": self.row_index,
            "needs_review": self.needs_review,
            "review_reasons": self.review_reasons,
        }


@dataclass
class SheetParseResult:
    """Result of parsing a single sheet."""
    sheet_name: str
    detected_format: str  # SIMPLE_ROW, HEADERLESS, WIDE_CONTACT, VERTICAL, UNKNOWN
    format_confidence: str  # high, medium, low
    column_mapping: ColumnMapping
    has_headers: bool
    headers: list[str]
    total_rows: int
    data_rows: int
    blank_rows: int
    rows_with_email: int
    rows_without_email: int
    parsed_rows: list[ParsedRow] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class FileParseResult:
    """Result of parsing an entire file."""
    filename: str
    file_size_bytes: int
    sheet_count: int
    sheet_names: list[str]
    sheets: list[SheetParseResult] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def total_parsed_rows(self) -> int:
        return sum(s.data_rows for s in self.sheets)

    @property
    def total_emails(self) -> int:
        return sum(s.rows_with_email for s in self.sheets)


# ──────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────

def _clean_val(v: Any) -> str | None:
    """Clean a cell value, returning None for empty/null-like values."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() in ("none", "n/a", "nan", "null", ""):
        return None
    return s


def _clean_email(v: Any) -> str | None:
    """Normalize an email address."""
    s = _clean_val(v)
    if not s:
        return None
    s = s.lower().strip()
    if EMAIL_RE.match(s):
        return s
    return None


def _clean_phone(v: Any) -> str | None:
    """Normalize a phone number, rejecting known bad placeholders."""
    s = _clean_val(v)
    if not s:
        return None
    cleaned = s.replace("-", "").replace(" ", "").replace("(", "").replace(")", "").replace("+", "").strip()
    if cleaned.lower() in BAD_PHONE_VALUES:
        return None
    if PHONE_RE.search(cleaned):
        return s.strip()  # Return original formatting but trimmed
    return None


def _normalize_header(h: str) -> str:
    """Normalize a header string for matching."""
    return re.sub(r"[^a-z0-9]", "", h.lower())


# ──────────────────────────────────────────────────────────────
# Column Mapping Detection
# ──────────────────────────────────────────────────────────────

def detect_column_mapping(headers: list[str], sample_rows: list[list[str]] | None = None) -> ColumnMapping:
    """
    Detect column mapping using:
    1. Header keyword matching (primary)
    2. Content-based analysis (fallback for unmatched columns)
    3. Positional heuristic (last resort for headerless data)
    """
    mapping = ColumnMapping()
    if not headers:
        return mapping

    normalized_headers = {_normalize_header(h): h for h in headers if h}
    mapped_logical: set[str] = set()
    mapped_headers: set[str] = set()

    # Pass 1: Exact and substring keyword match on headers
    for logical, keywords in COLUMN_KEYWORDS.items():
        if logical in mapped_logical:
            continue
        for kw in keywords:
            kw_norm = _normalize_header(kw)
            # Exact match
            if kw_norm in normalized_headers and normalized_headers[kw_norm] not in mapped_headers:
                setattr(mapping, logical, normalized_headers[kw_norm])
                mapped_logical.add(logical)
                mapped_headers.add(normalized_headers[kw_norm])
                break
            # Substring match
            for nh, original in normalized_headers.items():
                if kw_norm in nh and original not in mapped_headers:
                    setattr(mapping, logical, original)
                    mapped_logical.add(logical)
                    mapped_headers.add(original)
                    break
            if logical in mapped_logical:
                break

    # Pass 2: Content-based detection for unmapped columns
    if sample_rows and len(sample_rows) > 0:
        for idx, h in enumerate(headers):
            if h in mapped_headers or not h:
                continue
            sample_vals = [
                row[idx] for row in sample_rows[:50]
                if idx < len(row) and _clean_val(row[idx])
            ]
            if not sample_vals:
                continue

            email_hits = sum(1 for v in sample_vals if EMAIL_RE.match(str(v).strip().lower()))
            phone_hits = sum(
                1 for v in sample_vals
                if PHONE_RE.search(str(v).replace("-", "").replace(" ", "").replace("(", "").replace(")", ""))
            )
            ratio = len(sample_vals)

            if email_hits / max(ratio, 1) > 0.5:
                if "email" not in mapped_logical:
                    mapping.email = h
                    mapped_logical.add("email")
                    mapped_headers.add(h)
                    mapping.detection_method = "content_analysis"
                elif "email2" not in mapped_logical:
                    mapping.email2 = h
                    mapped_logical.add("email2")
                    mapped_headers.add(h)
                elif "email3" not in mapped_logical:
                    mapping.email3 = h
                    mapped_logical.add("email3")
                    mapped_headers.add(h)
                elif "email4" not in mapped_logical:
                    mapping.email4 = h
                    mapped_logical.add("email4")
                    mapped_headers.add(h)
            elif phone_hits / max(ratio, 1) > 0.5:
                if "phone" not in mapped_logical:
                    mapping.phone = h
                    mapped_logical.add("phone")
                    mapped_headers.add(h)
                    mapping.detection_method = "content_analysis"
                elif "phone2" not in mapped_logical:
                    mapping.phone2 = h
                    mapped_logical.add("phone2")
                    mapped_headers.add(h)
                elif "phone3" not in mapped_logical:
                    mapping.phone3 = h
                    mapped_logical.add("phone3")
                    mapped_headers.add(h)
                elif "phone4" not in mapped_logical:
                    mapping.phone4 = h
                    mapped_logical.add("phone4")
                    mapped_headers.add(h)

    # Track unmapped columns
    mapping.unmapped_columns = [h for h in headers if h and h not in mapped_headers]

    return mapping


def detect_positional_mapping(sample_rows: list[list[str]]) -> ColumnMapping:
    """
    For headerless data, detect column roles by analyzing content patterns.
    Returns a ColumnMapping with positional column references like 'Column_0'.
    """
    mapping = ColumnMapping(detection_method="positional")
    if not sample_rows:
        return mapping

    num_cols = max(len(row) for row in sample_rows)
    col_analysis: list[dict[str, int]] = [{"email": 0, "phone": 0, "text": 0} for _ in range(num_cols)]

    for row in sample_rows[:20]:
        for idx in range(min(len(row), num_cols)):
            val = _clean_val(row[idx])
            if not val:
                continue
            if EMAIL_RE.match(val.strip().lower()):
                col_analysis[idx]["email"] += 1
            elif PHONE_RE.search(val.replace("-", "").replace(" ", "").replace("(", "").replace(")", "")):
                col_analysis[idx]["phone"] += 1
            else:
                col_analysis[idx]["text"] += 1

    total_samples = len(sample_rows[:20])
    email_assigned = False
    phone_assigned = False

    for idx in range(num_cols):
        col_name = f"Column_{idx}"
        analysis = col_analysis[idx]

        if analysis["email"] > total_samples * 0.3 and not email_assigned:
            mapping.email = col_name
            email_assigned = True
        elif analysis["phone"] > total_samples * 0.3 and not phone_assigned:
            mapping.phone = col_name
            phone_assigned = True
        elif analysis["text"] > 0:
            # First unassigned text column → name, second → company
            if mapping.name is None:
                mapping.name = col_name
            elif mapping.company is None:
                mapping.company = col_name
            else:
                mapping.unmapped_columns.append(col_name)

    return mapping


# ──────────────────────────────────────────────────────────────
# Sheet Parsing
# ──────────────────────────────────────────────────────────────

def _detect_headers(first_row: list[str]) -> bool:
    """Heuristic: does the first row look like headers?"""
    if not first_row:
        return False

    non_empty = [v for v in first_row if v]
    if len(non_empty) < 2:
        return False

    # If any cell in the first row looks like an email, it's probably data, not headers
    if any(EMAIL_RE.match(str(v).strip().lower()) for v in first_row if v):
        return False

    # If all values are short text (< 50 chars), likely headers
    if all(len(str(v)) < 50 for v in first_row if v):
        return True

    return False


def parse_sheet_rows(
    rows: list[list[str]],
    sheet_name: str,
    source_file: str,
) -> SheetParseResult:
    """Parse all rows from a single sheet into structured data."""

    if not rows:
        return SheetParseResult(
            sheet_name=sheet_name, detected_format="EMPTY",
            format_confidence="high", column_mapping=ColumnMapping(),
            has_headers=False, headers=[], total_rows=0, data_rows=0,
            blank_rows=0, rows_with_email=0, rows_without_email=0,
        )

    # Detect headers
    has_headers = _detect_headers(rows[0])

    if has_headers:
        headers = [str(v).strip() if v else "" for v in rows[0]]
        data_rows_raw = rows[1:]
    else:
        headers = [f"Column_{i}" for i in range(len(rows[0]))]
        data_rows_raw = rows

    # Detect column mapping
    sample_data = data_rows_raw[:50]
    if has_headers:
        mapping = detect_column_mapping(headers, sample_data)
    else:
        mapping = detect_positional_mapping(sample_data)

    # Determine format
    if not has_headers:
        detected_format = "HEADERLESS"
    elif len([h for h in headers if h]) <= 5:
        detected_format = "SIMPLE_ROW"
    else:
        detected_format = "WIDE_CONTACT"

    # Parse data rows
    parsed_rows: list[ParsedRow] = []
    blank_count = 0
    email_count = 0
    no_email_count = 0

    col_map = mapping.to_dict()

    for row_idx, raw_row in enumerate(data_rows_raw):
        # Skip blank rows
        if all(not _clean_val(v) for v in raw_row):
            blank_count += 1
            continue

        # Build raw data dict
        raw_dict = {}
        for i, h in enumerate(headers):
            if i < len(raw_row) and h:
                raw_dict[h] = raw_row[i] if i < len(raw_row) else None

        # Map fields
        parsed = ParsedRow(
            raw_data=raw_dict,
            source_file=source_file,
            source_sheet=sheet_name,
            row_index=row_idx + (2 if has_headers else 1),  # 1-indexed, accounting for header
        )

        for logical_field, col_header in col_map.items():
            if col_header is None:
                continue
            # Find column index
            try:
                col_idx = headers.index(col_header)
            except ValueError:
                continue
            if col_idx >= len(raw_row):
                continue

            raw_val = _clean_val(raw_row[col_idx])

            if logical_field == "email":
                parsed.email = _clean_email(raw_val) if raw_val else None
            elif logical_field == "email2":
                parsed.email2 = _clean_email(raw_val) if raw_val else None
            elif logical_field == "email3":
                parsed.email3 = _clean_email(raw_val) if raw_val else None
            elif logical_field == "email4":
                parsed.email4 = _clean_email(raw_val) if raw_val else None
            elif logical_field == "phone":
                parsed.phone = _clean_phone(raw_val) if raw_val else None
            elif logical_field == "phone2":
                parsed.phone2 = _clean_phone(raw_val) if raw_val else None
            elif logical_field == "phone3":
                parsed.phone3 = _clean_phone(raw_val) if raw_val else None
            elif logical_field == "phone4":
                parsed.phone4 = _clean_phone(raw_val) if raw_val else None
            else:
                setattr(parsed, logical_field, raw_val)

        # Store unmapped columns in metadata
        unmapped_data = {}
        for col_name in mapping.unmapped_columns:
            try:
                col_idx = headers.index(col_name)
                val = _clean_val(raw_row[col_idx]) if col_idx < len(raw_row) else None
                if val:
                    unmapped_data[col_name] = val
            except (ValueError, IndexError):
                continue

        if unmapped_data:
            parsed.metadata["unmapped_fields"] = unmapped_data

        # Source tracking
        parsed.metadata["import_source_sheet"] = sheet_name
        parsed.metadata["import_source_file"] = source_file
        parsed.metadata["source_format"] = detected_format

        # Review flags
        if not parsed.email:
            parsed.needs_review = True
            parsed.review_reasons.append("missing_email")
            no_email_count += 1
        else:
            email_count += 1

        if parsed.name and re.match(r"^.+\s[A-Z]\.$", parsed.name):
            parsed.needs_review = True
            parsed.review_reasons.append("partial_name")

        if not has_headers:
            parsed.needs_review = True
            if "headerless_source" not in parsed.review_reasons:
                parsed.review_reasons.append("headerless_source")

        if not parsed.state and not parsed.location:
            parsed.needs_review = True
            if "missing_location" not in parsed.review_reasons:
                parsed.review_reasons.append("missing_location")

        parsed_rows.append(parsed)

    return SheetParseResult(
        sheet_name=sheet_name,
        detected_format=detected_format,
        format_confidence=mapping.confidence,
        column_mapping=mapping,
        has_headers=has_headers,
        headers=headers,
        total_rows=len(data_rows_raw),
        data_rows=len(parsed_rows),
        blank_rows=blank_count,
        rows_with_email=email_count,
        rows_without_email=no_email_count,
        parsed_rows=parsed_rows,
    )


# ──────────────────────────────────────────────────────────────
# File Parsing (top-level entry points)
# ──────────────────────────────────────────────────────────────

def parse_xlsx_file(filepath: str, max_rows_per_sheet: int | None = None) -> FileParseResult:
    """
    Parse an xlsx file with all its sheets.
    Returns structured results without touching any database.
    """
    filename = os.path.basename(filepath)
    file_size = os.path.getsize(filepath)

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return FileParseResult(
            filename=filename, file_size_bytes=file_size,
            sheet_count=0, sheet_names=[],
            errors=[f"Cannot open file: {e}"],
        )

    result = FileParseResult(
        filename=filename, file_size_bytes=file_size,
        sheet_count=len(wb.sheetnames), sheet_names=list(wb.sheetnames),
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows: list[list[str]] = []
        row_count = 0

        for row_values in ws.iter_rows(values_only=True):
            row_count += 1
            if max_rows_per_sheet and row_count > max_rows_per_sheet:
                break
            all_rows.append([str(v).strip() if v is not None else "" for v in row_values])

        sheet_result = parse_sheet_rows(all_rows, sheet_name, filename)
        result.sheets.append(sheet_result)

    wb.close()
    return result


def parse_csv_file(filepath: str, max_rows: int | None = None) -> FileParseResult:
    """Parse a CSV file. Returns structured results without touching any database."""
    filename = os.path.basename(filepath)
    file_size = os.path.getsize(filepath)

    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            all_rows: list[list[str]] = []
            for i, row in enumerate(reader):
                if max_rows and i > max_rows:
                    break
                all_rows.append([str(v).strip() for v in row])
    except Exception as e:
        return FileParseResult(
            filename=filename, file_size_bytes=file_size,
            sheet_count=0, sheet_names=[],
            errors=[f"Cannot read CSV: {e}"],
        )

    sheet_result = parse_sheet_rows(all_rows, "Sheet1", filename)

    return FileParseResult(
        filename=filename, file_size_bytes=file_size,
        sheet_count=1, sheet_names=["Sheet1"],
        sheets=[sheet_result],
    )


def parse_file(filepath: str, max_rows_per_sheet: int | None = None) -> FileParseResult:
    """Auto-detect file type and parse accordingly."""
    lower = filepath.lower()
    if lower.endswith(".csv"):
        return parse_csv_file(filepath, max_rows=max_rows_per_sheet)
    elif lower.endswith((".xlsx", ".xls")):
        return parse_xlsx_file(filepath, max_rows_per_sheet=max_rows_per_sheet)
    else:
        return FileParseResult(
            filename=os.path.basename(filepath),
            file_size_bytes=os.path.getsize(filepath) if os.path.exists(filepath) else 0,
            sheet_count=0, sheet_names=[],
            errors=[f"Unsupported file type: {filepath}"],
        )
