import csv
import json
import os
import traceback
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.models import Company, RawUpload, Recruiter, UploadJob
from app.utils.normalizer import extract_domain, normalize_text
from app.utils.state_mapper import normalize_state
from app.services.job_tracker import mark_progress, utc_now

def clean_val(v):
    if pd.isna(v): return None
    s = str(v).strip()
    return s if s and s.lower() not in ('none', 'n/a', 'nan', 'null', '') else None

def clean_email(email): 
    return str(email).lower().strip() if email else None

def clean_phone(phone): 
    if not phone: return None
    return str(phone).replace("-","").replace(" ","").replace("(","").replace(")","").strip()

def clean_name(name):
    if not name:
        return None
    value = str(name).strip()
    return value if value else None

def build_company_name(company_name: str | None) -> str | None:
    if not company_name:
        return None
    value = str(company_name).strip()
    return value if value else None


def count_file_rows(filepath: str) -> int:
    if filepath.lower().endswith(".csv"):
        try:
            with open(filepath, "r", encoding="utf-8", errors="ignore") as handle:
                row_count = sum(1 for _ in handle) - 1
                return max(row_count, 0)
        except Exception:
            return 0
    try:
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        worksheet = workbook.active
        row_count = max((worksheet.max_row or 1) - 1, 0)
        workbook.close()
        return row_count
    except Exception:
        return 0

def iter_file_batches(filepath: str, batch_size: int = 2000):
    if filepath.lower().endswith(".csv"):
        for chunk in pd.read_csv(filepath, dtype=str, keep_default_na=False, chunksize=batch_size):
            yield chunk.fillna("").to_dict(orient="records")
        return

    workbook = load_workbook(filepath, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(header).strip() if header is not None else "" for header in next(rows, [])]
    batch: list[dict] = []
    for values in rows:
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[index] if index < len(values) else None
        batch.append(row)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch
    workbook.close()

def get_or_create_company(
    db: Session,
    source_job_id: str | None,
    company_name: str | None,
    location: str | None,
    state: str | None,
    email: str | None,
    company_cache: dict[str, int],
):
    company_name = build_company_name(company_name)
    if not company_name:
        return None

    normalized_name = normalize_text(company_name)
    if not normalized_name:
        return None

    if normalized_name in company_cache:
        return company_cache[normalized_name]

    company = db.query(Company).filter(Company.normalized_company_name == normalized_name).first()
    if company:
        company_cache[normalized_name] = company.company_id
        return company.company_id

    domain = extract_domain(email) if email else ""
    if domain and domain not in {"gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "aol.com"}:
        company = db.query(Company).filter(
            (Company.website.ilike(f"%{domain}%")) |
            (Company.email_pattern.ilike(f"%{domain}%"))
        ).first()
        if company:
            company_cache[normalized_name] = company.company_id
            return company.company_id

    company = Company(
        company_name=company_name,
        normalized_company_name=normalized_name,
        location=location,
        state=state,
        data_source="etl",
        trust_score=80,
        is_active=True,
        source_job_id=source_job_id,
    )
    db.add(company)
    db.flush()
    company_cache[normalized_name] = company.company_id
    return company.company_id

def calculate_completeness(name: str | None, email: str | None, phone: str | None, company_id: int | None, location: str | None, state: str | None) -> int:
    score = 0
    if name: score += 25
    if email: score += 25
    if phone: score += 15
    if company_id: score += 15
    if location: score += 10
    if state: score += 10
    return min(score, 100)

def process_smart_import(job_id: str, filepath: str, column_map: dict):
    db: Session = SessionLocal()
    job = db.query(UploadJob).filter(UploadJob.job_id == job_id).first()
    
    if not job:
        db.close()
        return

    total_rows = count_file_rows(filepath)
    mark_progress(
        job,
        status="parsing",
        current_step="Parsing rows",
        progress_percent=8,
        total_rows=total_rows,
        processed_rows=0,
        valid_rows=0,
        warning_rows=0,
        duplicate_rows=0,
        possible_duplicate_rows=0,
        enriched_rows=0,
        failed_rows=0,
        inserted_rows=0,
        skipped_rows=0,
        error_count=0,
    )
    db.commit()

    try:
        source_job_id = job_id
        mark_progress(job, status="mapping", current_step="Mapping columns", progress_percent=15)
        db.commit()

        email_col = column_map.get('email')
        name_col = column_map.get('name')
        phone_col = column_map.get('phone')
        company_col = column_map.get('company')
        location_col = column_map.get('location')
        state_col = column_map.get('state')

        inserted = 0
        skipped = 0
        errors = 0
        processed = 0
        valid_rows = 0
        warning_rows = 0
        duplicate_rows = 0
        possible_duplicate_rows = 0
        enriched_rows = 0
        failed_rows = 0
        error_log = []
        existing_emails = {
            email for (email,) in db.query(Recruiter.email).yield_per(5000) if email
        }
        company_cache: dict[str, int] = {
            normalized: company_id
            for company_id, normalized in db.query(Company.company_id, Company.normalized_company_name).all()
            if normalized
        }

        batch_size = 2000
        if total_rows <= 0:
            total_rows = 0
        job.total_rows = total_rows
        db.commit()

        for batch in iter_file_batches(filepath, batch_size=batch_size):
            mark_progress(
                job,
                status="importing",
                current_step=f"Importing batch {max(processed // batch_size + 1, 1)}",
                progress_percent=20 if not total_rows else min(95, 20 + int((processed / max(total_rows, 1)) * 75)),
            )
            db.commit()
            pending_rows = []

            for row in batch:
                processed += 1
                try:
                    email = clean_email(clean_val(row.get(email_col, '')))
                    if not email or '@' not in email:
                        skipped += 1
                        failed_rows += 1
                        continue

                    if email in existing_emails:
                        skipped += 1
                        duplicate_rows += 1
                        continue

                    raw_name = clean_name(clean_val(row.get(name_col, ''))) if name_col else None
                    fallback_name = email.split('@')[0].replace('.', ' ').replace('_', ' ').title()
                    recruiter_name = raw_name or fallback_name
                    if not recruiter_name:
                        skipped += 1
                        failed_rows += 1
                        continue

                    company_name = clean_val(row.get(company_col, '')) if company_col else None
                    loc_val = clean_val(row.get(location_col, '')) if location_col else None
                    state_val = clean_val(row.get(state_col, '')) if state_col else None
                    normalized_state = normalize_state(state_val or loc_val) if (state_val or loc_val) else None
                    phone_val = clean_phone(clean_val(row.get(phone_col, ''))) if phone_col else None
                    try:
                        company_id = get_or_create_company(
                            db,
                            source_job_id,
                            company_name,
                            loc_val,
                            normalized_state,
                            email,
                            company_cache,
                        )
                    except Exception as company_error:
                        company_id = None
                        warning_rows += 1
                        error_log.append({"row": processed, "reason": f"Company lookup failed: {company_error}"})

                    pending_rows.append({
                        "raw_row": {
                            "job_id": job_id,
                            "raw_data": json.dumps(row, default=str),
                            "source_filename": job.filename,
                        },
                        "recruiter_row": {
                            "recruiter_name": recruiter_name,
                            "normalized_recruiter_name": normalize_text(recruiter_name),
                            "email": email,
                            "phone": phone_val,
                            "company_id": company_id,
                            "location": loc_val,
                            "state": normalized_state,
                            "normalized_city": normalize_text(loc_val) if loc_val else None,
                            "location_confidence": "high" if normalized_state or loc_val else "low",
                            "completeness_score": calculate_completeness(
                                recruiter_name,
                                email,
                                phone_val,
                                company_id,
                                loc_val,
                                normalized_state,
                            ),
                            "needs_review": not bool(company_id and normalized_state),
                            "data_source": "etl",
                            "trust_score": 80 if company_id else 65,
                            "is_active": True,
                            "source_job_id": job_id,
                        },
                    })
                    existing_emails.add(email)
                    inserted += 1
                    valid_rows += 1
                    if not normalized_state or not company_id:
                        warning_rows += 1
                except Exception as e:
                    errors += 1
                    failed_rows += 1
                    error_log.append({"row": processed, "reason": str(e)})

            try:
                if pending_rows:
                    db.bulk_insert_mappings(RawUpload, [item["raw_row"] for item in pending_rows])
                    db.bulk_insert_mappings(Recruiter, [item["recruiter_row"] for item in pending_rows])
                db.commit()
            except Exception as batch_error:
                db.rollback()
                row_failures = 0
                for item in pending_rows:
                    try:
                        raw_row = item["raw_row"]
                        recruiter_row = item["recruiter_row"]
                        raw_record = RawUpload(**raw_row)
                        recruiter_record = Recruiter(**recruiter_row)
                        db.add(raw_record)
                        db.add(recruiter_record)
                        db.commit()
                    except Exception as row_error:
                        db.rollback()
                        row_failures += 1
                        errors += 1
                        failed_rows += 1
                        error_log.append({"row": processed, "reason": f"{batch_error}; row retry failed: {row_error}"})
                if row_failures == 0:
                    error_log.append({"row": processed, "reason": str(batch_error)})

            mark_progress(
                job,
                status="importing",
                current_step=f"Importing rows {processed}/{max(total_rows, processed)}",
                progress_percent=20 if not total_rows else min(96, 20 + int((processed / max(total_rows, processed)) * 75)),
                processed_rows=processed,
                inserted_rows=inserted,
                skipped_rows=skipped,
                error_count=errors,
                valid_rows=valid_rows,
                warning_rows=warning_rows,
                duplicate_rows=duplicate_rows,
                possible_duplicate_rows=possible_duplicate_rows,
                enriched_rows=enriched_rows,
                failed_rows=failed_rows,
            )
            db.commit()

        mark_progress(
            job,
            status="completed",
            current_step="Import completed",
            progress_percent=100,
            processed_rows=processed,
            inserted_rows=inserted,
            skipped_rows=skipped,
            error_count=errors,
            valid_rows=valid_rows,
            warning_rows=warning_rows,
            duplicate_rows=duplicate_rows,
            possible_duplicate_rows=possible_duplicate_rows,
            enriched_rows=enriched_rows,
            failed_rows=failed_rows,
        )
        job.errors = json.dumps(error_log)
        job.completed_at = utc_now()
        db.commit()

    except Exception as e:
        db.rollback()
        mark_progress(job, status="failed", current_step="Import failed", progress_percent=100)
        job.errors = json.dumps([{"row": 0, "reason": f"Fatal pipeline error: {traceback.format_exc()}"}])
        job.error_message = str(e)
        job.completed_at = utc_now()
        db.commit()
    finally:
        db.close()

