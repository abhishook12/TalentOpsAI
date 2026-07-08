import os
import time
import json
import hashlib
import pyautogui
from PIL import Image
from google import genai
from google.genai import types
import traceback
import openpyxl
from openpyxl.styles import PatternFill

import win32gui
import win32ui
import win32con
from ctypes import windll

def capture_window_bg(hwnd, region):
    left, top, right, bot = win32gui.GetWindowRect(hwnd)
    w = right - left
    h = bot - top
    if w <= 0 or h <= 0:
        return None
        
    hwndDC = win32gui.GetWindowDC(hwnd)
    mfcDC  = win32ui.CreateDCFromHandle(hwndDC)
    saveDC = mfcDC.CreateCompatibleDC()

    saveBitMap = win32ui.CreateBitmap()
    saveBitMap.CreateCompatibleBitmap(mfcDC, w, h)
    saveDC.SelectObject(saveBitMap)

    # PW_RENDERFULLCONTENT = 2
    result = windll.user32.PrintWindow(hwnd, saveDC.GetSafeHdc(), 2)
    
    im = None
    if result == 1:
        bmpinfo = saveBitMap.GetInfo()
        bmpstr = saveBitMap.GetBitmapBits(True)
        try:
            im = Image.frombuffer(
                'RGB',
                (bmpinfo['bmWidth'], bmpinfo['bmHeight']),
                bmpstr, 'raw', 'BGRX', 0, 1)
            
            # Crop to region
            crop_x = region[0] - left
            crop_y = region[1] - top
            im = im.crop((crop_x, crop_y, crop_x + region[2], crop_y + region[3]))
        except:
            pass
            
    win32gui.DeleteObject(saveBitMap.GetHandle())
    saveDC.DeleteDC()
    mfcDC.DeleteDC()
    win32gui.ReleaseDC(hwnd, hwndDC)

    return im

def background_scroll(hwnd, amount, x, y):
    wparam = (amount & 0xFFFF) << 16
    lparam = (y << 16) | (x & 0xFFFF)
    win32gui.PostMessage(hwnd, win32con.WM_MOUSEWHEEL, wparam, lparam)

def merge_contact(all_contacts, new_contact):
    new_items = new_contact.get("contacts", [])
    if not new_items:
        return False
        
    new_identifiers = {str(i.get("value", "")).lower().strip() for i in new_items if i.get("value")}
    new_name = str(new_contact.get("name", "")).lower().strip()
    
    for existing in all_contacts:
        existing_items = existing.get("contacts", [])
        existing_identifiers = {str(i.get("value", "")).lower().strip() for i in existing_items if i.get("value")}
        existing_name = str(existing.get("name", "")).lower().strip()
        
        is_match = False
        if new_name and existing_name and new_name == existing_name:
            is_match = True
        if new_identifiers.intersection(existing_identifiers):
            is_match = True
            
        if is_match:
            for key in ["company", "title", "location", "linkedin", "notes"]:
                if not existing.get(key) and new_contact.get(key):
                    existing[key] = new_contact.get(key)
            for item in new_items:
                val = str(item.get("value", "")).lower().strip()
                if val and val not in existing_identifiers:
                    existing_items.append(item)
                    existing_identifiers.add(val)
            return True
            
    all_contacts.append(new_contact)
    return False

def build_excel(contacts, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Contacts"

    from openpyxl.styles import Font, Alignment
    
    headers = ["Name", "Title", "Company", "Location", "Email(s)", "Phone(s)", "LinkedIn", "Notes"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    current_row = 2
    for contact in contacts:
        name = contact.get("name", "")
        title = contact.get("title", "")
        company = contact.get("company", "")
        location = contact.get("location", "")
        linkedin = contact.get("linkedin", "")
        notes = contact.get("notes", "")
        
        emails = []
        phones = []
        for item in contact.get("contacts", []):
            ctype = str(item.get("type", "")).lower()
            val = str(item.get("value", ""))
            if "email" in ctype:
                emails.append(val)
            elif "phone" in ctype:
                phones.append(val)
        
        ws.cell(row=current_row, column=1, value=name)
        ws.cell(row=current_row, column=2, value=title)
        ws.cell(row=current_row, column=3, value=company)
        ws.cell(row=current_row, column=4, value=location)
        ws.cell(row=current_row, column=5, value="; ".join(emails))
        ws.cell(row=current_row, column=6, value="; ".join(phones))
        ws.cell(row=current_row, column=7, value=linkedin)
        ws.cell(row=current_row, column=8, value=notes)
        
        current_row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(50, max_length + 2)

    wb.save(out_path)

def _image_hash(img):
    """Compute a perceptual hash of a PIL image for stall detection."""
    small = img.resize((16, 16)).convert('L')
    pixels = list(small.getdata())
    return hashlib.md5(bytes(pixels)).hexdigest()


def _save_checkpoint(output_dir, session_id, data):
    """Save extraction checkpoint so we can resume after crashes."""
    cp_path = os.path.join(output_dir, f"checkpoint_{session_id}.json")
    with open(cp_path, "w") as f:
        json.dump(data, f, indent=2)


def _load_checkpoint(output_dir):
    """Load the most recent checkpoint if it exists."""
    import glob
    cp_pattern = os.path.join(output_dir, "checkpoint_*.json")
    files = glob.glob(cp_pattern)
    if not files:
        return None
    latest_cp = max(files, key=os.path.getctime)
    try:
        with open(latest_cp, "r") as f:
            return json.load(f)
    except:
        return None


def run_extraction_loop(ui, config):
    try:
        import datetime
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        debug_log_path = os.path.join(output_dir, "debug_log.txt")
        
        with open(debug_log_path, "a") as f:
            f.write(f"\n[{datetime.datetime.now()}] run_extraction_loop started\n")
            
        session_id = config.get("session_id", "default")
        
        now = datetime.datetime.now()
        friendly_time = now.strftime("%I-%M%p_%d%b").upper()
        
        region = ui.selected_region
        if not region:
            with open(debug_log_path, "a") as f: f.write("Error: no region\n")
            ui.updater.update_status.emit("Error: No region selected.")
            ui.is_running = False
            return
        with open(debug_log_path, "a") as f: f.write(f"Region: {region}\n")
        
        # --- Load API Keys ---
        api_keys = []
        keys_file = "api_keys.txt"
        if os.path.exists(keys_file):
            with open(keys_file, "r", encoding="utf-8") as f:
                api_keys = [line.encode('ascii', 'ignore').decode('ascii').strip() for line in f if line.strip()]
                api_keys = [k for k in api_keys if k]
        
        # Fallback to the hardcoded key if file is empty/missing
        if not api_keys:
            api_keys = ['AQ.Ab8RN6JeAR1jCZaGpC-HcXZBHrtG-TK0oZQL6aQCMMm3vuyHYQ']
            
        current_key_idx = 0
        
        def get_client(idx):
            key = api_keys[idx % len(api_keys)]
            return genai.Client(api_key=key)
            
        try:
            client = get_client(current_key_idx)
            with open(debug_log_path, "a") as f: f.write(f"Gemini client initialized with key {current_key_idx + 1}/{len(api_keys)}\n")
        except Exception as e:
            with open(debug_log_path, "a") as f: f.write(f"Gemini Init Error: {e}\n")
            ui.updater.update_status.emit(f"Gemini Init Error: {e}")
            ui.is_running = False
            return

        # --- Load checkpoint from previous run ---
        checkpoint = _load_checkpoint(output_dir)
        all_extracted_contacts = []
        if checkpoint and "all_extracted_contacts" in checkpoint:
            all_extracted_contacts = checkpoint["all_extracted_contacts"]
            with open(debug_log_path, "a") as f:
                f.write(f"Resuming from checkpoint: last_person={checkpoint.get('last_person_name')}, saved={len(all_extracted_contacts)}\n")
            ui.updater.update_status.emit(f"Resuming from: {checkpoint.get('last_person_name', '?')}")

        scroll_num = 0
        total_saved = 0
        error_msg = ""
        direction = config.get("direction", "Down")
        

        
        # --- Stall detection: track last 3 image hashes ---
        last_image_hash = None
        stall_count = 0
        MAX_STALLS = 3  # Stop after 3 identical frames
        
        scroll_amount = config.get("scroll_amount", int(region[3] * 0.8))
        if direction == "Down":
            scroll_amount = -abs(scroll_amount)
        else:
            scroll_amount = abs(scroll_amount)

        ui.updater.update_status.emit("Starting Vision AI Excel Pipeline...")

        while ui.is_running:
            if scroll_num >= config.get("max_scrolls", 500):
                ui.updater.update_status.emit(f"Reached Maximum Scrolls ({scroll_num})!")
                with open(debug_log_path, "a") as f: f.write(f"Reached max scrolls: {scroll_num}\n")
                ui.is_running = False
                break
                
            scroll_num += 1
            with open(debug_log_path, "a") as f: f.write(f"Starting scroll {scroll_num}\n")
            if not ui.is_running:
                break

            while ui.is_paused:
                time.sleep(0.5)
                if not ui.is_running:
                    break

            ui.updater.update_status.emit(f"Capturing screen {scroll_num}...")
            ui.updater.update_stats.emit({"scroll": scroll_num})
            ui.updater.update_action.emit("Analyzing with Gemini...")

            try:
                # Find the target window under the center of the region
                center_x = region[0] + region[2] // 2
                center_y = region[1] + region[3] // 2
                target_hwnd = win32gui.WindowFromPoint((center_x, center_y))
                
                # Get the top-level parent window to capture
                parent = win32gui.GetParent(target_hwnd)
                top_hwnd = target_hwnd
                while parent:
                    top_hwnd = parent
                    parent = win32gui.GetParent(top_hwnd)

                with open(debug_log_path, "a") as f: f.write(f"Capturing background window {top_hwnd} for region={region}...\n")
                img = capture_window_bg(top_hwnd, region)
                if not img:
                    with open(debug_log_path, "a") as f: f.write(f"Background capture failed, falling back to pyautogui!\n")
                    img = pyautogui.screenshot(region=region)
                
                debug_path = os.path.join(output_dir, f"vision_debug_{scroll_num}.png")
                img.save(debug_path)
                with open(debug_log_path, "a") as f: f.write(f"Saved debug img to {debug_path}\n")

                # --- Stall detection: compare image hash ---
                current_hash = _image_hash(img)
                if current_hash == last_image_hash:
                    stall_count += 1
                    with open(debug_log_path, "a") as f:
                        f.write(f"STALL DETECTED ({stall_count}/{MAX_STALLS}): identical frame to previous scroll\n")
                    if stall_count >= MAX_STALLS:
                        with open(debug_log_path, "a") as f:
                            f.write(f"STOPPING: Scroll stuck after {MAX_STALLS} identical frames. Saving progress.\n")
                        ui.updater.update_status.emit(f"Scroll stuck! Saved {total_saved} contacts. Checkpointed.")
                        ui.is_running = False
                        break
                else:
                    stall_count = 0
                last_image_hash = current_hash

                prompt = """You are an expert data extraction AI analyzing a continuous vertical scroll of a chat interface.
Return a valid JSON list of objects. Each object should have the following keys:
"name", "title", "company", "location", "linkedin", "notes", "contacts", "is_continuation_from_above"
The "contacts" field should be a list of objects with "type" and "value".
"type" should be one of: "Work Email", "Personal Email", "Phone".
If a person has multiple emails or phones, include all of them in the "contacts" list.
Extract any available title, location, linkedin URL, or notes (e.g. if an email association is uncertain). If not present, leave as an empty string.

CRITICALLY IMPORTANT: The names appearing in small text at the very top of each message block are the SENDERS/RECRUITERS (e.g., 'Shruti Tiwari', 'Gaurav Ojha', 'Chauhan', 'Sharma', 'chitransh tiwari', 'Yatin Rawat'). DO NOT extract the sender's name. 

You MUST extract the candidate information (Name, Email, Phone, Company) located INSIDE the actual message content. 
EXTREME ACCURACY REQUIRED: DO NOT SKIP ANY CANDIDATES. You must extract EVERY SINGLE candidate contact block that is visible in the image. If you see an email address or a phone number, you MUST extract it and the associated candidate name. Missing a candidate is an absolute failure.

IMPORTANT: The image is a segment of a continuous scroll. If the very FIRST contact info shown at the top of the screen is missing its Name/Company header because the header was scrolled off-screen (e.g. it just shows a phone number or email belonging to the person from the previous screen), set "is_continuation_from_above": true for that object. Otherwise, false.
Only output the raw JSON array.
"""
                with open(debug_log_path, "a") as f: f.write("Calling Gemini API...\n")
                
                retries = 3
                while retries > 0:
                    if not ui.is_running:
                        break
                    try:
                        response = client.models.generate_content(
                            model='gemini-2.5-flash',
                            contents=[prompt, img],
                            config=types.GenerateContentConfig(
                                response_mime_type='application/json',
                            )
                        )
                        break
                    except Exception as api_e:
                        err_str = str(api_e).lower()
                        if "429" in err_str or "quota" in err_str or "exhausted" in err_str or "503" in err_str or "unavailable" in err_str:
                            with open(debug_log_path, "a") as f: f.write(f"API Limit Hit on Key {current_key_idx + 1}: {api_e}.\n")
                            
                            if len(api_keys) > 1:
                                current_key_idx = (current_key_idx + 1) % len(api_keys)
                                ui.updater.update_status.emit(f"Rate Limit Hit! Swapping to Key {current_key_idx + 1}...")
                                client = get_client(current_key_idx)
                                with open(debug_log_path, "a") as f: f.write(f"Swapped to Key {current_key_idx + 1}.\n")
                                # Don't decrement retries for a simple key swap, let it keep looping through keys
                                time.sleep(1) # tiny pause before retry
                            else:
                                ui.updater.update_status.emit(f"API Rate Limit! Pausing for 30s... ({retries} retries left)")
                                time.sleep(30)
                                retries -= 1
                                
                            if retries <= 0 and len(api_keys) <= 1:
                                raise Exception(f"Failed after 3 API retries: {api_e}")
                            
                            # If we have multiple keys and looped through all of them, force a sleep
                            if len(api_keys) > 1 and retries > 0 and (current_key_idx == 0):
                                retries -= 1
                                if retries <= 0:
                                    raise Exception(f"All keys exhausted and retries failed: {api_e}")
                                ui.updater.update_status.emit("All keys rate limited! Pausing 30s...")
                                time.sleep(30)
                        else:
                            raise api_e
                            
                with open(debug_log_path, "a") as f: f.write("Gemini API succeeded.\n")
                
                ui.updater.update_action.emit("Processing data...")
                
                raw_text = response.text.strip()
                if raw_text.startswith("```json"):
                    raw_text = raw_text[7:]
                if raw_text.endswith("```"):
                    raw_text = raw_text[:-3]
                    
                try:
                    contacts = json.loads(raw_text.strip())
                    with open(debug_log_path, "a") as f: f.write(f"Parsed {len(contacts)} contacts\n")
                except Exception as e:
                    with open(debug_log_path, "a") as f: f.write(f"JSON parse error: {e}\n")
                    contacts = []

                for idx, contact in enumerate(contacts):
                    if idx == 0 and contact.get("is_continuation_from_above") and len(all_extracted_contacts) > 0:
                        with open(debug_log_path, "a") as f: f.write("Merging continuation into previous contact!\n")
                        last_contact = all_extracted_contacts[-1]
                        existing_items = last_contact.get("contacts", [])
                        existing_identifiers = {str(i.get("value", "")).lower().strip() for i in existing_items if i.get("value")}
                        
                        new_items = contact.get("contacts", [])
                        for item in new_items:
                            val = str(item.get("value", "")).lower().strip()
                            if val and val not in existing_identifiers:
                                existing_items.append(item)
                                existing_identifiers.add(val)
                        continue

                    merged = merge_contact(all_extracted_contacts, contact)
                    if not merged:
                        total_saved += 1

                ui.updater.update_stats.emit({
                    "saved": total_saved,
                })

                # --- Save checkpoint after each successful scroll ---
                last_person = all_extracted_contacts[-1].get("name", "Unknown") if all_extracted_contacts else "None"
                _save_checkpoint(output_dir, session_id, {
                    "scroll_num": scroll_num,
                    "total_saved": total_saved,
                    "last_person_name": last_person,
                    "all_extracted_contacts": all_extracted_contacts,
                    "timestamp": str(datetime.datetime.now()),
                })
                with open(debug_log_path, "a") as f:
                    f.write(f"Checkpoint saved: scroll={scroll_num}, saved={total_saved}, last_person={last_person}\n")

                ui.updater.update_action.emit("Scrolling...")
                with open(debug_log_path, "a") as f: f.write(f"Scrolling background window {target_hwnd} amount: {scroll_amount}\n")
                background_scroll(target_hwnd, scroll_amount, center_x, center_y)
                time.sleep(config.get("delay", 1.0))

            except Exception as e:
                error_msg = f"Error: {e}\n{traceback.format_exc()}"
                with open(debug_log_path, "a") as f: f.write(f"Loop Exception: {error_msg}\n")
                ui.updater.update_status.emit(f"Error: {e}")
                ui.is_running = False
                break

        ui.updater.stop_ui.emit()
        ui.updater.update_action.emit("Idle")

        if error_msg:
            with open(debug_log_path, "a") as f: f.write("Writing to crash_log.txt\n")
            with open(os.path.join(output_dir, "crash_log.txt"), "w") as f:
                f.write(error_msg)
            ui.updater.update_status.emit("Crashed! See output/crash_log.txt")
        else:
            out_excel = os.path.join(output_dir, f"vision_extracted_data_{friendly_time}.xlsx")
            with open(debug_log_path, "a") as f: f.write(f"Building excel at {out_excel}\n")
            build_excel(all_extracted_contacts, out_excel)
            ui.updater.update_status.emit(f"Done! {total_saved} unique contacts saved to Excel.")
            with open(debug_log_path, "a") as f: f.write("Done building excel\n")
            
    except Exception as outer_e:
        debug_log_path = os.path.join("output", "debug_log.txt")
        with open(debug_log_path, "a") as f: 
            f.write(f"FATAL OUTER ERROR: {outer_e}\n{traceback.format_exc()}\n")
        try:
            ui.updater.update_status.emit(f"Fatal Error: {outer_e}")
            ui.is_running = False
            ui.updater.stop_ui.emit()
        except:
            pass
